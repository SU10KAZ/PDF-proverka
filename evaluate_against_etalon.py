"""
evaluate_against_etalon.py
---------------------------
Сравнение findings из 03_findings.json с эталонным xlsx (ручная проверка пользователя).

Эталон: отчет/etalon.xlsx
- Колонка B: текст замечания (всё что когда-либо находила нейросеть)
- Колонка D: вердикт пользователя ("Оставляем" / "Убираем" / "не проверял")

Использование:
    python evaluate_against_etalon.py projects/EOM/133_23-ГК-ГРЩ
    python evaluate_against_etalon.py projects/EOM/133_23-ГК-ГРЩ --etalon отчет/etalon.xlsx
    python evaluate_against_etalon.py projects/EOM/133_23-ГК-ГРЩ --threshold 50
    python evaluate_against_etalon.py projects/EOM/133_23-ГК-ГРЩ --json out.json

Что считается:
    - TP (true positive): finding мэтчится с эталонным "Оставляем"
    - FP (false positive): finding мэтчится с эталонным "Убираем"
    - Unknown: finding мэтчится с эталонным "не проверял" (нельзя судить)
    - Novel: finding не нашёл аналога в эталоне (новая находка)
    - Missed: эталонный "Оставляем" без соответствующего finding (пропуск)

Метрики:
    - Recall на definite KEEPs = TP / (TP + Missed)
    - Precision (definite) = TP / (TP + FP)
    - Precision (с потенциалом) = (TP + Unknown) / (TP + FP + Unknown)
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl не установлен. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from rapidfuzz import fuzz
except ImportError:
    print("ERROR: rapidfuzz не установлен. pip install rapidfuzz", file=sys.stderr)
    sys.exit(1)


# ─── Парсинг эталона ───────────────────────────────────────────────────────

VERDICT_KEEP = "keep"
VERDICT_REMOVE = "remove"
VERDICT_UNKNOWN = "unknown"
VERDICT_PARTIAL = "partial"


def parse_verdict(d_text: str) -> str:
    """Распарсить вердикт из колонки D эталона."""
    if not d_text:
        return VERDICT_UNKNOWN
    d = d_text.strip().lower()
    if "оставляем" in d:
        return VERDICT_KEEP
    if "убираем" in d or "удаляем" in d:
        return VERDICT_REMOVE
    if "не проверял" in d:
        return VERDICT_UNKNOWN
    # частично валидное (например "проблема есть но решить необосновано")
    return VERDICT_PARTIAL


def load_etalon(etalon_path: Path) -> list[dict]:
    """Загрузить эталон в список словарей."""
    wb = load_workbook(etalon_path, data_only=True)
    ws = wb.active
    items = []
    for row in range(2, ws.max_row + 1):
        n = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        d = ws.cell(row, 4).value
        if not b:
            continue
        items.append({
            "n": n,
            "text": str(b).strip(),
            "verdict_raw": str(d or "").strip(),
            "verdict": parse_verdict(d or ""),
        })
    return items


# ─── Парсинг findings ──────────────────────────────────────────────────────

def load_findings(findings_path: Path) -> list[dict]:
    """Загрузить findings из 03_findings.json."""
    with open(findings_path, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("findings", [])


def finding_text(f: dict) -> str:
    """Собрать текст finding для матчинга (problem + description)."""
    parts = []
    if f.get("problem"):
        parts.append(f["problem"])
    if f.get("description"):
        parts.append(f["description"][:300])
    if not parts and f.get("finding"):
        parts.append(f["finding"])
    return " ".join(parts)


# ─── Нормализация и матчинг ────────────────────────────────────────────────

def normalize(text: str) -> str:
    """Привести текст к виду для сравнения."""
    text = text.lower()
    # Унификация пунктуации
    text = re.sub(r"[«»\"'`]", "", text)
    text = re.sub(r"[—–-]", " ", text)
    # Разделители — в пробелы
    text = re.sub(r"[/\\,;:()\[\]]", " ", text)
    # Множественные пробелы
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def all_scores(finding: dict, etalon: list[dict]) -> list[tuple[int, dict]]:
    """Вернуть все scores этого finding со всеми эталонными записями.

    Returns:
        [(score, etalon_item), ...] отсортированный по убыванию score
    """
    f_text = normalize(finding_text(finding))
    if not f_text:
        return []
    scored = []
    for item in etalon:
        e_text = normalize(item["text"])
        # token_set_ratio устойчив к разному порядку слов и доп. тексту
        score = int(fuzz.token_set_ratio(f_text, e_text))
        scored.append((score, item))
    scored.sort(key=lambda x: x[0], reverse=True)
    return scored


def greedy_assign(
    findings: list[dict], etalon: list[dict], threshold: int
) -> tuple[list[tuple[dict, dict, int]], list[dict]]:
    """Greedy assignment: каждое finding ↔ один уникальный эталон.

    Сначала все пары (finding, etalon, score) сортируются по убыванию score.
    Затем жадно: берём лучшую пару, если оба свободны — фиксируем.

    Returns:
        (matched, novel) где matched = [(finding, etalon_item, score), ...],
        novel = [finding, ...] без матча
    """
    pairs = []  # (score, finding_idx, etalon_idx)
    for fi, f in enumerate(findings):
        scores = all_scores(f, etalon)
        for score, item in scores:
            if score < threshold:
                continue
            ei = etalon.index(item)
            pairs.append((score, fi, ei))

    # Сортируем по убыванию score
    pairs.sort(key=lambda x: x[0], reverse=True)

    used_findings: set[int] = set()
    used_etalon: set[int] = set()
    matched_indices = []  # (fi, ei, score)
    for score, fi, ei in pairs:
        if fi in used_findings or ei in used_etalon:
            continue
        used_findings.add(fi)
        used_etalon.add(ei)
        matched_indices.append((fi, ei, score))

    matched = [(findings[fi], etalon[ei], s) for fi, ei, s in matched_indices]
    novel = [f for i, f in enumerate(findings) if i not in used_findings]
    return matched, novel


# ─── Оценка ────────────────────────────────────────────────────────────────

def evaluate(
    findings: list[dict], etalon: list[dict], threshold: int
) -> dict:
    """Сопоставить findings с эталоном, посчитать метрики."""
    # Greedy assignment — каждое finding ↔ один эталон
    matched, novel = greedy_assign(findings, etalon, threshold)

    # Распределяем matched по вердиктам
    tp = []  # True Positives (matched + keep)
    fp = []  # False Positives (matched + remove)
    unknown_matches = []  # matched + "не проверял"
    partial = []  # matched + частично валидное

    for f, e, s in matched:
        if e["verdict"] == VERDICT_KEEP:
            tp.append((f, e, s))
        elif e["verdict"] == VERDICT_REMOVE:
            fp.append((f, e, s))
        elif e["verdict"] == VERDICT_UNKNOWN:
            unknown_matches.append((f, e, s))
        else:
            partial.append((f, e, s))

    # Missed KEEPs — эталонные "Оставляем" без матча
    keeps_in_etalon = [e for e in etalon if e["verdict"] == VERDICT_KEEP]
    matched_keep_ids = {e["n"] for f, e, s in tp}
    missed_keeps = [e for e in keeps_in_etalon if e["n"] not in matched_keep_ids]

    # Метрики
    n_tp = len(tp)
    n_fp = len(fp)
    n_unknown = len(unknown_matches)
    n_partial = len(partial)
    n_novel = len(novel)
    n_missed = len(missed_keeps)
    n_keeps_total = len(keeps_in_etalon)

    recall_definite = n_tp / n_keeps_total if n_keeps_total else 0
    precision_definite = n_tp / (n_tp + n_fp) if (n_tp + n_fp) else 0
    precision_potential = (
        (n_tp + n_unknown) / (n_tp + n_fp + n_unknown)
        if (n_tp + n_fp + n_unknown)
        else 0
    )

    return {
        "totals": {
            "findings": len(findings),
            "etalon_entries": len(etalon),
            "etalon_keeps": n_keeps_total,
            "tp": n_tp,
            "fp": n_fp,
            "partial": n_partial,
            "unknown": n_unknown,
            "novel": n_novel,
            "missed_keeps": n_missed,
        },
        "metrics": {
            "recall_definite": recall_definite,
            "precision_definite": precision_definite,
            "precision_potential": precision_potential,
        },
        "tp": tp,
        "fp": fp,
        "partial": partial,
        "unknown": unknown_matches,
        "novel": novel,
        "missed_keeps": missed_keeps,
    }


# ─── Вывод ────────────────────────────────────────────────────────────────

def color_supported() -> bool:
    return sys.stdout.isatty()


class C:
    """ANSI color codes (no-op if terminal не поддерживает)."""

    if color_supported():
        GREEN = "\033[92m"
        RED = "\033[91m"
        YELLOW = "\033[93m"
        BLUE = "\033[94m"
        DIM = "\033[2m"
        BOLD = "\033[1m"
        RESET = "\033[0m"
    else:
        GREEN = RED = YELLOW = BLUE = DIM = BOLD = RESET = ""


def fid(f: dict) -> str:
    return f.get("id", "?")


def short(text: str, n: int = 100) -> str:
    text = text.replace("\n", " ").strip()
    if len(text) > n:
        return text[: n - 1] + "…"
    return text


def print_report(result: dict, project_name: str):
    t = result["totals"]
    m = result["metrics"]

    print(f"\n{C.BOLD}=== Оценка качества аудита ==={C.RESET}")
    print(f"Проект: {project_name}")
    print(
        f"Findings: {t['findings']} | Эталон: {t['etalon_entries']} (из них KEEP: {t['etalon_keeps']})"
    )
    print()

    # Метрики
    print(f"{C.BOLD}Метрики:{C.RESET}")
    recall_pct = m["recall_definite"] * 100
    prec_d_pct = m["precision_definite"] * 100
    prec_p_pct = m["precision_potential"] * 100
    recall_color = (
        C.GREEN if recall_pct >= 80 else C.YELLOW if recall_pct >= 50 else C.RED
    )
    prec_color = (
        C.GREEN if prec_d_pct >= 60 else C.YELLOW if prec_d_pct >= 30 else C.RED
    )
    print(
        f"  Recall (definite KEEPs):   {recall_color}{t['tp']}/{t['etalon_keeps']} = {recall_pct:.0f}%{C.RESET}"
    )
    print(
        f"  Precision (definite):      {prec_color}{prec_d_pct:.0f}%{C.RESET}  "
        f"(TP/(TP+FP) = {t['tp']}/{t['tp']+t['fp']})"
    )
    print(
        f"  Precision (с потенциалом): {prec_p_pct:.0f}%  "
        f"((TP+Unknown)/(TP+FP+Unknown))"
    )
    print()

    # Распределение
    print(f"{C.BOLD}Распределение findings:{C.RESET}")
    print(f"  {C.GREEN}✅ TP        {t['tp']:3d}{C.RESET}  (matched + 'Оставляем')")
    print(f"  {C.RED}❌ FP        {t['fp']:3d}{C.RESET}  (matched + 'Убираем')")
    print(f"  {C.YELLOW}⚠ Partial   {t['partial']:3d}{C.RESET}  (matched + полу-валидное)")
    print(f"  {C.BLUE}? Unknown   {t['unknown']:3d}{C.RESET}  (matched + 'не проверял')")
    print(f"  {C.DIM}+ Novel     {t['novel']:3d}{C.RESET}  (нет в эталоне)")
    print(f"  {C.RED}- Missed    {t['missed_keeps']:3d}{C.RESET}  (KEEP без finding)")
    print()

    # TP details
    if result["tp"]:
        print(f"{C.BOLD}{C.GREEN}✅ True Positives ({len(result['tp'])}):{C.RESET}")
        for f, e, s in result["tp"]:
            print(f"  {C.GREEN}✅{C.RESET} {fid(f)} → эталон #{e['n']} (score {s})")
            print(f"     Finding: {short(finding_text(f), 110)}")
            print(f"     Эталон:  {short(e['text'], 110)}")
        print()

    # Partial
    if result["partial"]:
        print(f"{C.BOLD}{C.YELLOW}⚠ Partial ({len(result['partial'])}):{C.RESET}")
        for f, e, s in result["partial"]:
            print(f"  {C.YELLOW}⚠{C.RESET} {fid(f)} → эталон #{e['n']} (score {s})")
            print(f"     Finding: {short(finding_text(f), 110)}")
            print(f"     Эталон:  {short(e['text'], 110)}")
            print(f"     Вердикт: {short(e['verdict_raw'], 110)}")
        print()

    # Missed KEEPs — самое важное
    if result["missed_keeps"]:
        print(
            f"{C.BOLD}{C.RED}❌ Пропущенные KEEPs ({len(result['missed_keeps'])}):{C.RESET}"
        )
        for e in result["missed_keeps"]:
            print(f"  {C.RED}❌{C.RESET} #{e['n']}: {short(e['text'], 130)}")
            # Показать ближайший Novel finding (может быть это «он, но иначе сформулирован»)
            best_novel = None
            best_score = 0
            for f in result["novel"]:
                f_norm = normalize(finding_text(f))
                e_norm = normalize(e["text"])
                if f_norm and e_norm:
                    sc = int(fuzz.token_set_ratio(f_norm, e_norm))
                    if sc > best_score:
                        best_score = sc
                        best_novel = f
            if best_novel and best_score >= 30:
                print(
                    f"     {C.DIM}↳ ближайший Novel: {fid(best_novel)} score {best_score} — "
                    f"{short(finding_text(best_novel), 100)}{C.RESET}"
                )
        print()

    # FP
    if result["fp"]:
        print(f"{C.BOLD}{C.RED}❌ False Positives ({len(result['fp'])}):{C.RESET}")
        for f, e, s in result["fp"]:
            print(f"  {C.RED}❌{C.RESET} {fid(f)} → эталон #{e['n']} (score {s})")
            print(f"     Finding: {short(finding_text(f), 110)}")
            print(f"     Причина исключения: {short(e['verdict_raw'], 110)}")
        print()

    # Unknown
    if result["unknown"]:
        print(f"{C.BOLD}{C.BLUE}? Unknown matches ({len(result['unknown'])}):{C.RESET}")
        for f, e, s in result["unknown"]:
            print(f"  {C.BLUE}?{C.RESET} {fid(f)} → эталон #{e['n']} (score {s})")
            print(f"     Finding: {short(finding_text(f), 110)}")
        print()

    # Novel
    if result["novel"]:
        print(f"{C.BOLD}+ Novel findings ({len(result['novel'])}):{C.RESET}")
        for f in result["novel"]:
            print(f"  + {fid(f)}: {short(finding_text(f), 130)}")
        print()


# ─── Main ──────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Сравнить findings проекта с эталонным xlsx"
    )
    parser.add_argument(
        "project",
        help="Путь к проекту (например projects/EOM/133_23-ГК-ГРЩ)",
    )
    parser.add_argument(
        "--etalon",
        default="отчет/etalon.xlsx",
        help="Путь к эталону (default: отчет/etalon.xlsx)",
    )
    parser.add_argument(
        "--threshold",
        type=int,
        default=50,
        help="Минимальный score (0-100) для совпадения, default 50",
    )
    parser.add_argument(
        "--show-near-misses",
        action="store_true",
        help="Для Novel findings показать ближайшие эталонные записи (даже ниже threshold)",
    )
    parser.add_argument(
        "--json",
        help="Сохранить результат в JSON-файл",
    )
    args = parser.parse_args()

    project_path = Path(args.project)
    findings_path = project_path / "_output" / "03_findings.json"
    etalon_path = Path(args.etalon)

    if not findings_path.exists():
        print(f"ERROR: не найден {findings_path}", file=sys.stderr)
        sys.exit(2)
    if not etalon_path.exists():
        print(f"ERROR: не найден {etalon_path}", file=sys.stderr)
        sys.exit(2)

    findings = load_findings(findings_path)
    etalon = load_etalon(etalon_path)

    if not findings:
        print(f"WARN: в {findings_path} нет findings", file=sys.stderr)
    if not etalon:
        print(f"WARN: эталон {etalon_path} пуст", file=sys.stderr)
        sys.exit(2)

    result = evaluate(findings, etalon, args.threshold)
    print_report(result, project_path.name)

    # Опционально — для Novel показать ближайший эталон ниже threshold
    if args.show_near_misses and result["novel"]:
        print(f"{C.BOLD}Near-misses (ближайший эталон даже ниже threshold {args.threshold}):{C.RESET}")
        for f in result["novel"]:
            scores = all_scores(f, etalon)
            if scores:
                best_score, best_item = scores[0]
                print(f"  + {fid(f)} ~ #{best_item['n']} (score {best_score})")
                print(f"     Finding: {short(finding_text(f), 110)}")
                print(f"     Эталон:  {short(best_item['text'], 110)}  [{best_item['verdict']}]")
        print()

    if args.json:
        # Преобразуем для сериализации (убираем dict объекты findings)
        out = {
            "project": str(project_path),
            "totals": result["totals"],
            "metrics": result["metrics"],
            "tp": [
                {"finding_id": fid(f), "etalon_n": e["n"], "score": s}
                for f, e, s in result["tp"]
            ],
            "fp": [
                {"finding_id": fid(f), "etalon_n": e["n"], "score": s}
                for f, e, s in result["fp"]
            ],
            "partial": [
                {"finding_id": fid(f), "etalon_n": e["n"], "score": s}
                for f, e, s in result["partial"]
            ],
            "unknown": [
                {"finding_id": fid(f), "etalon_n": e["n"], "score": s}
                for f, e, s in result["unknown"]
            ],
            "novel": [{"finding_id": fid(f)} for f in result["novel"]],
            "missed_keeps": [
                {"etalon_n": e["n"], "text": e["text"]} for e in result["missed_keeps"]
            ],
        }
        Path(args.json).write_text(
            json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"\nJSON сохранён: {args.json}")


if __name__ == "__main__":
    main()
