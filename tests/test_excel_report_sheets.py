"""reserc.md #43 — тесты Excel-генерации (severity-нормализация + скрытый project_id).

generate_excel_report не имел тестов на построение листов. Покрываем
детерминированные части (без изменения прод-кода): нормализацию severity EN→RU и
старо-русских форм, скрытую ячейку project_id (нужна для round-trip импорта
решений), устойчивость к пустым данным.
"""
from __future__ import annotations

from openpyxl import Workbook

import backend.app.pipeline.stages.report.generate_excel_report as ger


# ─── severity нормализация ───────────────────────────────────────────────────

def test_normalize_sev_en_to_ru():
    assert ger.normalize_sev("CRITICAL") == "КРИТИЧЕСКОЕ"
    assert ger.normalize_sev("ECONOMIC") == "ЭКОНОМИЧЕСКОЕ"
    assert ger.normalize_sev("OPERATIONAL") == "ЭКСПЛУАТАЦИОННОЕ"
    assert ger.normalize_sev("CHECK_RELATED") == "ПРОВЕРИТЬ ПО СМЕЖНЫМ"


def test_normalize_sev_old_russian_to_new():
    assert ger.normalize_sev("КРИТИЧНО") == "КРИТИЧЕСКОЕ"
    assert ger.normalize_sev("ПРОВЕРИТЬ") == "ПРОВЕРИТЬ ПО СМЕЖНЫМ"


def test_normalize_sev_passthrough():
    assert ger.normalize_sev("КРИТИЧЕСКОЕ") == "КРИТИЧЕСКОЕ"   # уже канон
    assert ger.normalize_sev("нечто_иное") == "нечто_иное"      # неизвестное — как есть


def test_get_sev_cfg_known_and_fallback():
    assert ger.get_sev_cfg("CRITICAL") == ger.SEVERITY_CONFIG["КРИТИЧЕСКОЕ"]
    # неизвестная severity → безопасный fallback
    assert ger.get_sev_cfg("zzz") == ger.SEVERITY_CONFIG["ПРОВЕРИТЬ ПО СМЕЖНЫМ"]


# ─── скрытый project_id (round-trip импорта решений) ─────────────────────────

def _proj_entry(pid, findings):
    return {
        "project_id": pid,
        "project_info": {"object": "Тестовый дом"},
        "findings_json": {"findings": findings, "meta": {"total_findings": len(findings)}},
        "sheet_name": "ЭМ1",
    }


def test_build_project_sheet_hidden_project_id_cell():
    wb = Workbook()
    pid = "EOM/13АВ-ЭМ1"
    ger.build_project_sheet(wb, _proj_entry(pid, [
        {"severity": "CRITICAL", "finding": "Нет заземления", "page": 1},
    ]))
    ws = wb["ЭМ1"]
    pid_col = len(ger.PROJ_COLUMNS) + 1
    # Скрытая ячейка хранит ПОЛНЫЙ project_id для обратной загрузки решений.
    assert ws.cell(row=2, column=pid_col).value == pid
    assert ws.column_dimensions[ger.get_column_letter(pid_col)].hidden is True


def test_build_project_sheet_empty_findings_no_crash():
    wb = Workbook()
    ger.build_project_sheet(wb, _proj_entry("EOM/empty", []))
    ws = wb["ЭМ1"]
    pid_col = len(ger.PROJ_COLUMNS) + 1
    assert ws.cell(row=2, column=pid_col).value == "EOM/empty"


def test_build_optimization_project_sheet_hidden_project_id():
    wb = Workbook()
    pid = "EOM/13АВ-ЭМ1"
    entry = {
        "project_id": pid,
        "project_info": {"object": "Дом"},
        "optimization_json": {"items": [
            {"title": "Кабель дешевле", "savings_pct": 10, "spec_items": ["Поз. 5"]},
        ], "meta": {}},
        "sheet_name": "ЭМ1",
    }
    ger.build_optimization_project_sheet(wb, entry)
    ws = wb["ОПТ ЭМ1"]
    pid_col = len(ger.OPT_COLUMNS) + 1
    assert ws.cell(row=2, column=pid_col).value == pid


# ─── служебные вкладки в книге ───────────────────────────────────────────────

def test_no_instruction_sheet_builder():
    """Лист «ИНСТРУКЦИЯ» удалён из отчёта — билдера больше нет.

    Пакет аудита (`GET /api/export/audit-package/...`) кладёт внутрь
    `audit_report.xlsx`; лишние служебные вкладки в нём не нужны.
    """
    assert not hasattr(ger, "build_instruction_sheet")


def test_empty_default_sheet_removed_with_no_summary(tmp_path, monkeypatch, capsys):
    """`--no-summary` (так вызывает audit-package) больше не оставляет пустой
    дефолтный лист openpyxl «Sheet»: СВОДКА его не занимает, значит удаляем.
    """
    import sys
    from openpyxl import load_workbook

    out = tmp_path / "report.xlsx"
    project = {
        "project_id": "TEST-PRJ", "folder": str(tmp_path),
        "findings_path": str(tmp_path / "03_findings.json"),
        "optimization_path": str(tmp_path / "optimization.json"),
        "info_path": str(tmp_path / "project_info.json"),
        "has_findings": False, "has_optimization": False,
        "sheet_name": "TEST-PRJ",
    }
    monkeypatch.setattr(ger, "find_projects", lambda *a, **k: [project])
    monkeypatch.setattr(sys, "argv",
                        ["generate_excel_report.py", "--out", str(out), "--no-summary"])
    ger.main()

    names = load_workbook(out).sheetnames
    assert "Sheet" not in names, f"пустой дефолтный лист остался: {names}"
    assert "ИНСТРУКЦИЯ" not in names, f"лист инструкции остался: {names}"
    assert names, "книга не должна остаться без листов"
