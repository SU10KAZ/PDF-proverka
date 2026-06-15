"""Тесты V2-режима вкладки «Расхождения» — pair-scoped верификация инженера.

Ключевые инварианты:
  • V2 показывает расхождения ТОЛЬКО текущей PDF-пары (не всю сессию).
  • Ручные статусы хранятся отдельно в `pairs/<pid>/v2_review_status.json`.
  • `comparison_result.json` НИКОГДА не мутируется ручной разметкой.
  • Read-only: ни Qwen, ни Opus, ни unified-analysis не запускаются.
"""
from __future__ import annotations

import io
import json
from datetime import datetime
from pathlib import Path

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook


# ─── Фикстуры: tmp comparison root + сессия с двумя парами ────────────────


@pytest.fixture(autouse=True)
def _tmp_comparison_root(tmp_path, monkeypatch):
    root = tmp_path / "comparison_v2_test"
    root.mkdir()
    monkeypatch.setenv("COMPARISON_ROOT", str(root))
    yield root


def _paths():
    from backend.app.services.stage_comparison import paths as paths_mod
    return paths_mod


def _write_session(session_id: str, pair_ids: list[str]):
    paths_mod = _paths()
    session = {
        "id": session_id,
        "pair_order": list(pair_ids),
        "warnings": [],
        "created_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
    }
    paths_mod.session_json_path(session_id).write_text(
        json.dumps(session, ensure_ascii=False), encoding="utf-8")


def _write_pair(session_id: str, pair_id: str, left_name: str, right_name: str):
    paths_mod = _paths()
    pair = {
        "id": pair_id,
        "status": "matched",
        "left": {"filename": left_name, "pdf_path": f"/dev/null/{left_name}"},
        "right": {"filename": right_name, "pdf_path": f"/dev/null/{right_name}"},
    }
    paths_mod.pair_json_path(session_id, pair_id).write_text(
        json.dumps(pair, ensure_ascii=False), encoding="utf-8")


def _write_comparison_result(session_id: str, pair_id: str, changes: list[dict]):
    paths_mod = _paths()
    p = paths_mod.enriched_comparison_result_path(session_id, pair_id)
    p.parent.mkdir(parents=True, exist_ok=True)
    payload = {"version": 1, "status": "done", "changes": changes}
    p.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return p


def _change(cid, *, title, sev="medium", source="text", old="", new="",
            rhr=False, conf=0.9, cost="unknown"):
    return {
        "id": cid,
        "source": source,
        "type": "changed",
        "category": "general",
        "severity": sev,
        "title": title,
        "summary": f"summary {title}",
        "old_value": old,
        "new_value": new,
        "construction_impact": "влияние",
        "cost_impact": cost,
        "requires_human_review": rhr,
        "confidence": conf,
        "evidence_left": {"quote": old or "L", "section": "X", "approx_location": "стр. 1"},
        "evidence_right": {"quote": new or "R", "section": "X", "approx_location": "стр. 1"},
    }


@pytest.fixture
def session_with_two_pairs():
    sid = "sess_v2"
    p1, p2 = "pAAA", "pBBB"
    _write_session(sid, [p1, p2])
    _write_pair(sid, p1, "left1.pdf", "right1.pdf")
    _write_pair(sid, p2, "left2.pdf", "right2.pdf")
    # Пара 1 — три изменения (одно high, одно нужна ручная проверка, одно cost).
    _write_comparison_result(sid, p1, [
        _change("chg_p1_a", title="P1 высокая", sev="high", old="A1", new="A2"),
        _change("chg_p1_b", title="P1 ручная", sev="low", rhr=True, old="B1", new="B2"),
        _change("chg_p1_c", title="P1 стоимость", sev="medium", cost="likely", old="C1", new="C2"),
    ])
    # Пара 2 — два изменения с ДРУГИМИ id и заголовками.
    _write_comparison_result(sid, p2, [
        _change("chg_p2_x", title="P2 первое", sev="medium", old="X1", new="X2"),
        _change("chg_p2_y", title="P2 второе", sev="high", old="Y1", new="Y2"),
    ])
    return {"sid": sid, "p1": p1, "p2": p2}


def _client():
    from backend.app.api.routers import stage_comparison as router_mod
    app = FastAPI()
    app.include_router(router_mod.router)
    return TestClient(app)


def _v2_base(sid, pid):
    return f"/api/stage-comparison/sessions/{sid}/pairs/{pid}/v2"


# ─── 1/2. Pair-scoping ───────────────────────────────────────────────────


def test_v2_changes_returns_only_current_pair(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    r = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes")
    assert r.status_code == 200
    data = r.json()
    assert data["pair_id"] == s["p1"]
    titles = {it["title"] for it in data["items"]}
    assert titles == {"P1 высокая", "P1 ручная", "P1 стоимость"}
    assert all(it["pair_id"] == s["p1"] for it in data["items"])


def test_v2_changes_does_not_leak_other_pairs(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    r = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes")
    titles = {it["title"] for it in r.json()["items"]}
    # Заголовки из пары 2 не должны попасть в выдачу пары 1.
    assert "P2 первое" not in titles
    assert "P2 второе" not in titles


def test_v2_changes_pair2_independent(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    r = client.get(f"{_v2_base(s['sid'], s['p2'])}/changes")
    titles = {it["title"] for it in r.json()["items"]}
    assert titles == {"P2 первое", "P2 второе"}


# ─── 3. Summary только по текущей паре ───────────────────────────────────


def test_v2_summary_scoped_to_pair(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    r = client.get(f"{_v2_base(s['sid'], s['p1'])}/summary")
    assert r.status_code == 200
    summ = r.json()["summary"]
    assert summ["total"] == 3       # только 3 изменения пары 1, не 5
    assert summ["high"] == 1
    assert summ["medium"] == 1
    assert summ["low"] == 1
    assert summ["needs_human_review"] == 1
    assert summ["not_reviewed"] == 3
    assert summ["confirmed"] == 0


def test_v2_summary_embedded_in_changes(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    r = client.get(f"{_v2_base(s['sid'], s['p2'])}/changes")
    assert r.json()["summary"]["total"] == 2


# ─── 4. PATCH одного изменения ───────────────────────────────────────────


def test_v2_patch_saves_review_status(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    changes = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"]
    cid = changes[0]["id"]
    r = client.patch(f"{_v2_base(s['sid'], s['p1'])}/changes/{cid}",
                     json={"review_status": "confirmed", "review_comment": "Проверено инженером"})
    assert r.status_code == 200
    # Перечитываем — статус сохранился.
    again = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()
    target = next(it for it in again["items"] if it["id"] == cid)
    assert target["review_status"] == "confirmed"
    assert target["review_comment"] == "Проверено инженером"
    assert target["reviewed_at"]
    assert again["summary"]["confirmed"] == 1
    assert again["summary"]["not_reviewed"] == 2


def test_v2_patch_rejects_invalid_status(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    cid = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"][0]["id"]
    r = client.patch(f"{_v2_base(s['sid'], s['p1'])}/changes/{cid}",
                     json={"review_status": "totally_bogus"})
    assert r.status_code == 400


def test_v2_patch_unknown_change_404(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    r = client.patch(f"{_v2_base(s['sid'], s['p1'])}/changes/v2_doesnotexist",
                     json={"review_status": "confirmed"})
    assert r.status_code == 404


def test_v2_patch_change_of_other_pair_404(session_with_two_pairs):
    """id из пары 2 нельзя пропатчить через endpoint пары 1."""
    s = session_with_two_pairs
    client = _client()
    p2_id = client.get(f"{_v2_base(s['sid'], s['p2'])}/changes").json()["items"][0]["id"]
    r = client.patch(f"{_v2_base(s['sid'], s['p1'])}/changes/{p2_id}",
                     json={"review_status": "confirmed"})
    assert r.status_code == 404


# ─── 5. Bulk PATCH в рамках пары ─────────────────────────────────────────


def test_v2_bulk_patch_scoped_to_pair(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    p1_items = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"]
    p2_items = client.get(f"{_v2_base(s['sid'], s['p2'])}/changes").json()["items"]
    p1_ids = [it["id"] for it in p1_items]
    foreign_id = p2_items[0]["id"]
    r = client.patch(f"{_v2_base(s['sid'], s['p1'])}/changes",
                     json={"ids": p1_ids + [foreign_id],
                           "patch": {"review_status": "rejected",
                                     "review_comment": "Пакетно отклонено"}})
    assert r.status_code == 200
    body = r.json()
    assert set(body["updated"]) == set(p1_ids)
    assert foreign_id in body["skipped"]      # чужой id не применился
    again = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()
    assert again["summary"]["rejected"] == 3
    # Пара 2 осталась нетронутой.
    p2_again = client.get(f"{_v2_base(s['sid'], s['p2'])}/changes").json()
    assert p2_again["summary"]["rejected"] == 0


# ─── 6. Хранение в pairs/<pid>/v2_review_status.json ─────────────────────


def test_v2_status_stored_in_pair_file(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    cid = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"][0]["id"]
    client.patch(f"{_v2_base(s['sid'], s['p1'])}/changes/{cid}",
                 json={"review_status": "confirmed"})
    status_path = _paths().v2_review_status_path(s["sid"], s["p1"])
    assert status_path.exists()
    assert status_path.name == "v2_review_status.json"
    data = json.loads(status_path.read_text(encoding="utf-8"))
    assert cid in data["items"]
    assert data["items"][cid]["review_status"] == "confirmed"
    # Файл пары 2 не создан (мы её не размечали).
    assert not _paths().v2_review_status_path(s["sid"], s["p2"]).exists()


# ─── 7. comparison_result.json не мутируется ─────────────────────────────


def test_v2_does_not_mutate_comparison_result(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    cr_path = _paths().enriched_comparison_result_path(s["sid"], s["p1"])
    before = cr_path.read_bytes()
    cid = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"][0]["id"]
    client.patch(f"{_v2_base(s['sid'], s['p1'])}/changes/{cid}",
                 json={"review_status": "confirmed", "review_comment": "x"})
    client.patch(f"{_v2_base(s['sid'], s['p1'])}/changes",
                 json={"ids": [cid], "patch": {"review_status": "rejected"}})
    after = cr_path.read_bytes()
    assert before == after, "comparison_result.json не должен меняться при ручной разметке"


# ─── 8. Export XLSX только текущей пары ──────────────────────────────────


def test_v2_export_only_current_pair(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    r = client.get(f"{_v2_base(s['sid'], s['p1'])}/export.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == [
        "Summary", "All V2 changes", "Confirmed", "Needs clarification",
        "Rejected", "Cost impact", "Not reviewed",
    ]
    ws = wb["All V2 changes"]
    # Заголовок + 3 строки (только пара 1).
    titles = [ws.cell(row=ri, column=7).value for ri in range(2, ws.max_row + 1)]
    assert set(titles) == {"P1 высокая", "P1 ручная", "P1 стоимость"}
    assert "P2 первое" not in titles


def test_v2_export_confirmed_sheet_reflects_status(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    cid = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"][0]["id"]
    client.patch(f"{_v2_base(s['sid'], s['p1'])}/changes/{cid}",
                 json={"review_status": "confirmed"})
    r = client.get(f"{_v2_base(s['sid'], s['p1'])}/export.xlsx")
    wb = load_workbook(io.BytesIO(r.content))
    confirmed = wb["Confirmed"]
    assert confirmed.max_row == 2     # header + 1 confirmed row


# ─── id-стабильность ─────────────────────────────────────────────────────


def test_v2_ids_are_stable_across_rebuilds(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    a = {it["title"]: it["id"] for it in client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"]}
    b = {it["title"]: it["id"] for it in client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"]}
    assert a == b
    assert all(v.startswith("v2_") for v in a.values())


# ─── 404 для несуществующей пары/сессии ──────────────────────────────────


def test_v2_unknown_pair_404(session_with_two_pairs):
    s = session_with_two_pairs
    client = _client()
    r = client.get(f"{_v2_base(s['sid'], 'pNOPE')}/changes")
    assert r.status_code == 404


# ─── 11. Никаких Qwen/Opus вызовов ───────────────────────────────────────


def test_v2_does_not_invoke_llm_providers(session_with_two_pairs, monkeypatch):
    """V2 read-only: ни локальный Qwen, ни Opus-provider не дёргаются."""
    s = session_with_two_pairs

    from backend.app.services.stage_comparison import enriched_comparison as ec
    from backend.app.services.stage_comparison import graphic_llm_local as gl

    def _boom(*a, **k):
        raise AssertionError("LLM provider must NOT be called by V2 read flow")

    # Любая попытка реально сравнить/описать через модель = провал теста.
    monkeypatch.setattr(ec, "run_enriched_comparison", _boom, raising=False)
    monkeypatch.setattr(gl, "describe_image_local", _boom, raising=False)

    client = _client()
    assert client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").status_code == 200
    cid = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"][0]["id"]
    assert client.patch(f"{_v2_base(s['sid'], s['p1'])}/changes/{cid}",
                        json={"review_status": "confirmed"}).status_code == 200
    assert client.get(f"{_v2_base(s['sid'], s['p1'])}/export.xlsx").status_code == 200


# ─── UI parity (кнопка V2 + старая вкладка цела) ─────────────────────────

_ROOT = Path(__file__).resolve().parent.parent


def test_derive_quality_label_disputed_is_questionable():
    """r5: disputed=True → questionable (когда не requires_human_review)."""
    from backend.app.services.stage_comparison import v2_review as v2r
    assert v2r.derive_quality_label(
        {"disputed": True, "requires_human_review": False, "confidence": 0.9}
    ) == "questionable"
    # requires_human_review имеет приоритет
    assert v2r.derive_quality_label(
        {"disputed": True, "requires_human_review": True}
    ) == "needs_human_review"
    # без disputed и прочих сигналов — good
    assert v2r.derive_quality_label(
        {"disputed": False, "requires_human_review": False, "confidence": 0.9}
    ) == "good"


def test_ui_no_separate_v2_toggle():
    """Раздельного переключателя [Расхождения][V2] в UI быть не должно —
    V2 теперь единый основной режим. Старые always-visible кнопки убраны."""
    html = (_ROOT / "frontend" / "index.html").read_text(encoding="utf-8")
    # Двух отдельных кнопок-переключателей больше нет.
    assert "scSetV2View('v2')" not in html
    assert "scSetV2View('current')" not in html
    # Состояние режима осталось (используется только dev-переключателем).
    assert "scV2View" in html


def test_ui_v2_is_default_mode():
    """V2 — режим по умолчанию во вкладке «Расхождения»."""
    app_js = (_ROOT / "frontend" / "static" / "js" / "app.js").read_text(encoding="utf-8")
    norm = " ".join(app_js.split())
    assert "const scV2View = ref('v2')" in norm


def test_ui_old_discrepancies_tab_intact():
    """Старый unified-режим вкладки «Расхождения» оставлен для отладки (не удалён)."""
    html = (_ROOT / "frontend" / "index.html").read_text(encoding="utf-8")
    assert "scUnifiedItemsSorted" in html        # старая таблица расхождений
    assert "scDiffSubtab==='unified'" in html


def test_old_unified_diff_flat_endpoints_present():
    """Старые unified-diff-flat endpoints не удалены (read-only/debug), а V2
    endpoint существует и используется как основной источник."""
    from backend.app.api.routers import stage_comparison as router_mod
    paths = {getattr(r, "path", "") for r in router_mod.router.routes}
    assert any("/unified-diff-flat" in p for p in paths)
    assert any(p.endswith("/v2/changes") for p in paths)
    assert any(p.endswith("/v2/export.xlsx") for p in paths)


def test_ui_app_js_exposes_v2_methods():
    app_js = (_ROOT / "frontend" / "static" / "js" / "app.js").read_text(encoding="utf-8")
    assert "scLoadV2Changes" in app_js
    assert "scV2View" in app_js


# ═══════════════════════════════════════════════════════════════════════════
#  Impact classification: исключение admin / documentation / cosmetic из V2
# ═══════════════════════════════════════════════════════════════════════════


@pytest.fixture
def session_mixed_impact():
    """Пара с тремя инженерными и тремя исключаемыми изменениями."""
    sid, pid = "sess_impact", "pIMP"
    _write_session(sid, [pid])
    _write_pair(sid, pid, "old.pdf", "new.pdf")
    _write_comparison_result(sid, pid, [
        # инженерные (остаются):
        _change("chg_eng_beton", title="Класс бетона B25 на B30", sev="medium", old="B25", new="B30"),
        _change("chg_cost_guard", title="Корректировка штампа", sev="low", cost="likely", old="s1", new="s2"),
        _change("chg_rhr_eng", title="Пересмотр нагрузок по щиту", sev="low", rhr=True, old="n1", new="n2"),
        # исключаемые:
        _change("chg_admin_org", title="Сменилась проектная организация и ГИП, подпись", sev="low", old="o1", new="o2"),
        _change("chg_doc_shifr", title="Изменён шифр, номер листа и дата выпуска", sev="low", old="d1", new="d2"),
        _change("chg_cosm", title="Переформулировка: значение не изменилось", sev="low", old="c1", new="c2"),
    ])
    return {"sid": sid, "pid": pid}


def _items_by_raw(items):
    return {it["raw_id"]: it for it in items}


# 1/3/4. admin / documentation / cosmetic скрыты по умолчанию.
def test_v2_excludes_admin_doc_cosmetic_by_default(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    data = client.get(f"{_v2_base(s['sid'], s['pid'])}/changes").json()
    raws = {it["raw_id"] for it in data["items"]}
    assert "chg_admin_org" not in raws          # admin_only скрыт
    assert "chg_doc_shifr" not in raws          # documentation_only скрыт
    assert "chg_cosm" not in raws               # cosmetic_or_noise скрыт
    # инженерные остались:
    assert {"chg_eng_beton", "chg_cost_guard", "chg_rhr_eng"} <= raws
    assert len(data["items"]) == 3


# 2. include_excluded=true возвращает всё с флагами.
def test_v2_include_excluded_returns_all_with_flags(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    data = client.get(f"{_v2_base(s['sid'], s['pid'])}/changes?include_excluded=true").json()
    assert len(data["items"]) == 6
    by = _items_by_raw(data["items"])
    admin = by["chg_admin_org"]
    assert admin["impact_class"] == "admin_only"
    assert admin["excluded_from_main"] is True
    assert admin["exclusion_reason"]
    assert by["chg_doc_shifr"]["impact_class"] == "documentation_only"
    assert by["chg_cosm"]["impact_class"] == "cosmetic_or_noise"
    # инженерные не помечены excluded
    assert by["chg_eng_beton"]["excluded_from_main"] is False
    assert by["chg_eng_beton"]["impact_class"] not in v2_excluded_set()


def v2_excluded_set():
    from backend.app.services.stage_comparison import v2_review as v2r
    return v2r.EXCLUDED_IMPACT_CLASSES


# 5. инженерное изменение остаётся.
def test_v2_engineering_change_remains(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    data = client.get(f"{_v2_base(s['sid'], s['pid'])}/changes").json()
    by = _items_by_raw(data["items"])
    assert "chg_eng_beton" in by
    assert by["chg_eng_beton"]["impact_class"] == "construction_technical_impact"


# 5b. cost_direction (денежный эффект) присутствует у каждого изменения —
#     питает чип «удорожание/удешевление/нейтрально» в колонке «№».
def test_v2_items_have_cost_direction(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    data = client.get(f"{_v2_base(s['sid'], s['pid'])}/changes").json()
    assert data["items"], "ожидались инженерные изменения"
    allowed = {"increase", "decrease", "unknown", "neutral"}
    for it in data["items"]:
        assert it.get("cost_direction") in allowed


# 6. cost_impact=likely не скрывается (даже если выглядит как штамп).
def test_v2_cost_impact_not_excluded(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    data = client.get(f"{_v2_base(s['sid'], s['pid'])}/changes").json()
    by = _items_by_raw(data["items"])
    assert "chg_cost_guard" in by               # не исключено
    assert by["chg_cost_guard"]["impact_class"] == "manual_review_required"


# 7. requires_human_review инженерное остаётся.
def test_v2_rhr_engineering_remains(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    data = client.get(f"{_v2_base(s['sid'], s['pid'])}/changes").json()
    by = _items_by_raw(data["items"])
    assert "chg_rhr_eng" in by
    assert by["chg_rhr_eng"]["impact_class"] == "engineering_system_impact"


# Summary: разбивка исключённых.
def test_v2_summary_exclusion_breakdown(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    summ = client.get(f"{_v2_base(s['sid'], s['pid'])}/changes").json()["summary"]
    assert summ["engineering_total"] == 3
    assert summ["excluded_total"] == 3
    assert summ["excluded_admin_only"] == 1
    assert summ["excluded_documentation_only"] == 1
    assert summ["excluded_cosmetic_or_noise"] == 1


# 8. v2_excluded_changes.json создаётся.
def test_v2_excluded_changes_file_written(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    client.get(f"{_v2_base(s['sid'], s['pid'])}/changes")
    p = _paths().v2_excluded_changes_path(s["sid"], s["pid"])
    assert p.exists()
    assert p.name == "v2_excluded_changes.json"
    data = json.loads(p.read_text(encoding="utf-8"))
    # 3 исключённых изменения с impact_class + причиной + контентом.
    assert len(data["items"]) == 3
    entry = next(iter(data["items"].values()))
    assert entry["impact_class"] in v2_excluded_set()
    assert entry["exclusion_reason"]
    assert "source_title" in entry


# comparison_result.json не мутируется фильтром.
def test_v2_filter_does_not_mutate_comparison_result(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    cr = _paths().enriched_comparison_result_path(s["sid"], s["pid"])
    before = cr.read_bytes()
    client.get(f"{_v2_base(s['sid'], s['pid'])}/changes")
    client.get(f"{_v2_base(s['sid'], s['pid'])}/changes?include_excluded=true")
    assert cr.read_bytes() == before


# 9. review_status для excluded item сохраняется и виден при include_excluded.
def test_v2_review_status_preserved_for_excluded(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    # id исключённого изменения берём из include_excluded.
    alld = client.get(f"{_v2_base(s['sid'], s['pid'])}/changes?include_excluded=true").json()
    admin = _items_by_raw(alld["items"])["chg_admin_org"]
    cid = admin["id"]
    # PATCH разрешён даже для исключённого.
    r = client.patch(f"{_v2_base(s['sid'], s['pid'])}/changes/{cid}",
                     json={"review_status": "rejected", "review_comment": "оформление"})
    assert r.status_code == 200
    # статус хранится в v2_review_status.json
    sp = _paths().v2_review_status_path(s["sid"], s["pid"])
    assert cid in json.loads(sp.read_text(encoding="utf-8"))["items"]
    # при include_excluded=true статус виден
    again = client.get(f"{_v2_base(s['sid'], s['pid'])}/changes?include_excluded=true").json()
    assert _items_by_raw(again["items"])["chg_admin_org"]["review_status"] == "rejected"
    # в основной таблице строка скрыта, но не потеряна
    main = client.get(f"{_v2_base(s['sid'], s['pid'])}/changes").json()
    assert "chg_admin_org" not in {it["raw_id"] for it in main["items"]}


# 10. export.xlsx по умолчанию не содержит excluded.
def test_v2_export_default_excludes(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    r = client.get(f"{_v2_base(s['sid'], s['pid'])}/export.xlsx")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert "Excluded admin-doc-noise" not in wb.sheetnames
    ws = wb["All V2 changes"]
    titles = [ws.cell(row=ri, column=7).value for ri in range(2, ws.max_row + 1)]
    assert "Сменилась проектная организация и ГИП, подпись" not in titles
    assert "Класс бетона B25 на B30" in titles
    # Summary показывает excluded breakdown.
    sumvals = {ws2.value for ws2 in wb["Summary"]["A"]}
    assert "Исключено всего" in sumvals


# 11. export.xlsx?include_excluded=true содержит excluded sheet.
def test_v2_export_include_excluded_sheet(session_mixed_impact):
    s = session_mixed_impact
    client = _client()
    r = client.get(f"{_v2_base(s['sid'], s['pid'])}/export.xlsx?include_excluded=true")
    wb = load_workbook(io.BytesIO(r.content))
    assert "Excluded admin-doc-noise" in wb.sheetnames
    ws = wb["Excluded admin-doc-noise"]
    titles = [ws.cell(row=ri, column=7).value for ri in range(2, ws.max_row + 1)]
    assert set(titles) == {
        "Сменилась проектная организация и ГИП, подпись",
        "Изменён шифр, номер листа и дата выпуска",
        "Переформулировка: значение не изменилось",
    }


# 12. UI больше НЕ содержит переключатель «Показать административные /
#     оформление / шум»: исключённые изменения просто не показываются.
def test_ui_has_no_show_excluded_toggle():
    html = (_ROOT / "frontend" / "index.html").read_text(encoding="utf-8")
    assert "Показать административные / оформление / шум" not in html
    assert "scV2ShowExcluded" not in html
    app_js = (_ROOT / "frontend" / "static" / "js" / "app.js").read_text(encoding="utf-8")
    assert "scV2ShowExcluded" not in app_js
    assert "scV2ToggleShowExcluded" not in app_js
    # колонка «№» больше не запрашивает исключённые
    assert "include_excluded=true" not in app_js


# 12b. Колонка «№» показывает только источник сравнения (текст/изображение)
#      и денежный эффект (удорожание/удешевление/нейтрально).
def test_ui_no_column_shows_source_and_cost_direction():
    html = (_ROOT / "frontend" / "index.html").read_text(encoding="utf-8")
    assert "scCostDirectionLabel(it.cost_direction)" in html
    assert "scUnifiedSourceLabel(it.source_layer)" in html
    # старые бейджи убраны из ячейки «№» V2-ведомости (номер/severity/impact)
    assert "№{{ idx + 1 }}" not in html
    assert "scV2ImpactBadgeStyle(it.impact_class)" not in html
    app_js = (_ROOT / "frontend" / "static" / "js" / "app.js").read_text(encoding="utf-8")
    assert "scCostDirectionLabel" in app_js
    assert "scCostDirectionStyle" in app_js


# 13. No Qwen/Opus при фильтрации.
def test_v2_impact_filter_no_llm(session_mixed_impact, monkeypatch):
    s = session_mixed_impact
    from backend.app.services.stage_comparison import enriched_comparison as ec
    from backend.app.services.stage_comparison import graphic_llm_local as gl

    def _boom(*a, **k):
        raise AssertionError("LLM must NOT be called by V2 impact filter")

    monkeypatch.setattr(ec, "run_enriched_comparison", _boom, raising=False)
    monkeypatch.setattr(gl, "describe_image_local", _boom, raising=False)
    client = _client()
    assert client.get(f"{_v2_base(s['sid'], s['pid'])}/changes").status_code == 200
    assert client.get(f"{_v2_base(s['sid'], s['pid'])}/changes?include_excluded=true").status_code == 200
    assert client.get(f"{_v2_base(s['sid'], s['pid'])}/export.xlsx?include_excluded=true").status_code == 200


# ═══════════════════════════════════════════════════════════════════════════
#  Canonical review storage + legacy expert_review.json fallback (ТЗ Задача 2)
#  Приоритет: v2_review_status.json → expert_review.json → not_reviewed.
# ═══════════════════════════════════════════════════════════════════════════


def _write_expert_review(session_id: str, key: str, decision: str, reason: str = ""):
    """Записать legacy-решение эксперта `<pair_id>::<raw_id>` в expert_review.json."""
    paths_mod = _paths()
    p = paths_mod.expert_review_path(session_id)
    p.parent.mkdir(parents=True, exist_ok=True)
    existing = {}
    if p.exists():
        existing = json.loads(p.read_text(encoding="utf-8"))
    decisions = existing.get("decisions") or {}
    decisions[key] = {"decision": decision, "rejection_reason": reason,
                      "reviewer": "tester", "timestamp": "2026-06-04T00:00:00Z"}
    existing.update({"version": 2, "decisions": decisions, "updated_at": "2026-06-04T00:00:00Z"})
    p.write_text(json.dumps(existing, ensure_ascii=False), encoding="utf-8")


def test_v2_falls_back_to_expert_review_by_v2_id(session_with_two_pairs):
    """Решение эксперта по v2_id (V2-native) → review_status в V2-строке."""
    s = session_with_two_pairs
    client = _client()
    items = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"]
    target = items[0]
    _write_expert_review(s["sid"], f"{s['p1']}::{target['id']}", "accepted", "ок инженер")
    again = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()
    row = next(it for it in again["items"] if it["id"] == target["id"])
    assert row["review_status"] == "confirmed"
    assert row["review_source"] == "expert_review"
    assert row["review_comment"] == "ок инженер"
    assert again["summary"]["confirmed"] == 1


def test_v2_falls_back_to_expert_review_by_classic_raw_id(session_with_two_pairs):
    """Старое решение из «Расхождений» по классическому chg_-id видно в V2."""
    s = session_with_two_pairs
    client = _client()
    items = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"]
    target = next(it for it in items if it["raw_id"] == "chg_p1_a")
    # Эксперт размечал в классических «Расхождениях» — ключ по raw_id (chg_…).
    _write_expert_review(s["sid"], f"{s['p1']}::chg_p1_a", "rejected", "не подтверждается")
    again = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()
    row = next(it for it in again["items"] if it["id"] == target["id"])
    assert row["review_status"] == "rejected"
    assert row["review_source"] == "expert_review"


def test_v2_status_takes_priority_over_expert_review(session_with_two_pairs):
    """Канонический v2_review_status.json НЕ перезаписывается expert fallback."""
    s = session_with_two_pairs
    client = _client()
    items = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"]
    target = items[0]
    # expert говорит rejected, но инженер в V2 поставил confirmed → V2 побеждает.
    _write_expert_review(s["sid"], f"{s['p1']}::{target['id']}", "rejected", "old")
    client.patch(f"{_v2_base(s['sid'], s['p1'])}/changes/{target['id']}",
                 json={"review_status": "confirmed", "review_comment": "новое"})
    again = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()
    row = next(it for it in again["items"] if it["id"] == target["id"])
    assert row["review_status"] == "confirmed"
    assert row["review_source"] == "v2_review_status"
    assert row["review_comment"] == "новое"


def test_v2_no_expert_review_stays_not_reviewed(session_with_two_pairs):
    """Без expert_review.json и без v2-статуса строка остаётся not_reviewed."""
    s = session_with_two_pairs
    client = _client()
    data = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()
    assert all(it["review_status"] == "not_reviewed" for it in data["items"])
    assert all(it["review_source"] == "none" for it in data["items"])


def test_v2_expert_fallback_does_not_mutate_expert_review(session_with_two_pairs):
    """Чтение V2 не мутирует expert_review.json (read-only fallback)."""
    s = session_with_two_pairs
    client = _client()
    items = client.get(f"{_v2_base(s['sid'], s['p1'])}/changes").json()["items"]
    _write_expert_review(s["sid"], f"{s['p1']}::{items[0]['id']}", "accepted")
    p = _paths().expert_review_path(s["sid"])
    before = p.read_bytes()
    client.get(f"{_v2_base(s['sid'], s['p1'])}/changes")
    client.get(f"{_v2_base(s['sid'], s['p1'])}/changes?include_excluded=true")
    assert p.read_bytes() == before
