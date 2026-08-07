from __future__ import annotations

import importlib.util
import json
from pathlib import Path

from openpyxl import load_workbook


_MODULE_PATH = Path(__file__).resolve().parents[1] / "scripts" / "export_rejected_findings_excel.py"
_SPEC = importlib.util.spec_from_file_location("export_rejected_findings_excel", _MODULE_PATH)
exporter = importlib.util.module_from_spec(_SPEC)
assert _SPEC.loader is not None
_SPEC.loader.exec_module(exporter)
build_workbook = exporter.build_workbook


def _write_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def _write_jsonl(path: Path, rows: list[dict]) -> None:
    path.write_text(
        "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows),
        encoding="utf-8",
    )


def test_all_results_are_traceable_and_use_normalized_outcomes(tmp_path: Path) -> None:
    first_id = "RF-20260701-aaaaaaaaaaaaaaaa"
    second_id = "RF-20260702-bbbbbbbbbbbbbbbb"
    results = [
        {
            "case_id": first_id,
            "status": "success",
            "verdict": "expert_correct",
            "raw_verdict": "expert_may_be_wrong",
            "confidence": "high",
            "recommended_action": "keep_rejected",
            "decision_origin": "human",
            "binding_status": "exact",
            "factual_verdict": "contradicted",
            "report_value": "remove",
            "reason_quality": "substantiated",
            "object_name": "Объект 1",
            "discipline": "KJ",
            "document": "DOC-1",
            "version_id": "v001",
            "item_id": "F-001",
            "expert_timestamp": "2026-07-01T09:00:00+00:00",
            "finding_problem": "Замечание 1",
            "expert_reason": "Причина 1",
            "reason_assessment": "Оценка причины 1",
            "finding_assessment": "Оценка замечания 1",
            "norm_assessment": "Оценка нормы 1",
            "decisive_evidence": [],
            "missing_context": [],
            "guard_adjustments": [],
        },
        {
            "case_id": second_id,
            "status": "success",
            "verdict": "insufficient_evidence",
            "raw_verdict": "expert_correct",
            "confidence": "medium",
            "recommended_action": "manual_recheck",
            "decision_origin": "human",
            "binding_status": "missing",
            "factual_verdict": "unclear",
            "report_value": "unclear",
            "reason_quality": "missing",
            "object_name": "Объект 2",
            "discipline": "AR",
            "document": "DOC-2",
            "version_id": "v002",
            "item_id": "F-002",
            "expert_timestamp": "2026-07-02T10:00:00+00:00",
            "finding_problem": "Замечание 2",
            "expert_reason": "",
            "reason_assessment": "Оценка причины 2",
            "finding_assessment": "Оценка замечания 2",
            "norm_assessment": "Оценка нормы 2",
            "decisive_evidence": [],
            "missing_context": ["Нужен лист 2"],
            "guard_adjustments": ["verdict downgraded: exact finding binding is missing"],
        },
    ]
    manifest = [
        {
            "case_id": row["case_id"],
            "expert_timestamp_local": row["expert_timestamp"],
            "source_item_path": f"/tmp/{row['item_id']}.json",
            "review_path": f"/tmp/{row['item_id']}-review.json",
        }
        for row in results
    ]
    summary = {
        "period": "2026-07",
        "selected_cases": 2,
        "completed": 2,
        "latest_errors": 0,
        "generated_at": "2026-08-05T00:00:00+00:00",
    }

    _write_json(tmp_path / "summary.json", summary)
    _write_jsonl(tmp_path / "results.jsonl", results)
    _write_jsonl(tmp_path / "manifest.jsonl", manifest)
    _write_json(
        tmp_path / "candidates.json",
        {"candidate_count": 1, "candidates": [{"case_id": second_id}]},
    )

    output = tmp_path / "report.xlsx"
    assert build_workbook(tmp_path, output) == (2, 1)

    workbook = load_workbook(output, read_only=True, data_only=False)
    assert workbook.sheetnames == ["Сводка", "Все результаты", "Кандидаты"]
    sheet = workbook["Все результаты"]
    headers = [cell.value for cell in sheet[1]]
    required = {
        "ID кейса",
        "Итог аудита",
        "Рекомендуемое действие",
        "Уверенность",
        "Статус анализа",
        "Объект",
        "Раздел",
        "Документ",
        "Версия",
        "ID замечания",
        "Дата решения (МСК)",
    }
    assert required <= set(headers)
    assert "raw_verdict" not in headers
    assert "Сырой вердикт" not in headers


def test_append_only_retry_log_exports_latest_success(tmp_path: Path) -> None:
    case_id = "RF-20260703-cccccccccccccccc"
    input_hash = "sha256-current"
    failed_attempt = {
        "case_id": case_id,
        "input_hash": input_hash,
        "status": "error",
        "error": "temporary runner failure",
    }
    successful_retry = {
        "case_id": case_id,
        "input_hash": input_hash,
        "status": "success",
        "verdict": "insufficient_evidence",
        "confidence": "medium",
        "recommended_action": "collect_context",
        "review_priority": "none",
        "decision_origin": "human",
        "binding_status": "exact",
        "factual_verdict": "unclear",
        "report_value": "unclear",
        "reason_quality": "unsubstantiated",
        "object_name": "Объект",
        "discipline": "KJ",
        "document": "DOC-3",
        "version_id": "v003",
        "item_id": "F-003",
        "expert_timestamp": "2026-07-03T09:00:00+00:00",
        "finding_problem": "Замечание",
        "expert_reason": "Причина",
        "reason_assessment": "Недостаточно данных",
        "finding_assessment": "Нужен полный лист",
        "norm_assessment": "Норма не проверена",
        "decisive_evidence": [],
        "missing_context": ["Полный лист"],
        "guard_adjustments": [],
    }
    manifest = [
        {
            "case_id": case_id,
            "input_hash": input_hash,
            "expert_timestamp_local": successful_retry["expert_timestamp"],
            "source_item_path": "/tmp/F-003.json",
            "review_path": "/tmp/F-003-review.json",
        }
    ]

    _write_json(
        tmp_path / "summary.json",
        {
            "period": "2026-07",
            "selected_cases": 1,
            "completed": 1,
            "latest_errors": 0,
            "generated_at": "2026-08-05T00:00:00+00:00",
        },
    )
    _write_jsonl(tmp_path / "results.jsonl", [failed_attempt, successful_retry])
    _write_jsonl(tmp_path / "manifest.jsonl", manifest)
    _write_json(tmp_path / "candidates.json", {"candidate_count": 0, "candidates": []})

    output = tmp_path / "retry-report.xlsx"
    assert build_workbook(tmp_path, output) == (1, 0)

    workbook = load_workbook(output, read_only=True, data_only=False)
    sheet = workbook["Все результаты"]
    headers = [cell.value for cell in sheet[1]]
    row = [cell.value for cell in sheet[2]]
    assert dict(zip(headers, row))["ID кейса"] == case_id
    assert dict(zip(headers, row))["Статус анализа"] == "Успешно"
