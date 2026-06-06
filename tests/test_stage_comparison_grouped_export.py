"""Тесты grouped XLSX-экспорта вкладки «Отчёт».

Кнопка «⬇ Экспортировать всё в XLSX» зовёт
`/sessions/{sid}/unified-diff-flat/export.xlsx?accepted_only=true&grouped=true`.
Отчёт должен повторять UI: каждая PDF-пара — сворачиваемый раздел
(Excel-группировка, открывается «плюсиком»), внутри — таблица сравнения
с колонками №, Место, Изменение, Было, Стало, Влияние.
"""
import io

from openpyxl import load_workbook

from backend.app.api.routers.stage_comparison import (
    _build_grouped_comparison_workbook,
    _sc_split_lines,
    _sc_page_str,
)


def _sample_items():
    return [
        {
            "pair_id": "pf06effb7",
            "pair_label": "OLD.pdf ↔ NEW.pdf",
            "sheet": "Лист PDF stage 1: 46 / Лист PDF stage 2: 43",
            "page": 46,
            "severity": "medium",
            "source_layer": "image_enrichment",
            "change_direction": "unknown",
            "title": "Уточнены модели башенных кранов",
            "summary": "Кранам присвоены конкретные модели POTAIN.",
            "old_value": "Башенный кран №1–№4 без указания моделей",
            "new_value": "Кран 1 POTAIN MD 208; Кран 2 POTAIN MCT 385",
            "construction_impact": "Конкретизация моделей кранов.",
        },
        {
            "pair_id": "pf06effb7",
            "pair_label": "OLD.pdf ↔ NEW.pdf",
            "sheet": "Лист 3",
            "page": [3, 4],
            "severity": "high",
            "source_layer": "text",
            "change_direction": "complication",
            "title": "Критическое изменение",
            "summary": "детали",
            "old_value": "a; b; c",
            "new_value": "x; y",
            "construction_impact": "влияние",
        },
        {
            "pair_id": "pOTHER",
            "pair_label": "X.pdf ↔ Y.pdf",
            "sheet": "Лист 1",
            "page": 1,
            "severity": "low",
            "source_layer": "table",
            "title": "Другая пара",
            "summary": "",
            "old_value": "",
            "new_value": "",
            "construction_impact": "",
        },
    ]


def _roundtrip(items):
    return _roundtrip_with_order(items, pair_order=None)


def _roundtrip_with_order(items, pair_order):
    wb = _build_grouped_comparison_workbook(items, pair_order=pair_order)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf).active


def test_grouped_export_has_collapsible_pair_sections():
    ws = _roundtrip(_sample_items())
    # «+» (toggle) над группой деталей — summaryBelow выключен.
    assert ws.sheet_properties.outlinePr.summaryBelow is False

    # Заголовки пар (level 0) — свёрнуты по умолчанию (открываются плюсиком).
    pair_headers = [
        ri for ri in range(1, ws.max_row + 1)
        if ws.row_dimensions[ri].outline_level == 0
        and ws.row_dimensions[ri].collapsed
    ]
    # Две пары с принятыми изменениями → два сворачиваемых раздела.
    assert len(pair_headers) == 2
    labels = {ws.cell(row=ri, column=1).value for ri in pair_headers}
    assert any("OLD.pdf ↔ NEW.pdf" in (lbl or "") for lbl in labels)
    assert any("X.pdf ↔ Y.pdf" in (lbl or "") for lbl in labels)
    assert any("согласовано: 2" in (lbl or "") for lbl in labels)


def test_grouped_export_section_order_follows_pair_order():
    # pair_order повторяет порядок пар сессии (как вкладка «Отчёт»),
    # а не алфавит pair_label. Здесь pOTHER должна идти ПЕРВОЙ.
    items = _sample_items()
    ws = _roundtrip_with_order(items, pair_order=["pOTHER", "pf06effb7"])
    pair_headers = [
        ws.cell(row=ri, column=1).value
        for ri in range(1, ws.max_row + 1)
        if ws.row_dimensions[ri].outline_level == 0
        and ws.row_dimensions[ri].collapsed
    ]
    assert pair_headers[0].startswith("X.pdf ↔ Y.pdf")
    assert pair_headers[1].startswith("OLD.pdf ↔ NEW.pdf")


def test_grouped_export_pairs_outside_order_appended_by_appearance():
    items = _sample_items()
    # pair_order знает только про вторую пару — первая (по items) идёт следом.
    ws = _roundtrip_with_order(items, pair_order=["pOTHER"])
    pair_headers = [
        ws.cell(row=ri, column=1).value
        for ri in range(1, ws.max_row + 1)
        if ws.row_dimensions[ri].outline_level == 0
        and ws.row_dimensions[ri].collapsed
    ]
    assert pair_headers[0].startswith("X.pdf ↔ Y.pdf")
    assert pair_headers[1].startswith("OLD.pdf ↔ NEW.pdf")


def test_grouped_export_detail_rows_are_grouped_and_hidden():
    ws = _roundtrip(_sample_items())
    detail = [ri for ri in range(1, ws.max_row + 1) if ws.row_dimensions[ri].outline_level == 1]
    assert detail, "должны быть строки деталей на уровне группировки 1"
    # Все детали свёрнуты (раздел открывается плюсиком).
    assert all(ws.row_dimensions[ri].hidden for ri in detail)


def test_grouped_export_columns_match_ui():
    ws = _roundtrip(_sample_items())
    # Под-заголовок таблицы (первая level-1 строка) повторяет UI-колонки.
    subhdr_row = next(ri for ri in range(1, ws.max_row + 1) if ws.row_dimensions[ri].outline_level == 1)
    headers = [ws.cell(row=subhdr_row, column=c).value for c in range(1, 7)]
    assert headers == ["№", "Место", "Изменение", "Было", "Стало", "Влияние"]


def test_grouped_export_sorts_by_severity_within_pair():
    ws = _roundtrip(_sample_items())
    # В паре OLD↔NEW high ("Критическое") должен идти раньше medium ("Уточнены...").
    change_cells = [
        ws.cell(row=ri, column=3).value
        for ri in range(1, ws.max_row + 1)
        if ws.row_dimensions[ri].outline_level == 1
        and (ws.cell(row=ri, column=3).value or "") not in ("Изменение",)
    ]
    joined = "\n".join(c for c in change_cells if c)
    assert joined.index("Критическое изменение") < joined.index("Уточнены модели")


def test_grouped_export_number_cell_carries_severity_and_source():
    ws = _roundtrip(_sample_items())
    no_cells = [
        ws.cell(row=ri, column=1).value
        for ri in range(1, ws.max_row + 1)
        if ws.row_dimensions[ri].outline_level == 1
        and str(ws.cell(row=ri, column=1).value or "").startswith("№")
    ]
    text = "\n".join(c for c in no_cells if c)
    # severity + источник (схлопнут до текст/изображение) + направление.
    assert "medium" in text and "изображение" in text
    assert "high" in text and "усложнение" in text


def test_grouped_export_empty_items_produces_valid_workbook():
    ws = _roundtrip([])
    assert ws.max_row >= 1  # пустой лист валиден


def test_sc_split_lines_mirrors_frontend():
    assert _sc_split_lines("a; b; c") == "a;\nb;\nc"
    assert _sc_split_lines("single") == "single"
    assert _sc_split_lines("") == ""
    assert _sc_split_lines(None) == ""


def test_sc_page_str_handles_list_and_scalar():
    assert _sc_page_str([3, 4]) == "3, 4"
    assert _sc_page_str(46) == "46"
    assert _sc_page_str(None) == ""
