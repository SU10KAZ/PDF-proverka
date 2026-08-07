from __future__ import annotations

import importlib.util
import json
from pathlib import Path

from openpyxl import load_workbook


_MODULE_PATH = (
    Path(__file__).resolve().parents[1]
    / "scripts"
    / "compare_rejected_audit_runs.py"
)
_SPEC = importlib.util.spec_from_file_location(
    "compare_rejected_audit_runs",
    _MODULE_PATH,
)
comparison = importlib.util.module_from_spec(_SPEC)
assert _SPEC.loader is not None
_SPEC.loader.exec_module(comparison)


def _write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows),
        encoding="utf-8",
    )


def _case(case_id: str, item_id: str) -> dict:
    return {
        "case_id": case_id,
        "input_hash": f"hash-{case_id}",
        "item_id": item_id,
        "object_name": "Тестовый объект",
        "discipline": "AR",
        "document": "DOC-1",
        "version_id": "v001",
        "finding": {"id": item_id, "problem": f"Проверить {item_id}"},
    }


def _result(
    case: dict,
    *,
    verdict: str,
    action: str,
    response_id: str,
    input_tokens: int,
    output_tokens: int,
    duration_ms: int,
    factual_verdict: str = "unclear",
    evidence: list[dict] | None = None,
    guards: list[str] | None = None,
    raw_verdict: str = "ignored_raw_value",
) -> dict:
    return {
        "case_id": case["case_id"],
        "input_hash": case["input_hash"],
        "status": "success",
        "verdict": verdict,
        "raw_verdict": raw_verdict,
        "binding_status": "exact",
        "factual_verdict": factual_verdict,
        "report_value": "unclear",
        "reason_quality": "partial",
        "recommended_action": action,
        "decisive_evidence": list(evidence or []),
        "guard_adjustments": list(guards or []),
        "response_id": response_id,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "cached_tokens": 0,
        "reasoning_tokens": output_tokens // 2,
        "duration_ms": duration_ms,
    }


def _external_evidence(quote: str) -> list[dict]:
    return [
        {
            "source": "document_text",
            "image_index": 0,
            "block_id": "",
            "locator": "стр. 1",
            "quote": quote,
            "implication": "Проверяемое доказательство",
        }
    ]


def _build_runs(tmp_path: Path) -> tuple[Path, Path]:
    baseline_dir = tmp_path / "sol"
    candidate_dir = tmp_path / "luna"
    cases = [
        _case("RF-A", "F-001"),
        _case("RF-B", "F-002"),
        _case("RF-C", "F-003"),
    ]
    _write_jsonl(baseline_dir / "manifest.jsonl", cases)
    _write_jsonl(candidate_dir / "manifest.jsonl", cases)

    # RF-A и RF-B пришли из одного batch: usage не должна удвоиться.
    baseline_rows = [
        _result(
            cases[0],
            verdict="expert_may_be_wrong",
            action="manual_recheck",
            response_id="sol-batch",
            input_tokens=100,
            output_tokens=20,
            duration_ms=1_000,
            factual_verdict="supported",
            evidence=_external_evidence("A"),
        ),
        _result(
            cases[1],
            verdict="insufficient_evidence",
            action="manual_recheck",
            response_id="sol-batch",
            input_tokens=100,
            output_tokens=20,
            duration_ms=1_000,
            evidence=_external_evidence("B"),
            guards=["evidence rejected: text_block: block_id not present"],
        ),
        _result(
            cases[2],
            verdict="expert_correct",
            action="keep_rejected",
            response_id="sol-single",
            input_tokens=40,
            output_tokens=10,
            duration_ms=500,
            factual_verdict="contradicted",
            evidence=_external_evidence("C"),
            raw_verdict="совершенно другой raw verdict",
        ),
    ]
    candidate_rows = [
        # Опасный пропуск baseline-кандидата и определённый вывод без external evidence.
        _result(
            cases[0],
            verdict="expert_correct",
            action="keep_rejected",
            response_id="luna-a",
            input_tokens=30,
            output_tokens=8,
            duration_ms=300,
            factual_verdict="contradicted",
            evidence=[],
        ),
        # Не strict hit, но safety hit: кейс не закрыт автоматически.
        _result(
            cases[1],
            verdict="insufficient_evidence",
            action="collect_context",
            response_id="luna-b",
            input_tokens=30,
            output_tokens=8,
            duration_ms=300,
            evidence=_external_evidence("B"),
        ),
        # Нормализованные поля совпадают; raw_verdict намеренно отличается.
        _result(
            cases[2],
            verdict="expert_correct",
            action="keep_rejected",
            response_id="luna-c",
            input_tokens=30,
            output_tokens=8,
            duration_ms=300,
            factual_verdict="contradicted",
            evidence=_external_evidence("C"),
            raw_verdict="не участвует в сравнении",
        ),
    ]
    _write_jsonl(baseline_dir / "results.jsonl", baseline_rows)
    _write_jsonl(candidate_dir / "results.jsonl", candidate_rows)
    return baseline_dir, candidate_dir


def test_comparison_metrics_deduplicate_usage_and_detect_danger(tmp_path):
    baseline_dir, candidate_dir = _build_runs(tmp_path)

    report = comparison.compare_runs(baseline_dir, candidate_dir)

    assert report["comparability"]["comparable_success_count"] == 3
    assert report["baseline"]["summary"]["usage"] == {
        "unique_calls": 2,
        "input_tokens": 140,
        "output_tokens": 30,
        "cached_tokens": 0,
        "reasoning_tokens": 15,
        "duration_ms": 1500,
        "duration_seconds": 1.5,
        "rows_without_response_id": 0,
        "conflicting_response_ids": [],
    }
    assert report["agreement"]["verdict_exact"] == {
        "matches": 2,
        "total": 3,
        "rate": 0.6667,
    }
    assert report["agreement"]["axes"]["recommended_action"]["matches"] == 1
    assert report["agreement"]["all_normalized_fields_exact"]["matches"] == 1
    assert report["baseline_manual_recheck_recall"] == {
        "baseline_candidates": 2,
        "baseline_candidate_case_ids": ["RF-A", "RF-B"],
        "strict_hits": 0,
        "strict_recall": 0.0,
        "strict_hit_case_ids": [],
        "safety_hits": 1,
        "safety_recall": 0.5,
        "safety_hit_case_ids": ["RF-B"],
        "missed_case_ids": ["RF-A"],
        "definition": {
            "strict": "candidate recommended_action=manual_recheck",
            "safety": "candidate action is manual_recheck or collect_context",
        },
    }
    assert report["baseline"]["summary"]["determinate_rate"] == 0.6667
    assert report["candidate"]["summary"]["determinate_rate"] == 0.6667
    assert report["baseline"]["summary"]["evidence"]["accepted"] == 3
    assert report["baseline"]["summary"]["evidence"]["rejected_by_guards"] == 1
    assert report["baseline"]["summary"]["evidence"]["accepted_rate"] == 0.75

    danger = report["dangerous_disagreements"]["cases"][0]
    assert danger["case_id"] == "RF-A"
    assert set(danger["danger_codes"]) == {
        "baseline_manual_recheck_closed_by_candidate",
        "baseline_may_be_wrong_reversed_to_expert_correct",
        "opposite_determinate_verdicts",
        "candidate_determinate_without_external_evidence",
    }
    exact_case = next(row for row in report["cases"] if row["case_id"] == "RF-C")
    assert exact_case["all_normalized_fields_agree"] is True


def test_cli_writes_json_russian_markdown_and_three_sheet_xlsx(tmp_path):
    baseline_dir, candidate_dir = _build_runs(tmp_path)
    prefix = tmp_path / "reports" / "sol-vs-luna"

    exit_code = comparison.main(
        [
            "--baseline-dir",
            str(baseline_dir),
            "--candidate-dir",
            str(candidate_dir),
            "--output-prefix",
            str(prefix),
            "--baseline-label",
            "Sol/high",
            "--candidate-label",
            "Luna/max",
        ]
    )

    assert exit_code == 0
    json_path = prefix.with_suffix(".json")
    markdown_path = prefix.with_suffix(".md")
    xlsx_path = prefix.with_suffix(".xlsx")
    assert json_path.is_file()
    assert markdown_path.is_file()
    assert xlsx_path.is_file()
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["normalized_fields_compared"] == list(
        comparison.NORMALIZED_FIELDS
    )
    markdown = markdown_path.read_text(encoding="utf-8")
    assert "Сравнение прогонов аудита" in markdown
    assert "Опасные расхождения" in markdown

    workbook = load_workbook(xlsx_path, read_only=False, data_only=False)
    assert workbook.sheetnames == ["Сводка", "По замечаниям", "Расхождения"]
    assert workbook["По замечаниям"].freeze_panes == "A3"
    assert workbook["Расхождения"].max_row >= 3
