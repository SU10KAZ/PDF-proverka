#!/usr/bin/env python3
"""Сравнить два read-only прогона аудита отклонённых замечаний.

Скрипт не запускает модели и не изменяет каталоги аудита. Он читает
``manifest.jsonl`` и append-only ``results.jsonl`` двух прогонов, выбирает
последний результат с актуальным ``input_hash`` и пишет три отчёта:

    python scripts/compare_rejected_audit_runs.py \
      --baseline-dir comparison/.../sol-high/retrieval-pilot-10-v3 \
      --candidate-dir comparison/.../luna-max/retrieval-pilot-10-v1 \
      --output-prefix comparison/.../sol-vs-luna

Результат: ``sol-vs-luna.json``, ``sol-vs-luna.md`` и ``sol-vs-luna.xlsx``.
Семантическое сравнение выполняется только по нормализованным полям. Поля
``raw_verdict`` и свободный текст заключений намеренно не участвуют.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NORMALIZED_FIELDS = (
    "verdict",
    "binding_status",
    "factual_verdict",
    "report_value",
    "reason_quality",
    "recommended_action",
)
DETERMINATE_VERDICTS = {"expert_correct", "expert_may_be_wrong"}
EXTERNAL_EVIDENCE_SOURCES = {
    "graphic_block",
    "text_block",
    "document_text",
    "norm_context",
}
USAGE_FIELDS = (
    "input_tokens",
    "output_tokens",
    "cached_tokens",
    "reasoning_tokens",
    "duration_ms",
)

FIELD_RU = {
    "verdict": "Вердикт",
    "binding_status": "Привязка причины",
    "factual_verdict": "Фактическая оценка",
    "report_value": "Ценность замечания",
    "reason_quality": "Качество причины",
    "recommended_action": "Рекомендуемое действие",
}
VALUE_RU = {
    "expert_correct": "Эксперт прав",
    "expert_may_be_wrong": "Эксперт мог ошибиться",
    "insufficient_evidence": "Недостаточно данных",
    "exact": "Точная",
    "conflict": "Конфликт",
    "missing": "Отсутствует",
    "supported": "Подтверждён",
    "unsupported": "Не подтверждён",
    "contradicted": "Опровергнут",
    "unclear": "Неясно",
    "include": "Оставить",
    "merge": "Объединить",
    "downgrade": "Снизить значимость",
    "remove": "Убрать",
    "substantiated": "Обоснована",
    "partial": "Частично обоснована",
    "unsubstantiated": "Не обоснована",
    "keep_rejected": "Оставить отклонённым",
    "manual_recheck": "Ручная перепроверка",
    "collect_context": "Собрать контекст",
    "success": "Успешно",
    "error": "Ошибка",
    "pending": "Нет результата",
}
DANGER_RU = {
    "baseline_manual_recheck_closed_by_candidate": (
        "Кандидат закрыл как отклонённое замечание из baseline manual_recheck"
    ),
    "baseline_may_be_wrong_reversed_to_expert_correct": (
        "Вердикт «эксперт мог ошибиться» заменён на «эксперт прав»"
    ),
    "opposite_determinate_verdicts": "Противоположные определённые вердикты",
    "candidate_determinate_without_external_evidence": (
        "Определённый вердикт кандидата без принятого внешнего доказательства"
    ),
    "baseline_manual_recheck_not_comparable": (
        "Кандидат baseline manual_recheck не получил сопоставимого результата"
    ),
}

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="305496")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
DANGER_FILL = PatternFill("solid", fgColor="FCE4D6")
DIFF_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E1F2")
CELL_BORDER = Border(bottom=THIN_GRAY)
MAX_EXCEL_TEXT = 32_760


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_jsonl(path: Path, *, tolerate_malformed: bool) -> tuple[list[dict], int]:
    rows: list[dict] = []
    malformed = 0
    if not path.is_file():
        raise FileNotFoundError(f"Не найден обязательный файл: {path}")
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line_number, raw in enumerate(handle, start=1):
            line = raw.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError as exc:
                if tolerate_malformed:
                    malformed += 1
                    continue
                raise ValueError(
                    f"Некорректный JSONL {path}:{line_number}: {exc}"
                ) from exc
            if not isinstance(item, dict):
                if tolerate_malformed:
                    malformed += 1
                    continue
                raise ValueError(f"Ожидался JSON-объект {path}:{line_number}")
            item = dict(item)
            item["__line_number"] = line_number
            rows.append(item)
    return rows, malformed


def _manifest_index(rows: Sequence[dict], path: Path) -> tuple[dict[str, dict], list[str]]:
    index: dict[str, dict] = {}
    order: list[str] = []
    for row in rows:
        case_id = str(row.get("case_id") or "").strip()
        if not case_id:
            raise ValueError(
                f"В manifest отсутствует case_id: {path}:{row.get('__line_number')}"
            )
        if case_id in index:
            raise ValueError(f"Повторяющийся case_id в manifest {path}: {case_id}")
        clean = {key: value for key, value in row.items() if key != "__line_number"}
        index[case_id] = clean
        order.append(case_id)
    return index, order


def _current_results(
    result_rows: Sequence[dict],
    manifest: dict[str, dict],
) -> tuple[dict[str, dict], list[str], list[str]]:
    latest: dict[str, dict] = {}
    for row in result_rows:
        case_id = str(row.get("case_id") or "").strip()
        if case_id:
            latest[case_id] = {
                key: value for key, value in row.items() if key != "__line_number"
            }
    stale: list[str] = []
    current: dict[str, dict] = {}
    for case_id, case in manifest.items():
        row = latest.get(case_id)
        if row is None:
            continue
        if str(row.get("input_hash") or "") != str(case.get("input_hash") or ""):
            stale.append(case_id)
            continue
        current[case_id] = row
    outside_manifest = sorted(case_id for case_id in latest if case_id not in manifest)
    return current, sorted(stale), outside_manifest


def _number(value: Any) -> int:
    try:
        return max(0, int(value or 0))
    except (TypeError, ValueError):
        return 0


def aggregate_usage(result_rows: Sequence[dict]) -> dict[str, Any]:
    """Суммировать usage один раз на response_id.

    В batch-ответе одна и та же usage записывается для каждого кейса. Если
    response_id отсутствует, строка считается отдельным вызовом. При
    противоречивых счётчиках одного response_id берётся максимум и конфликт
    отражается в диагностике.
    """

    grouped: dict[str, list[dict]] = defaultdict(list)
    missing_response_id = 0
    for ordinal, row in enumerate(result_rows, start=1):
        response_id = str(row.get("response_id") or "").strip()
        if not response_id:
            missing_response_id += 1
            response_id = f"__without_response_id_{ordinal}"
        grouped[response_id].append(row)

    totals = {field: 0 for field in USAGE_FIELDS}
    conflicting_ids: list[str] = []
    for response_id, rows in grouped.items():
        conflict = False
        for field in USAGE_FIELDS:
            values = {_number(row.get(field)) for row in rows}
            if len(values) > 1:
                conflict = True
            totals[field] += max(values, default=0)
        if conflict and not response_id.startswith("__without_response_id_"):
            conflicting_ids.append(response_id)

    return {
        "unique_calls": len(grouped),
        **totals,
        "duration_seconds": round(totals["duration_ms"] / 1000, 3),
        "rows_without_response_id": missing_response_id,
        "conflicting_response_ids": sorted(conflicting_ids),
    }


def _evidence_stats(rows: Iterable[dict]) -> dict[str, Any]:
    accepted = 0
    external = 0
    rejected = 0
    guard_cases: list[str] = []
    determinate_without_external: list[str] = []
    per_case: dict[str, dict[str, int]] = {}
    for row in rows:
        case_id = str(row.get("case_id") or "")
        evidence = row.get("decisive_evidence")
        evidence = evidence if isinstance(evidence, list) else []
        accepted_case = sum(1 for item in evidence if isinstance(item, dict))
        external_case = sum(
            1
            for item in evidence
            if isinstance(item, dict)
            and str(item.get("source") or "") in EXTERNAL_EVIDENCE_SOURCES
        )
        adjustments = row.get("guard_adjustments")
        adjustments = adjustments if isinstance(adjustments, list) else []
        rejected_case = sum(
            1
            for adjustment in adjustments
            if str(adjustment).startswith("evidence rejected:")
        )
        accepted += accepted_case
        external += external_case
        rejected += rejected_case
        if rejected_case:
            guard_cases.append(case_id)
        if (
            str(row.get("verdict") or "") in DETERMINATE_VERDICTS
            and external_case == 0
        ):
            determinate_without_external.append(case_id)
        per_case[case_id] = {
            "accepted": accepted_case,
            "accepted_external": external_case,
            "rejected_by_guards": rejected_case,
        }
    attempted = accepted + rejected
    return {
        "accepted": accepted,
        "accepted_external": external,
        "rejected_by_guards": rejected,
        "attempted": attempted,
        "accepted_rate": round(accepted / attempted, 4) if attempted else None,
        "guard_affected_cases": sorted(guard_cases),
        "determinate_without_external_evidence": sorted(
            determinate_without_external
        ),
        "per_case": per_case,
    }


def load_run(directory: Path) -> dict[str, Any]:
    directory = Path(directory).resolve()
    manifest_path = directory / "manifest.jsonl"
    results_path = directory / "results.jsonl"
    manifest_rows, _ = _read_jsonl(manifest_path, tolerate_malformed=False)
    manifest, manifest_order = _manifest_index(manifest_rows, manifest_path)
    result_rows, malformed_results = _read_jsonl(
        results_path, tolerate_malformed=True
    )
    current, stale_ids, outside_ids = _current_results(result_rows, manifest)
    successful = {
        case_id: row
        for case_id, row in current.items()
        if row.get("status") == "success"
    }
    errors = sorted(
        case_id for case_id, row in current.items() if row.get("status") == "error"
    )
    pending = sorted(case_id for case_id in manifest if case_id not in successful)
    determinate = sum(
        1
        for row in successful.values()
        if row.get("verdict") in DETERMINATE_VERDICTS
    )
    manual = sorted(
        case_id
        for case_id, row in successful.items()
        if row.get("recommended_action") == "manual_recheck"
    )
    evidence = _evidence_stats(successful.values())
    return {
        "directory": str(directory),
        "manifest_path": str(manifest_path),
        "results_path": str(results_path),
        "manifest_sha256": _sha256(manifest_path),
        "manifest": manifest,
        "manifest_order": manifest_order,
        "all_result_rows": result_rows,
        "current_results": current,
        "successful_results": successful,
        "summary": {
            "selected_cases": len(manifest),
            "current_results": len(current),
            "successful_cases": len(successful),
            "errors": len(errors),
            "error_case_ids": errors,
            "pending_cases": len(pending),
            "pending_case_ids": pending,
            "stale_results": len(stale_ids),
            "stale_case_ids": stale_ids,
            "results_outside_manifest": outside_ids,
            "malformed_result_lines": malformed_results,
            "determinate_cases": determinate,
            "determinate_rate": (
                round(determinate / len(successful), 4) if successful else None
            ),
            "manual_recheck_candidates": len(manual),
            "manual_recheck_case_ids": manual,
            "verdicts": dict(
                Counter(str(row.get("verdict") or "") for row in successful.values())
            ),
            "evidence": evidence,
            "usage": aggregate_usage(result_rows),
        },
    }


def _case_metadata(case_id: str, baseline: dict, candidate: dict) -> dict[str, str]:
    case = baseline.get("manifest", {}).get(case_id) or candidate.get("manifest", {}).get(case_id) or {}
    finding = case.get("finding") if isinstance(case.get("finding"), dict) else {}
    return {
        "case_id": case_id,
        "item_id": str(case.get("item_id") or finding.get("id") or ""),
        "object_name": str(case.get("object_name") or ""),
        "discipline": str(case.get("discipline") or ""),
        "document": str(case.get("document") or ""),
        "version_id": str(case.get("version_id") or ""),
        "finding_problem": str(
            finding.get("problem") or finding.get("description") or ""
        ),
    }


def _success_status(row: dict | None) -> str:
    return str((row or {}).get("status") or "pending")


def _ratio(count: int, total: int) -> float | None:
    return round(count / total, 4) if total else None


def compare_runs(
    baseline_dir: Path,
    candidate_dir: Path,
    *,
    baseline_label: str = "Sol baseline",
    candidate_label: str = "Luna candidate",
) -> dict[str, Any]:
    baseline = load_run(Path(baseline_dir))
    candidate = load_run(Path(candidate_dir))

    baseline_ids = set(baseline["manifest"])
    candidate_ids = set(candidate["manifest"])
    paired_ids = baseline_ids & candidate_ids
    union_order = list(baseline["manifest_order"])
    union_order.extend(
        case_id
        for case_id in candidate["manifest_order"]
        if case_id not in baseline_ids
    )

    input_hash_mismatches = sorted(
        case_id
        for case_id in paired_ids
        if str(baseline["manifest"][case_id].get("input_hash") or "")
        != str(candidate["manifest"][case_id].get("input_hash") or "")
    )
    comparable_ids = [
        case_id
        for case_id in union_order
        if case_id in paired_ids
        and case_id not in input_hash_mismatches
        and _success_status(baseline["current_results"].get(case_id)) == "success"
        and _success_status(candidate["current_results"].get(case_id)) == "success"
    ]
    comparable_set = set(comparable_ids)

    axis_matches = {field: 0 for field in NORMALIZED_FIELDS}
    rows: list[dict[str, Any]] = []
    dangerous: list[dict[str, Any]] = []
    all_axes_matches = 0

    baseline_evidence = baseline["summary"]["evidence"]["per_case"]
    candidate_evidence = candidate["summary"]["evidence"]["per_case"]

    for case_id in union_order:
        b_row = baseline["current_results"].get(case_id)
        c_row = candidate["current_results"].get(case_id)
        comparable = case_id in comparable_set
        field_agreement: dict[str, bool | None] = {}
        different_fields: list[str] = []
        for field in NORMALIZED_FIELDS:
            agrees = (
                str((b_row or {}).get(field) or "")
                == str((c_row or {}).get(field) or "")
                if comparable
                else None
            )
            field_agreement[field] = agrees
            if agrees is True:
                axis_matches[field] += 1
            elif agrees is False:
                different_fields.append(field)
        all_axes = comparable and all(value is True for value in field_agreement.values())
        if all_axes:
            all_axes_matches += 1

        danger_codes: list[str] = []
        if comparable:
            b_action = str((b_row or {}).get("recommended_action") or "")
            c_action = str((c_row or {}).get("recommended_action") or "")
            b_verdict = str((b_row or {}).get("verdict") or "")
            c_verdict = str((c_row or {}).get("verdict") or "")
            if b_action == "manual_recheck" and c_action == "keep_rejected":
                danger_codes.append("baseline_manual_recheck_closed_by_candidate")
            if b_verdict == "expert_may_be_wrong" and c_verdict == "expert_correct":
                danger_codes.append(
                    "baseline_may_be_wrong_reversed_to_expert_correct"
                )
            if {b_verdict, c_verdict} == {
                "expert_correct",
                "expert_may_be_wrong",
            }:
                danger_codes.append("opposite_determinate_verdicts")
            if (
                c_verdict in DETERMINATE_VERDICTS
                and candidate_evidence.get(case_id, {}).get("accepted_external", 0) == 0
            ):
                danger_codes.append(
                    "candidate_determinate_without_external_evidence"
                )
        elif (
            str((b_row or {}).get("status") or "") == "success"
            and str((b_row or {}).get("recommended_action") or "")
            == "manual_recheck"
        ):
            danger_codes.append("baseline_manual_recheck_not_comparable")

        metadata = _case_metadata(case_id, baseline, candidate)
        row = {
            **metadata,
            "in_baseline_manifest": case_id in baseline_ids,
            "in_candidate_manifest": case_id in candidate_ids,
            "input_hash_match": (
                case_id in paired_ids and case_id not in input_hash_mismatches
            ),
            "comparable": comparable,
            "baseline_status": _success_status(b_row),
            "candidate_status": _success_status(c_row),
            "baseline": {
                field: str((b_row or {}).get(field) or "")
                for field in NORMALIZED_FIELDS
            },
            "candidate": {
                field: str((c_row or {}).get(field) or "")
                for field in NORMALIZED_FIELDS
            },
            "agreement": field_agreement,
            "all_normalized_fields_agree": all_axes if comparable else None,
            "different_fields": different_fields,
            "baseline_evidence": baseline_evidence.get(
                case_id,
                {"accepted": 0, "accepted_external": 0, "rejected_by_guards": 0},
            ),
            "candidate_evidence": candidate_evidence.get(
                case_id,
                {"accepted": 0, "accepted_external": 0, "rejected_by_guards": 0},
            ),
            "danger_codes": danger_codes,
        }
        rows.append(row)
        if danger_codes:
            dangerous.append(
                {
                    "case_id": case_id,
                    "item_id": metadata["item_id"],
                    "danger_codes": danger_codes,
                    "different_fields": different_fields,
                    "baseline": row["baseline"],
                    "candidate": row["candidate"],
                }
            )

    comparable_total = len(comparable_ids)
    axis_agreement = {
        field: {
            "matches": axis_matches[field],
            "total": comparable_total,
            "rate": _ratio(axis_matches[field], comparable_total),
        }
        for field in NORMALIZED_FIELDS
    }

    baseline_manual = set(
        baseline["summary"]["manual_recheck_case_ids"]
    )
    strict_hits: list[str] = []
    safety_hits: list[str] = []
    missed: list[str] = []
    for case_id in sorted(baseline_manual):
        c_row = candidate["current_results"].get(case_id)
        eligible = (
            case_id in comparable_set
            and str((c_row or {}).get("status") or "") == "success"
        )
        action = str((c_row or {}).get("recommended_action") or "")
        if eligible and action == "manual_recheck":
            strict_hits.append(case_id)
        if eligible and action in {"manual_recheck", "collect_context"}:
            safety_hits.append(case_id)
        else:
            missed.append(case_id)

    baseline_public = {
        key: value
        for key, value in baseline.items()
        if key not in {
            "manifest",
            "manifest_order",
            "all_result_rows",
            "current_results",
            "successful_results",
        }
    }
    candidate_public = {
        key: value
        for key, value in candidate.items()
        if key not in {
            "manifest",
            "manifest_order",
            "all_result_rows",
            "current_results",
            "successful_results",
        }
    }
    return {
        "schema_version": 1,
        "generated_at": _now_iso(),
        "labels": {
            "baseline": baseline_label,
            "candidate": candidate_label,
        },
        "normalized_fields_compared": list(NORMALIZED_FIELDS),
        "baseline": baseline_public,
        "candidate": candidate_public,
        "comparability": {
            "baseline_case_count": len(baseline_ids),
            "candidate_case_count": len(candidate_ids),
            "paired_case_count": len(paired_ids),
            "comparable_success_count": comparable_total,
            "same_case_set": baseline_ids == candidate_ids,
            "same_manifest_sha256": (
                baseline["manifest_sha256"] == candidate["manifest_sha256"]
            ),
            "matching_input_hash_count": len(paired_ids) - len(input_hash_mismatches),
            "input_hash_mismatch_case_ids": input_hash_mismatches,
            "baseline_only_case_ids": sorted(baseline_ids - candidate_ids),
            "candidate_only_case_ids": sorted(candidate_ids - baseline_ids),
            "comparable_case_ids": comparable_ids,
        },
        "agreement": {
            "verdict_exact": axis_agreement["verdict"],
            "all_normalized_fields_exact": {
                "matches": all_axes_matches,
                "total": comparable_total,
                "rate": _ratio(all_axes_matches, comparable_total),
            },
            "axes": axis_agreement,
        },
        "baseline_manual_recheck_recall": {
            "baseline_candidates": len(baseline_manual),
            "baseline_candidate_case_ids": sorted(baseline_manual),
            "strict_hits": len(strict_hits),
            "strict_recall": _ratio(len(strict_hits), len(baseline_manual)),
            "strict_hit_case_ids": strict_hits,
            "safety_hits": len(safety_hits),
            "safety_recall": _ratio(len(safety_hits), len(baseline_manual)),
            "safety_hit_case_ids": safety_hits,
            "missed_case_ids": missed,
            "definition": {
                "strict": "candidate recommended_action=manual_recheck",
                "safety": "candidate action is manual_recheck or collect_context",
            },
        },
        "dangerous_disagreements": {
            "count": len(dangerous),
            "cases": dangerous,
        },
        "cases": rows,
    }


def _pct(value: Any) -> str:
    if value is None:
        return "—"
    return f"{100 * float(value):.1f}%"


def _human(value: Any) -> str:
    if value in (None, ""):
        return "—"
    return VALUE_RU.get(str(value), str(value))


def _md(value: Any) -> str:
    return str(value if value not in (None, "") else "—").replace("|", "\\|").replace("\n", " ")


def render_markdown(report: dict[str, Any]) -> str:
    labels = report["labels"]
    b = report["baseline"]["summary"]
    c = report["candidate"]["summary"]
    comp = report["comparability"]
    agreement = report["agreement"]
    recall = report["baseline_manual_recheck_recall"]
    dangers = report["dangerous_disagreements"]

    lines = [
        "# Сравнение прогонов аудита отклонённых замечаний",
        "",
        f"Сформировано: `{report['generated_at']}`",
        "",
        f"- Базовый прогон: **{_md(labels['baseline'])}** — `{_md(report['baseline']['directory'])}`",
        f"- Кандидат: **{_md(labels['candidate'])}** — `{_md(report['candidate']['directory'])}`",
        f"- Сопоставимых успешных кейсов: **{comp['comparable_success_count']}** из {comp['paired_case_count']}",
        f"- Совпадают `input_hash`: **{comp['matching_input_hash_count']}** из {comp['paired_case_count']}",
        "",
        "## Сводка",
        "",
        f"| Метрика | {_md(labels['baseline'])} | {_md(labels['candidate'])} |",
        "|---|---:|---:|",
        f"| Выбрано кейсов | {b['selected_cases']} | {c['selected_cases']} |",
        f"| Успешно | {b['successful_cases']} | {c['successful_cases']} |",
        f"| Определённые вердикты | {b['determinate_cases']} ({_pct(b['determinate_rate'])}) | {c['determinate_cases']} ({_pct(c['determinate_rate'])}) |",
        f"| Кандидаты manual_recheck | {b['manual_recheck_candidates']} | {c['manual_recheck_candidates']} |",
        f"| Evidence принято | {b['evidence']['accepted']} | {c['evidence']['accepted']} |",
        f"| Evidence отклонено guards | {b['evidence']['rejected_by_guards']} | {c['evidence']['rejected_by_guards']} |",
        f"| Валидность evidence | {_pct(b['evidence']['accepted_rate'])} | {_pct(c['evidence']['accepted_rate'])} |",
        f"| Уникальные вызовы | {b['usage']['unique_calls']} | {c['usage']['unique_calls']} |",
        f"| Входные токены | {b['usage']['input_tokens']} | {c['usage']['input_tokens']} |",
        f"| Выходные токены | {b['usage']['output_tokens']} | {c['usage']['output_tokens']} |",
        f"| Reasoning-токены | {b['usage']['reasoning_tokens']} | {c['usage']['reasoning_tokens']} |",
        f"| Суммарное модельное время, с | {b['usage']['duration_seconds']} | {c['usage']['duration_seconds']} |",
        "",
        "## Совпадение нормализованных полей",
        "",
        "| Поле | Совпало | Всего | Доля |",
        "|---|---:|---:|---:|",
    ]
    for field in NORMALIZED_FIELDS:
        metric = agreement["axes"][field]
        lines.append(
            f"| {_md(FIELD_RU[field])} | {metric['matches']} | {metric['total']} | {_pct(metric['rate'])} |"
        )
    exact = agreement["all_normalized_fields_exact"]
    lines.extend(
        [
            f"| Все шесть полей одновременно | {exact['matches']} | {exact['total']} | {_pct(exact['rate'])} |",
            "",
            "## Сохранение кандидатов baseline",
            "",
            f"- Строгий recall (`manual_recheck`): **{recall['strict_hits']}/{recall['baseline_candidates']} ({_pct(recall['strict_recall'])})**.",
            f"- Safety recall (`manual_recheck` или `collect_context`): **{recall['safety_hits']}/{recall['baseline_candidates']} ({_pct(recall['safety_recall'])})**.",
            f"- Пропущенные case_id: `{', '.join(recall['missed_case_ids']) or 'нет'}`.",
            "",
            f"## Опасные расхождения — {dangers['count']}",
            "",
        ]
    )
    if dangers["cases"]:
        lines.extend(
            [
                "| Case ID | ID замечания | Причина | Базовый вердикт | Вердикт кандидата |",
                "|---|---|---|---|---|",
            ]
        )
        for item in dangers["cases"]:
            reasons = "; ".join(
                DANGER_RU.get(code, code) for code in item["danger_codes"]
            )
            lines.append(
                "| {case} | {item} | {reason} | {base} | {candidate} |".format(
                    case=_md(item["case_id"]),
                    item=_md(item["item_id"]),
                    reason=_md(reasons),
                    base=_md(_human(item["baseline"].get("verdict"))),
                    candidate=_md(_human(item["candidate"].get("verdict"))),
                )
            )
    else:
        lines.append("Опасных расхождений не обнаружено.")
    lines.extend(
        [
            "",
            "## Методика",
            "",
            "Сравниваются только нормализованные поля: `verdict`, `binding_status`, "
            "`factual_verdict`, `report_value`, `reason_quality`, `recommended_action`. "
            "Usage и время дедуплицируются по `response_id`. Опасное расхождение — "
            "это сигнал для ручной проверки, а не автоматическое признание одной модели правой.",
            "",
        ]
    )
    return "\n".join(lines)


def _excel_value(value: Any) -> Any:
    if value is None or value == "":
        return "—"
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if isinstance(value, (int, float)):
        return value
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value)).strip()
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text[:MAX_EXCEL_TEXT] or "—"


def _style_title(ws, title: str, columns: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = ws.cell(1, 1, title)
    cell.fill = TITLE_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def _style_header(ws, row_number: int) -> None:
    for cell in ws[row_number]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_table(ws, name: str, header_row: int, last_row: int, last_col: int) -> None:
    if last_row <= header_row:
        return
    ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _write_summary_sheet(ws, report: dict[str, Any]) -> None:
    labels = report["labels"]
    b = report["baseline"]["summary"]
    c = report["candidate"]["summary"]
    _style_title(ws, "Сравнение прогонов аудита", 4)
    ws.append(["Метрика", labels["baseline"], labels["candidate"], "Сравнение"])
    _style_header(ws, 2)
    rows = [
        ("Каталог", report["baseline"]["directory"], report["candidate"]["directory"], ""),
        ("SHA-256 manifest", report["baseline"]["manifest_sha256"], report["candidate"]["manifest_sha256"], "Совпадает" if report["comparability"]["same_manifest_sha256"] else "Различается"),
        ("Выбрано кейсов", b["selected_cases"], c["selected_cases"], ""),
        ("Успешно", b["successful_cases"], c["successful_cases"], ""),
        ("Определённые вердикты", b["determinate_cases"], c["determinate_cases"], ""),
        ("Доля определённых", _pct(b["determinate_rate"]), _pct(c["determinate_rate"]), ""),
        ("Кандидаты manual_recheck", b["manual_recheck_candidates"], c["manual_recheck_candidates"], ""),
        ("Evidence принято", b["evidence"]["accepted"], c["evidence"]["accepted"], ""),
        ("Evidence отклонено guards", b["evidence"]["rejected_by_guards"], c["evidence"]["rejected_by_guards"], ""),
        ("Валидность evidence", _pct(b["evidence"]["accepted_rate"]), _pct(c["evidence"]["accepted_rate"]), ""),
        ("Уникальные вызовы", b["usage"]["unique_calls"], c["usage"]["unique_calls"], ""),
        ("Входные токены", b["usage"]["input_tokens"], c["usage"]["input_tokens"], ""),
        ("Выходные токены", b["usage"]["output_tokens"], c["usage"]["output_tokens"], ""),
        ("Reasoning-токены", b["usage"]["reasoning_tokens"], c["usage"]["reasoning_tokens"], ""),
        ("Суммарное модельное время, с", b["usage"]["duration_seconds"], c["usage"]["duration_seconds"], ""),
        ("Сопоставимые успешные кейсы", "", "", report["comparability"]["comparable_success_count"]),
        ("Точное совпадение verdict", "", "", _pct(report["agreement"]["verdict_exact"]["rate"])),
        ("Совпали все шесть полей", "", "", _pct(report["agreement"]["all_normalized_fields_exact"]["rate"])),
        ("Строгий recall кандидатов", "", "", _pct(report["baseline_manual_recheck_recall"]["strict_recall"])),
        ("Safety recall кандидатов", "", "", _pct(report["baseline_manual_recheck_recall"]["safety_recall"])),
        ("Опасные расхождения", "", "", report["dangerous_disagreements"]["count"]),
    ]
    for row in rows:
        ws.append([_excel_value(value) for value in row])

    axis_start = ws.max_row + 2
    ws.cell(axis_start, 1, "Совпадение по осям")
    ws.cell(axis_start, 1).fill = SUBHEADER_FILL
    ws.cell(axis_start, 1).font = Font(bold=True)
    ws.append(["Поле", "Совпало", "Всего", "Доля"])
    _style_header(ws, axis_start + 1)
    for field in NORMALIZED_FIELDS:
        metric = report["agreement"]["axes"][field]
        ws.append([FIELD_RU[field], metric["matches"], metric["total"], _pct(metric["rate"])])

    ws.freeze_panes = "A3"
    widths = {"A": 34, "B": 42, "C": 42, "D": 28}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER


def _normalized_text(values: dict[str, str]) -> str:
    return "\n".join(
        f"{FIELD_RU[field]}: {_human(values.get(field))}"
        for field in NORMALIZED_FIELDS
    )


def _write_case_sheet(ws, report: dict[str, Any]) -> None:
    headers = [
        "Case ID",
        "ID замечания",
        "Объект",
        "Раздел",
        "Документ",
        "Версия",
        "Сопоставим",
        "Статус baseline",
        "Статус candidate",
    ]
    for field in NORMALIZED_FIELDS:
        headers.extend([f"Baseline: {FIELD_RU[field]}", f"Candidate: {FIELD_RU[field]}"])
    headers.extend(
        [
            "Совпали все поля",
            "Evidence baseline: принято",
            "Evidence baseline: отклонено",
            "Evidence candidate: принято",
            "Evidence candidate: отклонено",
            "Опасные расхождения",
        ]
    )
    _style_title(ws, "Сравнение по замечаниям", len(headers))
    ws.append(headers)
    _style_header(ws, 2)
    for item in report["cases"]:
        row: list[Any] = [
            item["case_id"],
            item["item_id"],
            item["object_name"],
            item["discipline"],
            item["document"],
            item["version_id"],
            item["comparable"],
            _human(item["baseline_status"]),
            _human(item["candidate_status"]),
        ]
        for field in NORMALIZED_FIELDS:
            row.extend(
                [
                    _human(item["baseline"].get(field)),
                    _human(item["candidate"].get(field)),
                ]
            )
        row.extend(
            [
                item["all_normalized_fields_agree"],
                item["baseline_evidence"]["accepted"],
                item["baseline_evidence"]["rejected_by_guards"],
                item["candidate_evidence"]["accepted"],
                item["candidate_evidence"]["rejected_by_guards"],
                "\n".join(DANGER_RU.get(code, code) for code in item["danger_codes"]),
            ]
        )
        ws.append([_excel_value(value) for value in row])
        if item["danger_codes"]:
            for cell in ws[ws.max_row]:
                cell.fill = DANGER_FILL
        elif item["different_fields"]:
            for cell in ws[ws.max_row]:
                cell.fill = DIFF_FILL

    _add_table(ws, "AuditCasesComparison", 2, ws.max_row, len(headers))
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    for index in range(1, len(headers) + 1):
        if index in {1, 2, 4, 6, 7, 8, 9}:
            width = 18
        elif index in {3, 5}:
            width = 34
        else:
            width = 25
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER


def _write_disagreement_sheet(ws, report: dict[str, Any]) -> None:
    headers = [
        "Case ID",
        "ID замечания",
        "Замечание",
        "Сопоставим",
        "Различающиеся поля",
        "Опасные расхождения",
        "Baseline",
        "Candidate",
    ]
    _style_title(ws, "Расхождения", len(headers))
    ws.append(headers)
    _style_header(ws, 2)
    disagreements = [
        item
        for item in report["cases"]
        if item["different_fields"] or item["danger_codes"] or not item["comparable"]
    ]
    for item in disagreements:
        ws.append(
            [
                _excel_value(item["case_id"]),
                _excel_value(item["item_id"]),
                _excel_value(item["finding_problem"]),
                _excel_value(item["comparable"]),
                _excel_value(
                    ", ".join(FIELD_RU[field] for field in item["different_fields"])
                ),
                _excel_value(
                    "\n".join(
                        DANGER_RU.get(code, code) for code in item["danger_codes"]
                    )
                ),
                _excel_value(_normalized_text(item["baseline"])),
                _excel_value(_normalized_text(item["candidate"])),
            ]
        )
        if item["danger_codes"]:
            for cell in ws[ws.max_row]:
                cell.fill = DANGER_FILL
    _add_table(ws, "AuditDisagreements", 2, ws.max_row, len(headers))
    ws.freeze_panes = "A3"
    widths = [32, 18, 48, 16, 35, 48, 45, 45]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER


def write_xlsx(report: dict[str, Any], path: Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_summary_sheet(workbook.create_sheet("Сводка"), report)
    _write_case_sheet(workbook.create_sheet("По замечаниям"), report)
    _write_disagreement_sheet(workbook.create_sheet("Расхождения"), report)
    workbook.properties.title = "Сравнение прогонов аудита отклонённых замечаний"
    temporary = path.with_name(f".{path.name}.{os.getpid()}.tmp.xlsx")
    workbook.save(temporary)
    os.replace(temporary, path)


def _atomic_write_text(path: Path, text: str) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{os.getpid()}.tmp")
    temporary.write_text(text, encoding="utf-8")
    os.replace(temporary, path)


def write_reports(report: dict[str, Any], output_prefix: Path) -> dict[str, str]:
    output_prefix = Path(output_prefix)
    if output_prefix.suffix.lower() in {".json", ".md", ".xlsx"}:
        output_prefix = output_prefix.with_suffix("")
    json_path = output_prefix.with_suffix(".json")
    markdown_path = output_prefix.with_suffix(".md")
    xlsx_path = output_prefix.with_suffix(".xlsx")
    _atomic_write_text(
        json_path,
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
    )
    _atomic_write_text(markdown_path, render_markdown(report))
    write_xlsx(report, xlsx_path)
    return {
        "json": str(json_path.resolve()),
        "markdown": str(markdown_path.resolve()),
        "xlsx": str(xlsx_path.resolve()),
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Read-only сравнение двух прогонов аудита отклонённых замечаний"
    )
    parser.add_argument(
        "--baseline-dir",
        type=Path,
        required=True,
        help="каталог базового прогона с manifest.jsonl и results.jsonl",
    )
    parser.add_argument(
        "--candidate-dir",
        type=Path,
        required=True,
        help="каталог сравниваемого прогона с manifest.jsonl и results.jsonl",
    )
    parser.add_argument(
        "--output-prefix",
        type=Path,
        required=True,
        help="префикс для .json, .md и .xlsx",
    )
    parser.add_argument("--baseline-label", default="Sol baseline")
    parser.add_argument("--candidate-label", default="Luna candidate")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    report = compare_runs(
        args.baseline_dir,
        args.candidate_dir,
        baseline_label=args.baseline_label,
        candidate_label=args.candidate_label,
    )
    outputs = write_reports(report, args.output_prefix)
    print(
        "Сравнение готово: "
        f"{report['comparability']['comparable_success_count']} сопоставимых кейсов, "
        f"{report['dangerous_disagreements']['count']} опасных расхождений."
    )
    for label, path in outputs.items():
        print(f"  {label}: {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
