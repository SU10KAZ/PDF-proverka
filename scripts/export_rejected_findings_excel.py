#!/usr/bin/env python3
"""Export a rejected-findings audit directory to a readable Russian XLSX."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Callable
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


MAX_EXCEL_TEXT = 32_760
MOSCOW_TZ = ZoneInfo("Europe/Moscow")

VERDICT_RU = {
    "expert_correct": "Эксперт прав",
    "expert_may_be_wrong": "Эксперт мог ошибиться",
    "insufficient_evidence": "Недостаточно данных",
}
CONFIDENCE_RU = {"high": "Высокая", "medium": "Средняя", "low": "Низкая"}
ACTION_RU = {
    "keep_rejected": "Оставить отклонённым",
    "manual_recheck": "Ручная перепроверка",
    "collect_context": "Собрать недостающий контекст",
}
BINDING_RU = {
    "exact": "Причина относится к этому замечанию",
    "conflict": "Причина относится к другому предмету",
    "missing": "Привязка отсутствует",
}
FACTUAL_RU = {
    "supported": "Подтверждён",
    "unsupported": "Не подтверждён",
    "contradicted": "Опровергнут",
    "unclear": "Неясно",
}
REPORT_VALUE_RU = {
    "include": "Оставить отдельным пунктом",
    "merge": "Объединить",
    "downgrade": "Снизить значимость",
    "remove": "Убрать из отчёта",
    "unclear": "Неясно",
}
REASON_QUALITY_RU = {
    "substantiated": "Обоснована",
    "partial": "Частично обоснована",
    "unsubstantiated": "Не обоснована",
    "contradicted": "Опровергнута",
    "missing": "Причина отсутствует",
}
DECISION_EFFECT_RU = {
    "supports_rejection": "Подтверждает отклонение",
    "changes_rejection": "Может изменить отклонение",
    "reason_only": "Неточность только в причине",
    "unclear": "Неясно",
}
REJECTION_BASIS_RU = {
    "factual": "Фактический спор",
    "scope_stage": "Область документа или стадия",
    "report_value": "Ценность для отчёта",
    "duplicate": "Дубликат",
    "construction_state": "Состояние строительства",
    "mixed": "Смешанное основание",
    "unknown": "Неизвестно",
}
PRACTICAL_IMPACT_RU = {
    "high": "Высокое",
    "medium": "Среднее",
    "low": "Низкое",
    "none": "Отсутствует",
    "unclear": "Неясно",
}
SOURCE_ALIGNMENT_RU = {
    "not_visual": "Визуальная сверка не требуется",
    "confirmed_by_raster": "Подтверждено растром",
    "ocr_only_visual_claim": "Визуальный вывод только по OCR",
    "raster_text_conflict": "Конфликт растра и текста/OCR",
    "unreadable": "Источник нечитаем",
}
SCOPE_CONTEXT_RU = {
    "not_needed": "Связанный контекст не требуется",
    "verified_same_version": "Проверен в той же версии",
    "missing": "Связанный контекст отсутствует",
    "version_uncertain": "Версия не подтверждена",
    "conflict": "Источники конфликтуют",
}
REVIEW_PRIORITY_RU = {
    "high": "Высокий",
    "medium": "Средний",
    "low": "Низкий",
    "none": "Не назначен",
}
DECISION_ORIGIN_RU = {
    "human": "Ручное решение",
    "suspected_carryover": "Предположительно перенесённое решение",
    "carried_over": "Перенесённое решение",
}
STATUS_RU = {"success": "Успешно", "error": "Ошибка"}
EFFORT_RU = {
    "low": "Низкий",
    "medium": "Средний",
    "high": "Высокий",
    "xhigh": "Очень высокий",
    "max": "Максимальный",
}
EVIDENCE_SOURCE_RU = {
    "finding": "исходное замечание",
    "expert_reason": "причина отклонения эксперта",
    "graphic_block": "графический блок",
    "text_block": "текстовый блок",
    "document_text": "текст документа",
    "related_document": "связанный документ",
    "norm_context": "нормативный контекст",
}
DISCIPLINE_RU = {
    "AR": "АР — архитектурные решения",
    "AS": "АС — архитектурно-строительные решения",
    "GP": "ГП — генеральный план",
    "KJ": "КЖ — железобетонные конструкции",
    "KZh": "КЖ — железобетонные конструкции",
    "KM": "КМ — металлические конструкции",
    "OV": "ОВ — отопление и вентиляция",
    "VK": "ВК — водоснабжение и канализация",
    "EOM": "ЭОМ — электрооборудование и освещение",
    "SS": "СС — сети связи",
    "POS": "ПОС — организация строительства",
    "PB": "ПБ — пожарная безопасность",
}

COMMON_FLAG_RU = {
    "images_truncated": "Изображения обрезаны",
    "norm_not_in_index": "Норма отсутствует в индексе",
    "norm_not_indexed": "Норма не проиндексирована",
    "attached_images_missing": "Приложенные изображения не переданы",
    "attached_images_absent": "Приложенные изображения отсутствуют",
    "attached_images_empty": "Приложенные изображения пусты",
    "no_attached_images": "Нет приложенных изображений",
    "reason_item_mismatch": "Причина относится к другому замечанию",
    "source_images_truncated": "Изображения источника обрезаны",
    "service_or_carried_reason": "Служебная или перенесённая причина",
    "source_artifact_conflict": "Конфликт исходного артефакта",
    "same_version_artifact_conflict": "Конфликт артефактов той же версии",
    "suspected_carryover": "Вероятно перенесённое решение",
    "norm_status_unverified": "Статус нормы не проверен",
    "expert_origin_not_human": "Источник экспертного решения не ручной",
    "non_human_decision_origin": "Источник решения не ручной",
    "referenced_sheet_not_provided": "Указанный лист не предоставлен",
    "duplicate_reference_not_provided": "Исходное дублирующее замечание не предоставлено",
    "norm_context_unsupported": "Нормативный контекст не подтверждает довод",
    "external_reference_not_provided": "Внешний источник не предоставлен",
    "document_excerpt_truncated": "Фрагмент документа обрезан",
    "graphic_evidence_not_attached": "Графическое доказательство не приложено",
    "missing_expert_reason": "Причина эксперта отсутствует",
    "expert_reason_missing": "Причина эксперта отсутствует",
    "norm_context_not_on_point": "Нормативный контекст не относится к предмету замечания",
}

FLAG_TOKEN_RU = {
    "missing": "отсутствует",
    "not": "не",
    "provided": "предоставлен",
    "unverified": "не проверен",
    "unsupported": "не подтверждён",
    "contradicted": "опровергнут",
    "truncated": "обрезан",
    "incomplete": "неполный",
    "irrelevant": "не относится к предмету",
    "mismatch": "несоответствие",
    "conflict": "конфликт",
    "ambiguous": "неоднозначный",
    "expert": "эксперт",
    "reason": "причина",
    "finding": "замечание",
    "source": "источник",
    "graphic": "графический",
    "block": "блок",
    "document": "документ",
    "text": "текст",
    "norm": "норма",
    "normative": "нормативный",
    "context": "контекст",
    "evidence": "доказательство",
    "referenced": "указанный",
    "sheet": "лист",
    "images": "изображения",
    "image": "изображение",
    "attached": "приложенный",
    "external": "внешний",
    "status": "статус",
    "scope": "область применения",
    "claim": "утверждение",
    "decision": "решение",
    "carryover": "перенос",
    "suspected": "предполагаемый",
    "ocr": "OCR",
}

GUARD_EXACT_RU = {
    "evidence rejected: document_text: locator does not match loaded page/path": (
        "Доказательство отклонено: локатор текста документа не соответствует "
        "загруженной странице или пути."
    ),
    "expert_correct downgraded: finding binding is not exact": (
        "Вердикт «Эксперт прав» понижен: нет точной привязки к замечанию."
    ),
    "expert_correct downgraded: no case-validated external evidence": (
        "Вердикт «Эксперт прав» понижен: нет внешнего доказательства, "
        "проверенного для этого кейса."
    ),
    "expert_correct/remove downgraded: decision-critical context is missing": (
        "Вердикт понижен: отсутствует критичный для решения контекст."
    ),
    "expert_may_be_wrong downgraded: no validated evidence or binding mismatch": (
        "Вердикт «Эксперт мог ошибиться» понижен: нет проверенного доказательства "
        "или строгого конфликта привязки."
    ),
    "verdict downgraded: conflicting finding binding": (
        "Вердикт понижен: конфликт привязки причины к замечанию."
    ),
    "verdict downgraded: exact finding binding is missing": (
        "Вердикт понижен: отсутствует точная привязка к замечанию."
    ),
    "verdict downgraded: reason is service/carryover text": (
        "Вердикт понижен: причина является служебным или перенесённым текстом."
    ),
}

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="305496")
CANDIDATE_HEADER_FILL = PatternFill("solid", fgColor="C65911")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
AMBER_FILL = PatternFill("solid", fgColor="FCE4D6")
GRAY_FILL = PatternFill("solid", fgColor="E7E6E6")
BLUE_FILL = PatternFill("solid", fgColor="DDEBF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E1F2")
CELL_BORDER = Border(bottom=THIN_GRAY)


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f"Некорректный JSONL {path}:{line_number}: {exc}") from exc
    return rows


def select_current_successful_results(
    rows: list[dict[str, Any]], manifest: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    """Collapse the append-only result log using the report generator semantics."""
    latest_by_case: dict[str, dict[str, Any]] = {}
    for row in rows:
        case_id = row.get("case_id")
        if case_id:
            latest_by_case[str(case_id)] = row

    selected: list[dict[str, Any]] = []
    for case in manifest:
        case_id = str(case.get("case_id") or "")
        result = latest_by_case.get(case_id)
        if not result or result.get("status") != "success":
            continue
        if result.get("input_hash") != case.get("input_hash"):
            continue
        selected.append(result)
    return selected


def translated(value: Any, mapping: dict[str, str]) -> str:
    if value in (None, ""):
        return "—"
    text = str(value)
    return mapping.get(text, text)


def excel_text(value: Any) -> Any:
    if value is None or value == "":
        return "—"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value)).strip()
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    if len(text) > MAX_EXCEL_TEXT:
        text = text[: MAX_EXCEL_TEXT - 40] + "\n… [текст сокращён для Excel]"
    return text or "—"


def format_datetime(value: Any) -> str:
    if not value:
        return "—"
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=MOSCOW_TZ)
        return parsed.astimezone(MOSCOW_TZ).strftime("%d.%m.%Y %H:%M:%S МСК")
    except ValueError:
        return str(value)


def format_list(value: Any) -> str:
    if value in (None, "", []):
        return "—"
    if not isinstance(value, list):
        return str(value)
    return "\n".join(f"• {item}" for item in value) if value else "—"


def format_evidence(value: Any) -> str:
    if not value:
        return "—"
    if not isinstance(value, list):
        return str(value)

    rendered: list[str] = []
    known_fields = {
        "source",
        "source_id",
        "image_index",
        "block_id",
        "locator",
        "quote",
        "implication",
        "observation_basis",
        "verification_state",
        "claim_type",
        "absence_scope",
        "page",
        "page_number",
        "path",
        "norm_id",
        "paragraph",
    }
    fallback_labels = {
        "sheet": "лист",
        "section": "раздел",
        "title": "заголовок",
        "text": "текст",
        "value": "значение",
        "item_id": "ID элемента",
    }
    for index, item in enumerate(value, start=1):
        if not isinstance(item, dict):
            rendered.append(f"{index}. {item}")
            continue
        source = translated(item.get("source"), EVIDENCE_SOURCE_RU)
        parts = [f"{index}. Источник: {source}"]
        if item.get("source_id"):
            parts.append(f"ID источника: {item['source_id']}")
        if item.get("image_index") not in (None, ""):
            parts.append(f"изображение № {item['image_index']}")
        if item.get("block_id"):
            parts.append(f"блок {item['block_id']}")
        if item.get("page_number") not in (None, ""):
            parts.append(f"страница {item['page_number']}")
        elif item.get("page") not in (None, ""):
            parts.append(f"страница {item['page']}")
        if item.get("locator"):
            parts.append(f"локатор: {item['locator']}")
        if item.get("quote"):
            parts.append(f"цитата: «{item['quote']}»")
        if item.get("implication"):
            parts.append(f"вывод: {item['implication']}")
        if item.get("observation_basis"):
            parts.append(f"способ наблюдения: {item['observation_basis']}")
        if item.get("verification_state"):
            parts.append(f"сверка: {item['verification_state']}")
        if item.get("claim_type"):
            parts.append(f"тип утверждения: {item['claim_type']}")
        if item.get("absence_scope") not in (None, "", "none"):
            parts.append(f"область отсутствия: {item['absence_scope']}")
        if item.get("norm_id"):
            parts.append(f"норма: {item['norm_id']}")
        if item.get("paragraph"):
            parts.append(f"пункт: {item['paragraph']}")
        if item.get("path"):
            parts.append(f"путь: {item['path']}")
        for key, nested_value in item.items():
            if key in known_fields or nested_value in (None, "", [], {}):
                continue
            label = fallback_labels.get(key, key.replace("_", " "))
            if isinstance(nested_value, (dict, list)):
                nested_value = json.dumps(nested_value, ensure_ascii=False)
            parts.append(f"{label}: {nested_value}")
        rendered.append("; ".join(parts))
    return "\n\n".join(rendered)


def translate_guard(value: Any) -> str:
    text = str(value).strip()
    if text in GUARD_EXACT_RU:
        return GUARD_EXACT_RU[text]

    match = re.fullmatch(r"evidence rejected: ([a-z_]+): quote not found in (RF-[^ ]+) context", text)
    if match:
        source = translated(match.group(1), EVIDENCE_SOURCE_RU)
        return f"Доказательство отклонено: цитата из источника «{source}» не найдена в контексте кейса {match.group(2)}."

    match = re.fullmatch(r"evidence rejected: graphic_block: image_index/block_id not attached to (RF-[^ ]+)", text)
    if match:
        return f"Доказательство отклонено: изображение или графический блок не приложен к кейсу {match.group(1)}."

    match = re.fullmatch(r"evidence rejected: text_block: block_id not present in (RF-[^ ]+)", text)
    if match:
        return f"Доказательство отклонено: текстовый блок отсутствует в кейсе {match.group(1)}."

    return f"Техническая корректировка: {text}"


def format_guards(value: Any) -> str:
    if not value:
        return "—"
    values = value if isinstance(value, list) else [value]
    return "\n".join(f"• {translate_guard(item)}" for item in values)


def translate_flag(value: Any) -> str:
    code = str(value)
    if code in COMMON_FLAG_RU:
        return COMMON_FLAG_RU[code]
    words = [FLAG_TOKEN_RU.get(token, token) for token in code.split("_")]
    description = " ".join(words)
    return f"{description[:1].upper() + description[1:]} (код: {code})"


def format_flags(value: Any) -> str:
    if not value:
        return "—"
    values = value if isinstance(value, list) else [value]
    return "\n".join(f"• {translate_flag(item)}" for item in values)


def discipline_name(value: Any) -> str:
    if not value:
        return "—"
    text = str(value)
    return DISCIPLINE_RU.get(text, text)


def model_name(value: Any) -> str:
    if not value:
        return "—"
    return {
        "codex/gpt-5.6-sol": "Codex Sol (GPT-5.6)",
        "codex/gpt-5.6-terra": "Codex Terra (GPT-5.6)",
    }.get(str(value), str(value))


def candidate_basis(row: dict[str, Any]) -> str:
    reasons: list[str] = []
    if row.get("verdict") == "expert_may_be_wrong":
        reasons.append("Есть признаки, что эксперт мог ошибиться")
    if row.get("binding_status") == "conflict":
        reasons.append("причина отклонения относится к другому предмету")
    elif row.get("binding_status") == "missing":
        reasons.append("не найдена привязка причины к замечанию")
    if row.get("decision_origin") == "suspected_carryover":
        reasons.append("решение предположительно перенесено с другой версии")
    if row.get("reason_quality") in {"unsubstantiated", "contradicted", "missing"}:
        reasons.append(f"качество причины: {translated(row.get('reason_quality'), REASON_QUALITY_RU).lower()}")
    if not reasons:
        reasons.append("Правила аудита назначили ручную перепроверку из-за недостатка решающих доказательств")
    return "; ".join(reasons) + "."


def merge_rows(
    results: list[dict[str, Any]], manifest_by_case: dict[str, dict[str, Any]]
) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    for result in results:
        manifest = manifest_by_case.get(str(result.get("case_id")), {})
        row = dict(manifest)
        row.update(result)
        merged.append(row)
    return merged


Column = tuple[str, Callable[[dict[str, Any]], Any], float]


def set_workbook_metadata(workbook: Workbook) -> None:
    workbook.properties.title = "Аудит отклонённых замечаний за июль 2026"
    workbook.properties.subject = "Независимая проверка решений эксперта"
    workbook.properties.creator = "Codex — PDF-проверка"
    workbook.properties.description = (
        "Нормализованные результаты read-only аудита. Исходные решения эксперта не изменены."
    )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True


def apply_sheet_layout(ws: Any, widths: list[float], header_fill: PatternFill) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 34
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    ws.sheet_view.zoomScale = 80


def verdict_fill(verdict: str) -> PatternFill:
    return {
        "expert_correct": GREEN_FILL,
        "expert_may_be_wrong": AMBER_FILL,
        "insufficient_evidence": GRAY_FILL,
    }.get(verdict, BLUE_FILL)


def add_table_sheet(
    workbook: Workbook,
    title: str,
    columns: list[Column],
    rows: list[dict[str, Any]],
    *,
    header_fill: PatternFill,
    table_name: str,
    table_style: str,
) -> Any:
    ws = workbook.create_sheet(title)
    ws.append([column[0] for column in columns])
    path_columns = {
        index
        for index, (header, _, _) in enumerate(columns, start=1)
        if header in {"Файл исходных замечаний", "Файл решения эксперта"}
    }
    verdict_column = next(
        (
            index
            for index, (header, _, _) in enumerate(columns, start=1)
            if header == "Итог аудита"
        ),
        None,
    )
    for row_number, row in enumerate(rows, start=2):
        ws.append([excel_text(accessor(row)) for _, accessor, _ in columns])
        for cell in ws[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER
        if verdict_column is not None:
            ws.cell(row=row_number, column=verdict_column).fill = verdict_fill(
                str(row.get("verdict", ""))
            )

        for path_column in path_columns:
            cell = ws.cell(row=row_number, column=path_column)
            path_value = cell.value
            if isinstance(path_value, str) and path_value not in {"", "—"}:
                try:
                    cell.hyperlink = Path(path_value).as_uri()
                    cell.style = "Hyperlink"
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                except ValueError:
                    pass

    apply_sheet_layout(ws, [column[2] for column in columns], header_fill)
    if rows:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    return ws


def build_summary_sheet(
    workbook: Workbook,
    summary: dict[str, Any],
    results: list[dict[str, Any]],
    candidate_count: int,
) -> Any:
    ws = workbook.active
    ws.title = "Сводка"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "Аудит отклонённых замечаний — Кульдяев Ф. С."
    title.fill = TITLE_FILL
    title.font = Font(color="FFFFFF", bold=True, size=16)
    title.alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws[1]:
        cell.fill = TITLE_FILL
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = "Период решений: июль 2026 · режим анализа: Codex Sol / высокий уровень рассуждения"
    ws["A2"].font = Font(italic=True, color="44546A")

    parameters = [
        ("Период", "Июль 2026"),
        ("Эксперт", "Кульдяев Ф. С."),
        ("Всего замечаний", int(summary.get("selected_cases", len(results)))),
        ("Проверено", int(summary.get("completed", len(results)))),
        ("Ошибки анализа", int(summary.get("latest_errors", 0))),
        ("Кандидаты на ручную перепроверку", candidate_count),
        ("Модель", model_name(results[0].get("model")) if results else "—"),
        ("Уровень рассуждения", translated(results[0].get("reasoning_effort"), EFFORT_RU) if results else "—"),
        ("Сформировано", format_datetime(summary.get("generated_at"))),
        ("Экспортировано", datetime.now(MOSCOW_TZ).strftime("%d.%m.%Y %H:%M:%S МСК")),
    ]
    ws.append([])
    ws.append(["Параметр", "Значение"])
    for cell in ws[4]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    for label, value in parameters:
        ws.append([label, value])
    ws["B7"].number_format = "0"
    ws["B10"].hyperlink = "#'Кандидаты'!A1"
    ws["B10"].style = "Hyperlink"

    verdict_counts = Counter(str(row.get("verdict")) for row in results)
    total = len(results) or 1
    start_row = 4
    start_col = 4
    verdict_headers = ["Итог аудита", "Количество", "Доля", "Что это означает"]
    for offset, header in enumerate(verdict_headers):
        cell = ws.cell(row=start_row, column=start_col + offset, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    meanings = {
        "expert_correct": "Причина отклонения подтверждается; оставить отклонённым.",
        "expert_may_be_wrong": "Есть основания сомневаться; решение не менять автоматически, перепроверить вручную.",
        "insufficient_evidence": "Для уверенного вывода не хватает документа, блока или нормативного контекста.",
    }
    for index, verdict in enumerate(("expert_correct", "expert_may_be_wrong", "insufficient_evidence"), start=1):
        row_number = start_row + index
        count = verdict_counts.get(verdict, 0)
        values = [VERDICT_RU[verdict], count, count / total, meanings[verdict]]
        for offset, value in enumerate(values):
            cell = ws.cell(row=row_number, column=start_col + offset, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER
            if offset == 0:
                cell.fill = verdict_fill(verdict)
        ws.cell(row=row_number, column=start_col + 2).number_format = "0.00%"

    action_counts = Counter(str(row.get("recommended_action")) for row in results)
    action_row = 10
    for offset, header in enumerate(("Рекомендуемое действие", "Количество", "Доля")):
        cell = ws.cell(row=action_row, column=start_col + offset, value=header)
        cell.fill = CANDIDATE_HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for index, action in enumerate(("keep_rejected", "manual_recheck", "collect_context"), start=1):
        row_number = action_row + index
        count = action_counts.get(action, 0)
        ws.cell(row=row_number, column=start_col, value=ACTION_RU[action])
        ws.cell(row=row_number, column=start_col + 1, value=count)
        ws.cell(row=row_number, column=start_col + 2, value=count / total).number_format = "0.00%"
        for column in range(start_col, start_col + 3):
            ws.cell(row=row_number, column=column).border = CELL_BORDER
            ws.cell(row=row_number, column=column).alignment = Alignment(vertical="top", wrap_text=True)

    note_row = 16
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
    ws.cell(row=note_row, column=1, value=(
        "Важно: «Эксперт мог ошибиться» и вкладка «Кандидаты» означают только необходимость "
        "ручной перепроверки. Исходные решения эксперта, замечания и PDF-файлы этим экспортом не изменялись."
    ))
    ws.cell(row=note_row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=note_row, column=1).font = Font(bold=True, color="1F1F1F")
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[note_row].height = 48

    ws["A18"] = "Перейти к полному списку"
    ws["A18"].hyperlink = "#'Все результаты'!A1"
    ws["A18"].style = "Hyperlink"
    ws["D18"] = "Перейти к кандидатам"
    ws["D18"].hyperlink = "#'Кандидаты'!A1"
    ws["D18"].style = "Hyperlink"

    widths = {"A": 39, "B": 30, "C": 3, "D": 31, "E": 14, "F": 12, "G": 62, "H": 3}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A4"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    return ws


def build_workbook(audit_dir: Path, output_path: Path) -> tuple[int, int]:
    summary = load_json(audit_dir / "summary.json")
    manifest = load_jsonl(audit_dir / "manifest.jsonl")
    result_log = load_jsonl(audit_dir / "results.jsonl")
    results = select_current_successful_results(result_log, manifest)
    candidates_payload = load_json(audit_dir / "candidates.json")
    candidates = candidates_payload.get("candidates", [])

    expected = int(summary.get("completed", len(results)))
    if len(results) != expected:
        raise ValueError(
            "После выбора последних актуальных успешных результатов "
            f"получено {len(results)} записей, в summary.json completed={expected}"
        )
    result_ids = [str(row.get("case_id")) for row in results]

    manifest_by_case = {str(row.get("case_id")): row for row in manifest}
    merged_results = merge_rows(results, manifest_by_case)
    candidate_ids = {str(row.get("case_id")) for row in candidates}
    manual_ids = {
        str(row.get("case_id"))
        for row in merged_results
        if row.get("recommended_action") == "manual_recheck"
    }
    if candidate_ids != manual_ids:
        raise ValueError("candidates.json не совпадает с recommended_action=manual_recheck")

    # Always use the normalized verdict from results.jsonl. raw_verdict is intentionally omitted.
    merged_candidates = [row for row in merged_results if str(row.get("case_id")) in candidate_ids]

    workbook = Workbook()
    set_workbook_metadata(workbook)
    build_summary_sheet(workbook, summary, merged_results, len(merged_candidates))

    all_columns: list[Column] = [
        ("№", lambda row: row["_index"], 7),
        ("ID кейса", lambda row: row.get("case_id"), 26),
        ("Итог аудита", lambda row: translated(row.get("verdict"), VERDICT_RU), 23),
        (
            "Рекомендуемое действие",
            lambda row: translated(row.get("recommended_action"), ACTION_RU),
            31,
        ),
        ("Приоритет перепроверки", lambda row: translated(row.get("review_priority"), REVIEW_PRIORITY_RU), 22),
        ("Уверенность", lambda row: translated(row.get("confidence"), CONFIDENCE_RU), 15),
        ("Статус анализа", lambda row: translated(row.get("status"), STATUS_RU), 17),
        ("Объект", lambda row: row.get("object_name"), 34),
        ("Раздел", lambda row: discipline_name(row.get("discipline")), 30),
        ("Документ", lambda row: row.get("document"), 32),
        ("Версия", lambda row: row.get("version_id"), 12),
        ("ID замечания", lambda row: row.get("item_id"), 17),
        (
            "Дата решения (МСК)",
            lambda row: format_datetime(
                row.get("expert_timestamp_local") or row.get("expert_timestamp")
            ),
            23,
        ),
        ("Источник решения", lambda row: translated(row.get("decision_origin"), DECISION_ORIGIN_RU), 31),
        ("Привязка причины", lambda row: translated(row.get("binding_status"), BINDING_RU), 38),
        ("Факт по документам", lambda row: translated(row.get("factual_verdict"), FACTUAL_RU), 20),
        ("Ценность для отчёта", lambda row: translated(row.get("report_value"), REPORT_VALUE_RU), 25),
        ("Качество причины", lambda row: translated(row.get("reason_quality"), REASON_QUALITY_RU), 23),
        ("Влияние на решение", lambda row: translated(row.get("decision_effect"), DECISION_EFFECT_RU), 27),
        ("Основание отклонения", lambda row: translated(row.get("rejection_basis"), REJECTION_BASIS_RU), 30),
        ("Практическое влияние", lambda row: translated(row.get("practical_impact"), PRACTICAL_IMPACT_RU), 22),
        ("Оценка практического влияния", lambda row: row.get("impact_assessment"), 52),
        ("Сверка OCR/растра", lambda row: translated(row.get("source_alignment"), SOURCE_ALIGNMENT_RU), 31),
        ("Связанный контекст", lambda row: translated(row.get("scope_context_status"), SCOPE_CONTEXT_RU), 31),
        ("Исходное замечание", lambda row: row.get("finding_problem"), 58),
        ("Причина отклонения эксперта", lambda row: row.get("expert_reason"), 58),
        ("Оценка причины эксперта", lambda row: row.get("reason_assessment"), 62),
        ("Оценка замечания", lambda row: row.get("finding_assessment"), 62),
        ("Оценка нормативного контекста", lambda row: row.get("norm_assessment"), 58),
        ("Решающие доказательства", lambda row: format_evidence(row.get("decisive_evidence")), 70),
        ("Недостающий контекст", lambda row: format_list(row.get("missing_context")), 58),
        ("Корректировки защитных правил", lambda row: format_guards(row.get("guard_adjustments")), 58),
    ]
    indexed_results = [dict(row, _index=index) for index, row in enumerate(merged_results, start=1)]
    add_table_sheet(
        workbook,
        "Все результаты",
        all_columns,
        indexed_results,
        header_fill=HEADER_FILL,
        table_name="AllAuditResults",
        table_style="TableStyleMedium2",
    )

    candidate_columns: list[Column] = [
        ("№", lambda row: row["_index"], 7),
        ("ID кейса", lambda row: row.get("case_id"), 26),
        ("Итог аудита", lambda row: translated(row.get("verdict"), VERDICT_RU), 23),
        ("Основание перепроверки", candidate_basis, 58),
        ("Приоритет перепроверки", lambda row: translated(row.get("review_priority"), REVIEW_PRIORITY_RU), 22),
        ("Уверенность", lambda row: translated(row.get("confidence"), CONFIDENCE_RU), 15),
        ("Объект", lambda row: row.get("object_name"), 34),
        ("Раздел", lambda row: discipline_name(row.get("discipline")), 30),
        ("Документ", lambda row: row.get("document"), 32),
        ("Версия", lambda row: row.get("version_id"), 12),
        ("ID замечания", lambda row: row.get("item_id"), 17),
        ("Дата решения (МСК)", lambda row: format_datetime(row.get("expert_timestamp_local") or row.get("expert_timestamp")), 23),
        ("Источник решения", lambda row: translated(row.get("decision_origin"), DECISION_ORIGIN_RU), 31),
        ("Привязка причины", lambda row: translated(row.get("binding_status"), BINDING_RU), 38),
        ("Факт по документам", lambda row: translated(row.get("factual_verdict"), FACTUAL_RU), 20),
        ("Ценность для отчёта", lambda row: translated(row.get("report_value"), REPORT_VALUE_RU), 25),
        ("Качество причины", lambda row: translated(row.get("reason_quality"), REASON_QUALITY_RU), 23),
        ("Влияние на решение", lambda row: translated(row.get("decision_effect"), DECISION_EFFECT_RU), 27),
        ("Основание отклонения", lambda row: translated(row.get("rejection_basis"), REJECTION_BASIS_RU), 30),
        ("Практическое влияние", lambda row: translated(row.get("practical_impact"), PRACTICAL_IMPACT_RU), 22),
        ("Оценка практического влияния", lambda row: row.get("impact_assessment"), 52),
        ("Сверка OCR/растра", lambda row: translated(row.get("source_alignment"), SOURCE_ALIGNMENT_RU), 31),
        ("Связанный контекст", lambda row: translated(row.get("scope_context_status"), SCOPE_CONTEXT_RU), 31),
        ("Исходное замечание", lambda row: row.get("finding_problem"), 58),
        ("Причина отклонения эксперта", lambda row: row.get("expert_reason"), 58),
        ("Оценка причины эксперта", lambda row: row.get("reason_assessment"), 62),
        ("Оценка замечания", lambda row: row.get("finding_assessment"), 62),
        ("Оценка нормативного контекста", lambda row: row.get("norm_assessment"), 58),
        ("Решающие доказательства", lambda row: format_evidence(row.get("decisive_evidence")), 70),
        ("Недостающий контекст", lambda row: format_list(row.get("missing_context")), 58),
        ("Корректировки защитных правил", lambda row: format_guards(row.get("guard_adjustments")), 58),
        ("Ограничения и служебные флаги", lambda row: format_flags(row.get("integrity_flags")), 58),
        ("Файл исходных замечаний", lambda row: row.get("source_item_path"), 55),
        ("Файл решения эксперта", lambda row: row.get("review_path"), 55),
    ]
    indexed_candidates = [dict(row, _index=index) for index, row in enumerate(merged_candidates, start=1)]
    add_table_sheet(
        workbook,
        "Кандидаты",
        candidate_columns,
        indexed_candidates,
        header_fill=CANDIDATE_HEADER_FILL,
        table_name="ManualRecheckCandidates",
        table_style="TableStyleMedium9",
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    # Reopen the binary and validate the user-visible contract.
    checked = load_workbook(output_path, read_only=True, data_only=False)
    if checked.sheetnames != ["Сводка", "Все результаты", "Кандидаты"]:
        raise ValueError(f"Некорректные вкладки: {checked.sheetnames}")
    if checked["Все результаты"].max_row != len(results) + 1:
        raise ValueError("Некорректное число строк на вкладке «Все результаты»")
    if checked["Кандидаты"].max_row != len(candidates) + 1:
        raise ValueError("Некорректное число строк на вкладке «Кандидаты»")
    all_results_sheet = checked["Все результаты"]
    headers = [
        cell.value
        for cell in next(all_results_sheet.iter_rows(min_row=1, max_row=1))
    ]
    if "raw_verdict" in headers or "Сырой вердикт" in headers:
        raise ValueError("В экспорт ошибочно попал raw_verdict")
    required_headers = {
        "ID кейса",
        "Итог аудита",
        "Рекомендуемое действие",
        "Уверенность",
        "Статус анализа",
        "Объект",
        "Раздел",
        "Документ",
        "Версия",
        "ID замечания",
        "Дата решения (МСК)",
    }
    missing_headers = sorted(required_headers - set(headers))
    if missing_headers:
        raise ValueError(
            "На вкладке «Все результаты» отсутствуют обязательные колонки: "
            + ", ".join(missing_headers)
        )

    def sheet_column_values(sheet: Any, header: str) -> list[Any]:
        column = headers.index(header) + 1
        return [
            row[0]
            for row in sheet.iter_rows(
                min_row=2,
                min_col=column,
                max_col=column,
                values_only=True,
            )
        ]

    if sheet_column_values(all_results_sheet, "ID кейса") != result_ids:
        raise ValueError("ID кейсов на вкладке «Все результаты» не совпадают с results.jsonl")
    expected_verdicts = [translated(row.get("verdict"), VERDICT_RU) for row in merged_results]
    if sheet_column_values(all_results_sheet, "Итог аудита") != expected_verdicts:
        raise ValueError("В Excel попали не нормализованные итоговые вердикты")
    expected_actions = [
        translated(row.get("recommended_action"), ACTION_RU) for row in merged_results
    ]
    if sheet_column_values(all_results_sheet, "Рекомендуемое действие") != expected_actions:
        raise ValueError("Рекомендуемые действия в Excel не совпадают с results.jsonl")

    candidate_sheet = checked["Кандидаты"]
    candidate_headers = [
        cell.value
        for cell in next(candidate_sheet.iter_rows(min_row=1, max_row=1))
    ]
    candidate_id_column = candidate_headers.index("ID кейса") + 1
    exported_candidate_ids = [
        row[0]
        for row in candidate_sheet.iter_rows(
            min_row=2,
            min_col=candidate_id_column,
            max_col=candidate_id_column,
            values_only=True,
        )
    ]
    expected_candidate_ids = [str(row.get("case_id")) for row in merged_candidates]
    if exported_candidate_ids != expected_candidate_ids:
        raise ValueError("ID кейсов на вкладке «Кандидаты» не совпадают с candidates.json")
    checked.close()
    return len(results), len(candidates)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("audit_dir", type=Path, help="Каталог завершённого аудита")
    parser.add_argument("--output", type=Path, help="Путь выходного XLSX")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    audit_dir = args.audit_dir.resolve()
    output_path = args.output or audit_dir / "Аудит_замечаний_Кульдяев_июль_2026.xlsx"
    result_count, candidate_count = build_workbook(audit_dir, output_path.resolve())
    print(f"Готово: {output_path.resolve()}")
    print(f"Все результаты: {result_count}; кандидаты: {candidate_count}")


if __name__ == "__main__":
    main()
