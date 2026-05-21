"""REST API для реестра внешних замечаний (письма заказчика).

Импорт markdown → LLM-сопоставление → подтверждение/отказ от match'ей →
Excel-экспорт coverage.
"""
from __future__ import annotations

import io
import json
import logging
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from backend.app.services.common.object_service import (
    get_current_id,
    get_object_by_id,
)
from backend.app.services.external_register import matcher, section_map, service
from backend.app.services.external_register.models import (
    CustomerResponse,
    FindingMatch,
    MatchStatus,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/external-register", tags=["external-register"])


# ─── Models ─────────────────────────────────────────────────────────


class ImportRequest(BaseModel):
    object_id: Optional[str] = None
    source_md_path: str
    register_id: str = "su10_2026-05-13"


class MatchRunRequest(BaseModel):
    only_section: Optional[str] = None
    model: Optional[str] = None


class ConfirmRequest(BaseModel):
    user: str = "operator"
    finding_id: Optional[str] = None
    project_id: Optional[str] = None


class RejectRequest(BaseModel):
    user: str = "operator"


# ─── Helpers ────────────────────────────────────────────────────────


def _resolve_object_id(object_id: Optional[str]) -> str:
    if object_id:
        if get_object_by_id(object_id) is None:
            raise HTTPException(404, f"Unknown object_id: {object_id}")
        return object_id
    current = get_current_id()
    if not current:
        raise HTTPException(400, "No current object selected and object_id not provided")
    return current


# ─── Endpoints ──────────────────────────────────────────────────────


@router.get("/{object_id}")
async def list_register_entries(object_id: str, register_id: str = "su10_2026-05-13"):
    obj_id = _resolve_object_id(object_id if object_id != "_" else None)
    register = service.load_register(obj_id, register_id)
    if register is None:
        raise HTTPException(404, f"Register {register_id} not found for {obj_id}")
    return json.loads(register.model_dump_json())


@router.get("/{object_id}/coverage")
async def get_coverage(object_id: str, register_id: str = "su10_2026-05-13"):
    obj_id = _resolve_object_id(object_id if object_id != "_" else None)
    cov = service.coverage(obj_id, register_id)
    if cov is None:
        raise HTTPException(404, f"Register {register_id} not found")
    return cov


@router.get("/{object_id}/registers")
async def list_registers(object_id: str):
    obj_id = _resolve_object_id(object_id if object_id != "_" else None)
    return {"registers": service.list_registers(obj_id)}


@router.post("/import")
async def import_register_endpoint(req: ImportRequest):
    obj_id = _resolve_object_id(req.object_id)
    src = Path(req.source_md_path)
    if not src.is_absolute():
        from backend.app.core.config import ROOT_DIR
        src = (ROOT_DIR / req.source_md_path).resolve()
    if not src.exists():
        raise HTTPException(404, f"Source file not found: {src}")
    try:
        register = service.import_register(obj_id, req.register_id, src)
    except Exception as e:
        logger.exception("import failed")
        raise HTTPException(500, f"Import failed: {e}")
    return {
        "ok": True,
        "register_id": register.register_id,
        "object_id": register.object_id,
        "entries_total": len(register.entries),
        "unmapped_sections": register.unmapped_sections,
    }


@router.post("/{object_id}/match")
async def run_match_endpoint(
    object_id: str,
    req: MatchRunRequest,
    background_tasks: BackgroundTasks,
    register_id: str = "su10_2026-05-13",
):
    """Запустить LLM-сопоставление как фоновую задачу.

    Прогресс не транслируется в этом эндпоинте — клиент сам делает GET /coverage
    для опроса (либо открывает /ws/global если потребуется).
    """
    obj_id = _resolve_object_id(object_id if object_id != "_" else None)
    register = service.load_register(obj_id, register_id)
    if register is None:
        raise HTTPException(404, f"Register {register_id} not imported")

    async def _bg():
        try:
            stats = await matcher.match_register(
                obj_id,
                register_id,
                only_section=req.only_section,
                model=req.model,
            )
            logger.info("[external_register] match done: %s", stats)
        except Exception as e:
            logger.exception("matcher failed: %s", e)

    background_tasks.add_task(_bg)
    return {"ok": True, "status": "started", "only_section": req.only_section}


@router.post("/{object_id}/entry/{entry_key:path}/confirm")
async def confirm_entry(
    object_id: str,
    entry_key: str,
    req: ConfirmRequest,
    register_id: str = "su10_2026-05-13",
):
    obj_id = _resolve_object_id(object_id if object_id != "_" else None)
    entry = service.confirm_match(
        object_id=obj_id,
        register_id=register_id,
        entry_key=entry_key,
        user=req.user,
        finding_id=req.finding_id,
        project_id=req.project_id,
    )
    if entry is None:
        raise HTTPException(404, "Entry not found")
    return json.loads(entry.model_dump_json())


@router.post("/{object_id}/entry/{entry_key:path}/reject")
async def reject_entry(
    object_id: str,
    entry_key: str,
    req: RejectRequest,
    register_id: str = "su10_2026-05-13",
):
    obj_id = _resolve_object_id(object_id if object_id != "_" else None)
    entry = service.reject_match(
        object_id=obj_id,
        register_id=register_id,
        entry_key=entry_key,
        user=req.user,
    )
    if entry is None:
        raise HTTPException(404, "Entry not found")
    return json.loads(entry.model_dump_json())


@router.get("/{object_id}/export.xlsx")
async def export_xlsx(
    object_id: str,
    register_id: str = "su10_2026-05-13",
):
    """Excel-отчёт: лист 1 — реестр + match, лист 2 — coverage сводка."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    obj_id = _resolve_object_id(object_id if object_id != "_" else None)
    register = service.load_register(obj_id, register_id)
    if register is None:
        raise HTTPException(404, f"Register {register_id} not found")
    cov = service.coverage(obj_id, register_id) or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр СУ-10"

    header = [
        "#", "Раздел", "Лист/Раздел", "Категория (СУ-10)",
        "Проблема", "Решение",
        "Ответ заказчика", "Комментарий заказчика",
        "Project", "Finding ID", "Confidence", "Status",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")

    status_color = {
        "auto_matched": "C6EFCE",
        "confirmed": "92D050",
        "needs_review": "FFEB9C",
        "unmatched": "FFFFFF",
        "rejected": "FFC7CE",
    }

    for i, entry in enumerate(register.entries, start=1):
        match = entry.match
        ws.append([
            i,
            entry.section_code,
            entry.sheet_ref,
            entry.cat_su10,
            entry.problem,
            entry.proposed_solution,
            entry.customer_response.value,
            entry.customer_comment,
            match.project_id if match else "",
            match.finding_id if match else "",
            round(match.confidence, 2) if match else "",
            entry.match_status.value,
        ])
        row = ws[ws.max_row]
        bg = status_color.get(entry.match_status.value)
        if bg:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=bg)
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [5, 22, 25, 18, 50, 50, 18, 30, 25, 12, 10, 14]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    # Coverage sheet
    ws2 = wb.create_sheet("Coverage")
    ws2.append(["Метрика", "Значение"])
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    for key in ("total", "matched", "needs_review", "unmatched"):
        ws2.append([key, cov.get(key, 0)])
    ws2.append([])
    ws2.append(["По статусу", ""])
    for k, v in (cov.get("by_status") or {}).items():
        ws2.append([k, v])
    ws2.append([])
    ws2.append(["По ответу заказчика", ""])
    for k, v in (cov.get("by_customer_response") or {}).items():
        ws2.append([k, v])
    ws2.append([])
    ws2.append(["По разделу: matched / total", ""])
    for sec, stat in sorted((cov.get("by_section") or {}).items()):
        ws2.append([sec, f"{stat.get('matched', 0)} / {stat.get('total', 0)}"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"external_register_{register_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/_/section-map")
async def get_section_map():
    """Полный список известных section_code → project_id (для UI)."""
    return {
        "map": dict(section_map.SECTION_TO_PROJECT),
        "known_codes": section_map.all_known_codes(),
    }
