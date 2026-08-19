"""REST API для раздела «Сравнение стадий».

MVP-набор endpoint'ов:
  • POST   /api/stage-comparison/sessions
  • GET    /api/stage-comparison/sessions
  • GET    /api/stage-comparison/sessions/{session_id}
  • GET    /api/stage-comparison/sessions/{session_id}/pairs/{pair_id}
  • GET    /api/stage-comparison/sessions/{session_id}/pairs/{pair_id}/page-image
  • DELETE /api/stage-comparison/sessions/{session_id}/pairs/{pair_id}/links
  • GET    /api/stage-comparison/sessions/{session_id}/pairs/{pair_id}/text-diff
  • GET    /api/stage-comparison/sessions/{session_id}/pairs/{pair_id}/graphic-summary
  • POST   /api/stage-comparison/sessions/{session_id}/pairs/{pair_id}/graphic-diff
  • GET    /api/stage-comparison/sessions/{session_id}/pairs/{pair_id}/block-image

Платные LLM-сравнения графики НЕ запускаются автоматически: только при
run_paid=true в POST .../graphic-diff и только через paid_api_guard.
"""
from __future__ import annotations

import asyncio
import base64
import io
import json
import logging
import re
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, File, Form, HTTPException, Query, UploadFile
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import FileResponse, Response, StreamingResponse
from pydantic import BaseModel, Field

from backend.app.services.stage_comparison import diff_text, store, jobs as jobs_mod, objects as objects_mod
from backend.app.services.stage_comparison import text_llm as text_llm_mod, text_llm_jobs as text_llm_jobs_mod
from backend.app.services.stage_comparison import text_llm_provider as text_llm_provider_mod
from backend.app.services.stage_comparison import text_llm_preflight as text_llm_preflight_mod
from backend.app.services.stage_comparison import text_llm_flat as text_llm_flat_mod
from backend.app.services.stage_comparison import md_image_enrichment as md_enrichment_mod
from backend.app.services.stage_comparison import pipeline_v2_payload_service as pipeline_v2_payload_mod
from backend.app.services.stage_comparison import pipeline_v2_run_jobs as pipeline_v2_run_jobs_mod
from backend.app.services.stage_comparison import visual_block_equivalence_jobs as vbe_jobs_mod
from backend.app.services.stage_comparison import clear_analysis as clear_analysis_mod
from backend.app.services.stage_comparison import opus_only as opus_only_mod
from backend.app.services.stage_comparison import enriched_comparison as enriched_compare_mod
from backend.app.services.stage_comparison import unified_analysis as unified_analysis_mod
from backend.app.services.stage_comparison import unified_analysis_jobs as unified_jobs_mod
from backend.app.services.stage_comparison import unified_findings as unified_findings_mod
from backend.app.services.stage_comparison import unified_grouping as unified_grouping_mod
from backend.app.services.stage_comparison import expert_review as expert_review_mod
from backend.app.services.stage_comparison import v2_review as v2_review_mod
from backend.app.services.stage_comparison import review_transfer as review_transfer_mod
from backend.app.services.stage_comparison import paths as sc_paths_mod
from backend.app.services.stage_comparison import saved_config as saved_config_mod
from backend.app.services.stage_comparison import stage_upload as stage_upload_mod
from backend.app.services.stage_comparison import diagnostic_new_pipeline

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/stage-comparison", tags=["stage-comparison"])


GRAPHIC_DIFF_PROMPT = (
    "Сравни два изображения проектной документации. "
    "Первое изображение относится к предыдущей стадии проекта, "
    "второе — к новой стадии. Найди все значимые отличия: новые элементы, "
    "удалённые элементы, изменение размеров, изменение подписей, "
    "изменение расположения, изменение условных обозначений, "
    "изменение таблиц или схем. Ответ дай структурированным списком "
    "на русском языке. Не выдумывай отличия, если их не видно."
)


# ─── Pydantic-модели тела запросов ───────────────────────────────────────


class CreateSessionRequest(BaseModel):
    stage_a_path: str = Field(..., description="Путь к папке первой стадии")
    stage_b_path: str = Field(..., description="Путь к папке второй стадии")


class DeleteLinkRequest(BaseModel):
    left_block_id: str
    right_block_id: str


class GraphicDiffRequest(BaseModel):
    left_block_id: str
    right_block_id: str
    run_paid: bool = False
    model: Optional[str] = None  # по умолчанию config-овый


class AlignmentItem(BaseModel):
    slot: Optional[int] = None
    left_page: Optional[int] = None
    right_page: Optional[int] = None
    mode: Optional[str] = None
    note: Optional[str] = ""


class SaveAlignmentRequest(BaseModel):
    items: list[AlignmentItem]
    force: bool = False


class InsertBlankRequest(BaseModel):
    slot: int = Field(..., ge=1)
    side: str = Field(..., pattern="^(left|right)$")


class MoveAlignmentRequest(BaseModel):
    slot: int = Field(..., ge=1)
    direction: str = Field(..., pattern="^(up|down)$")


class V2ReviewPatch(BaseModel):
    """Частичный patch ручного статуса верификации одного V2-изменения."""
    review_status: Optional[str] = None
    review_comment: Optional[str] = None
    reviewed_by: Optional[str] = None


class V2ReviewBulkPatch(BaseModel):
    ids: list[str] = Field(default_factory=list)
    patch: V2ReviewPatch = Field(default_factory=V2ReviewPatch)


class V2ReviewTransferRequest(BaseModel):
    """Запрос переноса решений из «Расхождений» в V2 (на всю сессию)."""
    use_claude: bool = True


class InsertBlankSideRequest(BaseModel):
    slot: int = Field(..., ge=1)
    side: str = Field(..., pattern="^(left|right)$")


class MovePageSideRequest(BaseModel):
    slot: int = Field(..., ge=1)
    side: str = Field(..., pattern="^(left|right)$")
    direction: str = Field(..., pattern="^(up|down)$")


class DeletePageSideRequest(BaseModel):
    slot: int = Field(..., ge=1)
    side: str = Field(..., pattern="^(left|right)$")


class UpdatePairMatchRequest(BaseModel):
    right_pdf: Optional[str] = None
    right_md: Optional[str] = None
    right_result_json: Optional[str] = None
    status: Optional[str] = "manual"


class CreateManualPairRequest(BaseModel):
    left_pdf: Optional[str] = None
    right_pdf: Optional[str] = None
    left_md: Optional[str] = None
    left_result_json: Optional[str] = None
    right_md: Optional[str] = None
    right_result_json: Optional[str] = None


class DeletePairRequest(BaseModel):
    hard: bool = False


class GraphicDiffJobItem(BaseModel):
    pair_id: str
    left_block_id: str
    right_block_id: str


class CreateGraphicDiffJobRequest(BaseModel):
    scope: str = Field(..., pattern="^(selected|pair|session)$")
    pair_id: Optional[str] = None
    items: Optional[list[GraphicDiffJobItem]] = None
    run_paid: bool = False
    confirm_paid: bool = False
    model: Optional[str] = None


# ─── Sessions ────────────────────────────────────────────────────────────


@router.post("/sessions")
async def create_session(req: CreateSessionRequest):
    if not (req.stage_a_path or "").strip() or not (req.stage_b_path or "").strip():
        raise HTTPException(400, "stage_a_path и stage_b_path обязательны")
    # Allowlist (Задача 8): если env задан — проверим, что обе папки внутри
    try:
        store.assert_path_in_allowlist(req.stage_a_path)
        store.assert_path_in_allowlist(req.stage_b_path)
    except PermissionError as exc:
        raise HTTPException(403, str(exc)) from exc
    try:
        session, warnings = store.create_session(req.stage_a_path, req.stage_b_path)
    except OSError as exc:
        raise HTTPException(400, f"Ошибка доступа к папкам: {exc}") from exc
    return {
        "session_id": session["id"],
        "pairs": session.get("pairs") or [],
        "warnings": warnings,
        "created_at": session.get("created_at"),
        "stage_a_path": session.get("stage_a_path"),
        "stage_b_path": session.get("stage_b_path"),
    }


@router.get("/objects")
async def list_comparison_objects():
    """Автосписок «объектов» под allowlist-root'ами.

    Объект = подпапка с минимум двумя `stage_*` директориями внутри.
    UI использует это для селекта вместо ручного ввода путей.
    """
    return objects_mod.list_objects()


@router.post("/objects/{object_id}/stages/{stage_name}/upload")
async def upload_stage_archive_endpoint(
    object_id: str,
    stage_name: str,
    file: UploadFile = File(...),
):
    """Импортировать ZIP в stage_1/stage_2 выбранного platform object.

    Архив сначала полностью проверяется и распаковывается во временную папку.
    Документы становятся новыми версиями; состояние до импорта дополнительно
    сохраняется в recoverable backup.
    """
    if stage_name not in stage_upload_mod.VALID_STAGES:
        raise HTTPException(400, "Разрешены только stage_1 и stage_2")
    try:
        return await run_in_threadpool(
            stage_upload_mod.replace_stage_from_zip,
            object_id,
            stage_name,
            file.file,
            file.filename,
        )
    except stage_upload_mod.StageUploadError as exc:
        raise HTTPException(400, str(exc)) from exc
    except OSError as exc:
        logger.exception("stage archive upload failed: %s/%s", object_id, stage_name)
        raise HTTPException(500, f"Не удалось сохранить архив стадии: {exc}") from exc


@router.post("/objects/{object_id}/stages/{stage_name}/upload-folder")
async def upload_stage_folder_endpoint(
    object_id: str,
    stage_name: str,
    files: list[UploadFile] = File(...),
    relative_paths: str = Form("[]"),
    folder_name: str = Form(""),
):
    """Загрузить одним действием целую папку с компьютера пользователя."""
    if stage_name not in stage_upload_mod.VALID_STAGES:
        raise HTTPException(400, "Разрешены только stage_1 и stage_2")
    if not files:
        raise HTTPException(422, "Выбранная папка не содержит файлов")
    try:
        parsed_paths = json.loads(relative_paths or "[]")
    except json.JSONDecodeError as exc:
        raise HTTPException(422, "Некорректный список путей файлов") from exc
    if not isinstance(parsed_paths, list) or len(parsed_paths) != len(files):
        raise HTTPException(422, "Количество файлов и относительных путей не совпадает")
    uploads = [
        (upload.file, str(parsed_paths[index] or upload.filename or ""))
        for index, upload in enumerate(files)
    ]
    try:
        return await run_in_threadpool(
            stage_upload_mod.replace_stage_from_folder,
            object_id,
            stage_name,
            uploads,
            folder_name,
        )
    except stage_upload_mod.StageUploadError as exc:
        raise HTTPException(400, str(exc)) from exc
    except OSError as exc:
        logger.exception("stage folder upload failed: %s/%s", object_id, stage_name)
        raise HTTPException(500, f"Не удалось сохранить папку стадии: {exc}") from exc


# ─── Saved configuration ──────────────────────────────────────────────────


class SaveConfigRequest(BaseModel):
    """Тело для PUT /saved-config (legacy path-only сохранение).

    Полная каноничная конфигурация сохраняется через
    POST /sessions/{sid}/save-canonical — там подтягиваются пары/режимы.
    """

    stage_a_path: str = Field(..., description="Абсолютный путь к stage_1 директории")
    stage_b_path: str = Field(..., description="Абсолютный путь к stage_2 директории")
    object_label: Optional[str] = Field(default=None, description="UI-имя объекта")
    stage_a_label: Optional[str] = Field(default=None, description="UI-имя стадии A")
    stage_b_label: Optional[str] = Field(default=None, description="UI-имя стадии B")
    note: Optional[str] = Field(default=None, description="Свободная пометка")


@router.get("/saved-config")
async def get_saved_config_endpoint():
    """Прочитать «сохранённую конфигурацию» Stage Comparison.

    Используется UI для кнопки «Применить сохранённую конфигурацию»:
    клик → fetch → автозаполнение stage_a_path/stage_b_path в форме
    создания сессии.

    Возвращает ``{"saved": false}`` если конфиг ещё не сохранён, иначе
    объект с полями stage_a_path / stage_b_path / object_label /
    stage_a_label / stage_b_label / saved_at / note.
    """
    cfg = saved_config_mod.load_saved_config()
    if cfg is None:
        return {"saved": False}
    return {"saved": True, **cfg}


@router.put("/saved-config")
async def put_saved_config_endpoint(req: SaveConfigRequest):
    """Сохранить «сохранённую конфигурацию» (перезаписать существующую).

    Проверяет allowlist для stage_a_path / stage_b_path, чтобы кто-то не
    закрепил ссылку на чужую папку вне разрешённого root'а.
    """
    try:
        store.assert_path_in_allowlist(req.stage_a_path)
        store.assert_path_in_allowlist(req.stage_b_path)
    except PermissionError as exc:
        raise HTTPException(403, str(exc)) from exc
    try:
        saved = saved_config_mod.save_saved_config(
            stage_a_path=req.stage_a_path,
            stage_b_path=req.stage_b_path,
            object_label=req.object_label,
            stage_a_label=req.stage_a_label,
            stage_b_label=req.stage_b_label,
            note=req.note,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except OSError as exc:
        raise HTTPException(500, f"Ошибка записи saved-config: {exc}") from exc
    return {"saved": True, **saved}


@router.delete("/saved-config")
async def delete_saved_config_endpoint():
    """Удалить «сохранённую конфигурацию» (UI скроет/задизейблит кнопку)."""
    deleted = saved_config_mod.clear_saved_config()
    return {"deleted": deleted}


# ─── Canonical configuration (UX-обёртка над saved-config) ───────────────
# «Каноничная конфигурация» = одна актуальная рабочая конфигурация объекта:
# пара stage_a/stage_b, canonical_session_id и компактный summary пар (без
# артефактов анализа). Старый saved-config endpoint оставлен для обратной
# совместимости (одни пути), но обычный UI работает через canonical-config.


def _canonical_payload(cfg: Optional[dict]) -> dict:
    """Привести cfg к payload для UI с разрешённой каноничной сессией.

    Если canonical_session_id указывает на несуществующую сессию,
    возвращается ``canonical_session_available=false`` — UI покажет
    предупреждение «Каноничная конфигурация недоступна».
    """
    if cfg is None:
        return {"saved": False}
    sid = (cfg.get("canonical_session_id") or "").strip() or None
    session_available = False
    if sid:
        try:
            session_available = store.get_session(sid) is not None
        except Exception:  # noqa: BLE001
            session_available = False
    return {
        "saved": True,
        "canonical_session_available": session_available if sid else None,
        **cfg,
    }


@router.get("/canonical-config")
async def get_canonical_config():
    """Каноничная конфигурация объекта (одна актуальная).

    Возвращает ``{"saved": false}`` если ничего не сохранено. Иначе
    включает все поля saved-config + ``canonical_session_available``
    (bool, есть ли canonical_session_id физически на диске).
    """
    return _canonical_payload(saved_config_mod.load_saved_config())


@router.post("/sessions/{session_id}/save-canonical")
async def save_session_as_canonical(session_id: str, req: Optional[dict] = None):
    """Сохранить текущую сессию как каноничную конфигурацию объекта.

    Тело: ``{object_label?, stage_a_label?, stage_b_label?, note?,
    updated_by?}``. Все поля опциональны — служат для UI badge'а.
    Перезаписывает предыдущую каноничную конфигурацию (история не
    ведётся).
    """
    sess = store.get_session(session_id)
    if sess is None:
        raise HTTPException(404, f"Сессия {session_id} не найдена")
    stage_a_path = (sess.get("stage_a_path") or "").strip()
    stage_b_path = (sess.get("stage_b_path") or "").strip()
    if not stage_a_path or not stage_b_path:
        raise HTTPException(400, "Сессия не содержит stage_a_path/stage_b_path")
    try:
        store.assert_path_in_allowlist(stage_a_path)
        store.assert_path_in_allowlist(stage_b_path)
    except PermissionError as exc:
        raise HTTPException(403, str(exc)) from exc

    body = req if isinstance(req, dict) else {}
    try:
        saved = saved_config_mod.save_saved_config(
            stage_a_path=stage_a_path,
            stage_b_path=stage_b_path,
            object_label=body.get("object_label"),
            stage_a_label=body.get("stage_a_label"),
            stage_b_label=body.get("stage_b_label"),
            note=body.get("note"),
            canonical_session_id=session_id,
            pairs=sess.get("pairs") or [],
            updated_by=body.get("updated_by"),
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except OSError as exc:
        raise HTTPException(500, f"Ошибка записи canonical-config: {exc}") from exc
    return _canonical_payload(saved)


@router.get("/canonical-config/open")
async def open_canonical_config():
    """Получить полные данные каноничной сессии (для автозагрузки UI).

    Контракт:
      * ``saved=false`` — ничего не сохранено.
      * ``canonical_session_available=false`` — canonical_session_id есть,
        но сессия не найдена (повреждена / удалена) → UI предложит
        перезапустить сопоставление.
      * иначе — полный объект сессии + ``canonical_config`` (метаданные
        канона) + ``config_hash_current`` (recompute по факту, для
        invalidation analysis artifacts).
    """
    cfg = saved_config_mod.load_saved_config()
    if cfg is None:
        return {"saved": False}
    sid = (cfg.get("canonical_session_id") or "").strip()
    if not sid:
        # legacy v1 config: только пути, сессии нет
        return {
            "saved": True,
            "canonical_session_id": None,
            "canonical_session_available": None,
            "canonical_config": cfg,
        }
    sess = store.get_session(sid)
    if sess is None:
        return {
            "saved": True,
            "canonical_session_id": sid,
            "canonical_session_available": False,
            "canonical_config": cfg,
        }
    # Пересчитываем config_hash по текущему состоянию пар сессии — если он
    # отличается от saved_hash, UI покажет «результаты могут быть устаревшими».
    current_pairs_summary = []
    for idx, p in enumerate(sess.get("pairs") or []):
        if isinstance(p, dict) and p.get("id"):
            current_pairs_summary.append({
                "pair_id": str(p.get("id")),
                "left_filename": ((p.get("left") or {}).get("filename") or None),
                "right_filename": ((p.get("right") or {}).get("filename") or None),
                "left_pdf_path": ((p.get("left") or {}).get("pdf_path") or None),
                "right_pdf_path": ((p.get("right") or {}).get("pdf_path") or None),
                "disabled": str(p.get("status") or "") == "disabled",
                "status": (p.get("status") or None),
                "manual_links_count": len(p.get("links") or []),
                "order": idx + 1,
            })
    current_hash = saved_config_mod._compute_config_hash(current_pairs_summary)
    return {
        "saved": True,
        "canonical_session_id": sid,
        "canonical_session_available": True,
        "config_hash_saved": cfg.get("config_hash"),
        "config_hash_current": current_hash,
        "config_stale": bool(cfg.get("config_hash") and cfg.get("config_hash") != current_hash),
        "canonical_config": cfg,
        "session": sess,
    }


@router.get("/sessions/{session_id}/unmatched")
async def list_unmatched_endpoint(session_id: str):
    try:
        return store.list_unmatched(session_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs")
async def create_pair(session_id: str, req: CreateManualPairRequest):
    try:
        # Allowlist для путей (на всякий случай)
        if req.left_pdf:
            store.assert_path_in_allowlist(req.left_pdf)
        if req.right_pdf:
            store.assert_path_in_allowlist(req.right_pdf)
    except PermissionError as exc:
        raise HTTPException(403, str(exc)) from exc
    try:
        pair = store.create_manual_pair(
            session_id,
            left_pdf=req.left_pdf, right_pdf=req.right_pdf,
            left_md=req.left_md, left_result_json=req.left_result_json,
            right_md=req.right_md, right_result_json=req.right_result_json,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return pair


@router.post("/sessions/{session_id}/pairs/confirm-all")
async def confirm_all_maybe_pairs(session_id: str):
    """Подтвердить все пары со статусом ``maybe`` (массовое «Сопоставить все»).

    Не меняет PDF, alignment, links — только перезаписывает статус.
    """
    try:
        return store.confirm_maybe_pairs(session_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc


@router.put("/sessions/{session_id}/pairs/{pair_id}/match")
async def update_pair_match_endpoint(session_id: str, pair_id: str, req: UpdatePairMatchRequest):
    try:
        if req.right_pdf:
            store.assert_path_in_allowlist(req.right_pdf)
    except PermissionError as exc:
        raise HTTPException(403, str(exc)) from exc
    try:
        return store.update_pair_match(
            session_id, pair_id,
            right_pdf=req.right_pdf, right_md=req.right_md,
            right_result_json=req.right_result_json, status=req.status,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc


@router.delete("/sessions/{session_id}/pairs/{pair_id}")
async def delete_pair_endpoint(session_id: str, pair_id: str, hard: bool = Query(False)):
    try:
        store.delete_pair(session_id, pair_id, hard=hard)
        return {"ok": True, "hard": hard}
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc


class PairOrderRequest(BaseModel):
    pair_ids: list[str] = Field(..., description="Ordered list of pair_ids (top → bottom)")


@router.put("/sessions/{session_id}/pair-order")
async def set_pair_order_endpoint(session_id: str, req: PairOrderRequest):
    """Drag-and-drop reorder пар в session.json → pair_order.

    UI шлёт список pair_id'ов в желаемом порядке (сверху вниз).
    Backend нормализует (отфильтровывает несуществующие, добавляет в
    конец потерянные) и сохраняет в session.json. Возвращает фактический
    порядок после нормализации.
    """
    try:
        new_order = store.set_pair_order(session_id, req.pair_ids or [])
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    return {"ok": True, "pair_order": new_order}


@router.get("/sessions")
async def list_sessions_endpoint():
    return {"sessions": store.list_sessions()}


@router.get("/sessions/{session_id}")
async def get_session_endpoint(session_id: str):
    session = store.get_session(session_id)
    if session is None:
        raise HTTPException(404, "Сессия не найдена")
    return session


@router.get("/sessions/{session_id}/comparison-statuses")
async def get_comparison_statuses_endpoint(session_id: str):
    """Персистентные статусы сравнения по всем парам сессии (read-only).

    Колонка «Сравнение» в UI берёт статус отсюда (источник истины —
    `comparison_result.json` на диске), а не из «активного» unified-job'а.
    Это чинит баг, когда одно-парный fallback/retry затенял полный
    результат сессии и пары показывались как «—» (не запускалось), хотя
    сравнение было выполнено.
    """
    session = store.get_session(session_id)
    if session is None:
        raise HTTPException(404, "Сессия не найдена")
    return {"statuses": enriched_compare_mod.get_session_comparison_statuses(session_id)}


# ─── Pair view ───────────────────────────────────────────────────────────


@router.get("/sessions/{session_id}/pairs/{pair_id}")
async def get_pair(session_id: str, pair_id: str):
    view = store.get_pair_view(session_id, pair_id)
    if view is None:
        raise HTTPException(404, "Пара не найдена или сессия отсутствует")
    return view


@router.get("/sessions/{session_id}/pairs/{pair_id}/page-image")
async def get_page_image(
    session_id: str,
    pair_id: str,
    side: str = Query(..., pattern="^(left|right)$"),
    page: int = Query(1, ge=1),
    target_long_side: int = Query(1400, ge=400, le=3500),
):
    try:
        png = store.render_pdf_page(
            session_id, pair_id, side, page,
            target_long_side=target_long_side,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        logger.exception("page-image render failed")
        raise HTTPException(500, f"Ошибка рендера PDF: {exc}") from exc
    return FileResponse(str(png), media_type="image/png")


@router.get("/sessions/{session_id}/pairs/{pair_id}/block-image")
async def get_block_image(
    session_id: str,
    pair_id: str,
    side: str = Query(..., pattern="^(left|right)$"),
    block_id: str = Query(...),
    target_long_side: int = Query(1200, ge=200, le=3500),
):
    try:
        png = store.render_block_crop(
            session_id, pair_id, side, block_id,
            target_long_side=target_long_side,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        logger.exception("block-image render failed")
        raise HTTPException(500, f"Ошибка кропа блока: {exc}") from exc
    return FileResponse(str(png), media_type="image/png")


# ─── Block links ─────────────────────────────────────────────────────────


@router.delete("/sessions/{session_id}/pairs/{pair_id}/links")
async def remove_link(session_id: str, pair_id: str, req: DeleteLinkRequest):
    try:
        removed = store.delete_link(session_id, pair_id, req.left_block_id, req.right_block_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    return {"removed": removed}


# ─── Page alignment ──────────────────────────────────────────────────────


@router.get("/sessions/{session_id}/pairs/{pair_id}/page-alignment")
async def get_page_alignment(session_id: str, pair_id: str):
    try:
        return store.get_alignment(session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc


@router.put("/sessions/{session_id}/pairs/{pair_id}/page-alignment")
async def put_page_alignment(
    session_id: str, pair_id: str, req: SaveAlignmentRequest,
    force: bool = Query(False, description="Сохранить карту даже при ошибках валидации"),
):
    """Сохранение карты страниц.

    Если есть validation_errors И не передан force=true (ни в query, ни в body) —
    backend НЕ сохраняет карту и возвращает HTTP 422 с {ok:false, validation_errors}.

    Если force=true — карта сохраняется, response: {ok:true, saved_with_warnings:true,
    validation_errors:[...]}.
    """
    try:
        items = [it.model_dump() for it in (req.items or [])]
        # force из body имеет приоритет над query
        effective_force = bool(req.force or force)
        payload = store.save_alignment(session_id, pair_id, items, force=effective_force)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    if not payload.get("ok"):
        # 422 Unprocessable Entity — клиент пусть покажет ошибки в модалке
        raise HTTPException(status_code=422, detail=payload)
    return payload


@router.post("/sessions/{session_id}/pairs/{pair_id}/sheet-matching")
async def run_sheet_matching_endpoint(session_id: str, pair_id: str):
    """Консервативно сопоставить листы по PreparedDocument.

    Никаких LLM, diff/overlay или связей блоков. Если карта уже изменена
    пользователем, результат сохраняется для просмотра, но карта не трогается.
    """
    try:
        return store.run_sheet_matching(session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/sheet-identity")
async def run_sheet_identity_endpoint(session_id: str, pair_id: str):
    """Проверить принятые пары на идентичность без изменения карты и findings."""
    try:
        return store.run_sheet_identity(session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/sheet-alignment")
async def run_sheet_alignment_endpoint(session_id: str, pair_id: str):
    """Построить V3→V2 transform для принятых пар, без diff/findings."""
    try:
        return store.run_sheet_alignment(session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/change-regions-pilot")
async def run_change_regions_pilot_endpoint(session_id: str, pair_id: str):
    """Запустить изолированный пилот 5Б для трёх aligned-пар."""
    try:
        return store.run_change_regions_pilot(session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/change-regions-cleanup-pilot")
async def run_change_regions_cleanup_pilot_endpoint(session_id: str, pair_id: str):
    try:
        return store.run_change_regions_cleanup_pilot(session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/change-groups-pilot")
async def run_change_groups_pilot_endpoint(session_id: str, pair_id: str):
    """5Б.3: построить change groups поверх atomic regions трёх пилотов."""
    try:
        return store.run_change_groups_pilot(session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/change-detection")
async def run_change_detection_endpoint(session_id: str, pair_id: str):
    """5Б.4: цепочка canonical diff → atomic regions → groups для aligned."""
    try:
        return store.run_change_detection(session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/semantic-diff-pilot")
async def run_semantic_diff_pilot_endpoint(session_id: str, pair_id: str):
    """6А: 12 локальных «Было → Стало», без findings и влияния."""
    try:
        return await run_in_threadpool(store.run_semantic_diff_pilot, session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/semantic-diff-v6a1-pilot")
async def run_semantic_diff_v6a1_pilot_endpoint(session_id: str, pair_id: str):
    """6А.1: deterministic table/entity/number context, без LLM/findings."""
    try:
        return await run_in_threadpool(store.run_semantic_diff_v6a1_pilot, session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/semantic-diff-v6a2-mass")
async def run_semantic_diff_v6a2_mass_endpoint(session_id: str, pair_id: str):
    """6А.2: массовый deterministic-анализ всех групп, без LLM/findings."""
    try:
        return await run_in_threadpool(store.run_semantic_diff_v6a2_mass, session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


# ─── Диагностика новой цепочки сравнения (read-only) ──────────────────────
#
# Витрина уже посчитанных артефактов 5Б.4 + 6А.2 для режима
# «Новый алгоритм (диагностика)» во вкладке «Расхождения». Эти endpoint'ы:
#   • ничего не запускают и ничего не записывают;
#   • не читают и не меняют comparison_result.json, findings и решения эксперта;
#   • не вызывают LLM, Vision и OCR; влияние не считают;
#   • не дедуплицируют результат — одинаковые «Было → Стало» видны все.
# Выключается флагом STAGE_COMPARISON_NEW_PIPELINE_DIAGNOSTIC_ENABLED=0.


@router.get("/sessions/{session_id}/pairs/{pair_id}/diagnostic/new-pipeline")
async def get_new_pipeline_diagnostic_endpoint(session_id: str, pair_id: str):
    """Все change groups новой цепочки со смыслом 6А.2, как есть."""
    if not diagnostic_new_pipeline.is_enabled():
        return {
            "available": False,
            "reason": "disabled_by_flag:STAGE_COMPARISON_NEW_PIPELINE_DIAGNOSTIC_ENABLED",
            "items": [], "summary": {},
        }
    try:
        return await run_in_threadpool(store.get_new_pipeline_diagnostic, session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.get("/sessions/{session_id}/pairs/{pair_id}/diagnostic/new-pipeline/crop")
async def get_new_pipeline_crop_endpoint(
    session_id: str,
    pair_id: str,
    left_page: int = Query(..., ge=1),
    right_page: int = Query(..., ge=1),
    group_id: str = Query(...),
    side: str = Query(..., pattern="^(v2|v3|overlay)$"),
    target_long_side: int = Query(1100, ge=300, le=2600),
):
    """Кроп change group: готовый пилотный PNG либо рендер из PDF в память."""
    if not diagnostic_new_pipeline.is_enabled():
        raise HTTPException(403, "disabled_by_flag:STAGE_COMPARISON_NEW_PIPELINE_DIAGNOSTIC_ENABLED")
    try:
        png, source = await run_in_threadpool(
            store.render_new_pipeline_crop, session_id, pair_id,
            left_page=left_page, right_page=right_page, group_id=group_id,
            side=side, target_long_side=target_long_side,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        logger.exception("new-pipeline crop failed")
        raise HTTPException(500, f"Ошибка кропа change group: {exc}") from exc
    headers = {"X-Crop-Source": source, "Cache-Control": "private, max-age=600"}
    if isinstance(png, (bytes, bytearray)):
        return Response(content=bytes(png), media_type="image/png", headers=headers)
    return FileResponse(str(png), media_type="image/png", headers=headers)


# ─── Visual block equivalence recompute (Stage 3B, mark-only, flag-gated) ──
#
# Recompute-only API поверх mark-only прекчека визуальной эквивалентности
# СВЯЗАННЫХ блоков (links.json). Эти endpoint'ы:
#   • НЕ запускают Qwen/Opus/LLM и НЕ трогают основной pipeline;
#   • выполняют только рендер кропов + cv2-сравнение (Stage 2 runner);
#   • не делают реального skip Qwen/MD/Opus (enforced=false, флаги
#     exclude_* информационны);
#   • запуск job защищён флагом
#     STAGE_COMPARISON_VISUAL_BLOCK_EQUIVALENCE_JOBS_ENABLED (default OFF) —
#     при OFF запуск безопасно отклоняется (403); read-only status/list/artifact
#     остаются доступными.

class VisualBlockEquivalenceJobRequest(BaseModel):
    scope: str = "selected"                       # "pair" | "selected" | "session"
    pair_ids: Optional[list[str]] = None
    write_artifact: bool = True                   # писать visual_block_equivalence.json
    write_debug: bool = False                     # default False — не плодить debug PNG


@router.post("/sessions/{session_id}/visual-block-equivalence/jobs")
async def start_visual_block_equivalence_job_endpoint(
    session_id: str, req: VisualBlockEquivalenceJobRequest = VisualBlockEquivalenceJobRequest(),
):
    """Запустить recompute-only job визуальной эквивалентности связанных блоков.

    scope=pair|selected|session. Создаёт background asyncio-job (НЕ блокирует
    event loop), который для каждой пары пересчитывает визуальную эквивалентность
    по `links.json` (рендер кропов + cv2). Qwen/Opus/основной pipeline НЕ
    запускаются. Защищён feature-флагом (default OFF).
    """
    if not vbe_jobs_mod.VisualBlockEquivalenceJobsConfig.from_env().enabled:
        raise HTTPException(
            403,
            "visual_block_equivalence jobs отключены "
            "(STAGE_COMPARISON_VISUAL_BLOCK_EQUIVALENCE_JOBS_ENABLED=false)",
        )
    try:
        job = vbe_jobs_mod.create_visual_block_equivalence_job(
            session_id, scope=req.scope, pair_ids=req.pair_ids,
            write_artifact=req.write_artifact, write_debug=req.write_debug,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    # Фоновый запуск (asyncio task) — прогресс читается GET-эндпоинтом.
    vbe_jobs_mod.start_job_in_background(job["job_id"])
    return vbe_jobs_mod.get_visual_block_equivalence_job(job["job_id"]) or job


@router.get("/sessions/{session_id}/visual-block-equivalence/jobs")
async def list_visual_block_equivalence_jobs_endpoint(session_id: str):
    """Список visual-block-equivalence job'ов сессии (read-only, без флага)."""
    return {"jobs": vbe_jobs_mod.list_visual_block_equivalence_jobs(session_id)}


@router.get("/sessions/{session_id}/visual-block-equivalence/jobs/{job_id}")
async def get_visual_block_equivalence_job_endpoint(session_id: str, job_id: str):
    """Статус visual-block-equivalence job (processed/failed/total + per-pair)."""
    job = vbe_jobs_mod.get_visual_block_equivalence_job(job_id)
    if job is None or job.get("session_id") != session_id:
        raise HTTPException(404, "Job не найден")
    return job


@router.post("/sessions/{session_id}/visual-block-equivalence/jobs/{job_id}/cancel")
async def cancel_visual_block_equivalence_job_endpoint(session_id: str, job_id: str):
    """Отменить visual-block-equivalence job (останавливает дальнейшие пары)."""
    job = vbe_jobs_mod.get_visual_block_equivalence_job(job_id)
    if job is None or job.get("session_id") != session_id:
        raise HTTPException(404, "Job не найден")
    cancelled = vbe_jobs_mod.cancel_visual_block_equivalence_job(job_id)
    return cancelled or job


@router.get("/sessions/{session_id}/pairs/{pair_id}/visual-block-equivalence")
async def get_pair_visual_block_equivalence_endpoint(session_id: str, pair_id: str):
    """Последний `visual_block_equivalence.json` пары (read-only, без флага).

    404 если артефакта нет, 500 если он битый. Путь резолвится только через
    path-helper (произвольный filesystem path от пользователя не принимается).
    """
    path = sc_paths_mod.visual_block_equivalence_report_path(session_id, pair_id)
    if not path.exists():
        raise HTTPException(404, "Артефакт visual_block_equivalence не найден")
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError) as exc:
        raise HTTPException(
            500, f"Битый артефакт visual_block_equivalence: {exc}") from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/page-alignment/insert-blank")
async def insert_blank_alignment(session_id: str, pair_id: str, req: InsertBlankRequest):
    try:
        return store.alignment_insert_blank(session_id, pair_id, req.slot, req.side)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/page-alignment/move")
async def move_alignment(session_id: str, pair_id: str, req: MoveAlignmentRequest):
    try:
        return store.alignment_move(session_id, pair_id, req.slot, req.direction)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/page-alignment/reset")
async def reset_alignment(session_id: str, pair_id: str):
    try:
        return store.alignment_reset(session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/page-alignment/insert-blank-side")
async def insert_blank_alignment_side(session_id: str, pair_id: str, req: InsertBlankSideRequest):
    """Вставить пустой лист только на выбранной стороне (left/right).

    Левая и правая «дорожки» страниц независимы. Другая сторона остаётся на месте,
    slot'ы пересчитываются, links автоматически переоцениваются на stale/cross-page.
    """
    try:
        return store.alignment_insert_blank_side(session_id, pair_id, req.slot, req.side)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/page-alignment/move-page-side")
async def move_page_alignment_side(session_id: str, pair_id: str, req: MovePageSideRequest):
    """Переместить страницу одной стороны вверх/вниз; другая сторона не меняется."""
    try:
        return store.alignment_move_page_side(
            session_id, pair_id, req.slot, req.side, req.direction
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/pairs/{pair_id}/page-alignment/delete-page-side")
async def delete_page_alignment_side(session_id: str, pair_id: str, req: DeletePageSideRequest):
    """Удалить страницу одной стороны в slot'е; другая сторона не меняется."""
    try:
        return store.alignment_delete_page_side(
            session_id, pair_id, req.slot, req.side
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


# ─── Text diff ───────────────────────────────────────────────────────────


@router.get("/sessions/{session_id}/pairs/{pair_id}/text-diff")
async def text_diff_endpoint(session_id: str, pair_id: str):
    """Legacy: построчный difflib MD-файлов. Используется как debug/fallback
    после переезда на семантический text-llm-diff. См. POST/GET .../text-llm-diff."""
    session = store.get_session(session_id)
    if session is None:
        raise HTTPException(404, "Сессия не найдена")
    pair = next((p for p in session.get("pairs") or [] if p.get("id") == pair_id), None)
    if pair is None:
        raise HTTPException(404, "Пара не найдена")
    left_md = (pair.get("left") or {}).get("md_path")
    right_md = (pair.get("right") or {}).get("md_path")
    diff = diff_text.build_text_diff(left_md, right_md)
    diff["left_md_path"] = left_md
    diff["right_md_path"] = right_md
    return diff


# ─── Text LLM diff (Claude Sonnet, explicit user-triggered) ──────────────


class CreateTextLLMJobRequest(BaseModel):
    scope: str = Field(..., pattern="^(pair|session|selected)$")
    pair_id: Optional[str] = None
    pair_ids: Optional[list[str]] = None
    confirm: bool = False
    force: bool = False  # сейчас runner всегда force=True; зарезервировано


@router.get("/sessions/{session_id}/pairs/{pair_id}/text-llm-diff")
async def get_text_llm_diff_endpoint(session_id: str, pair_id: str):
    session = store.get_session(session_id)
    if session is None:
        raise HTTPException(404, "Сессия не найдена")
    pair = next((p for p in session.get("pairs") or [] if p.get("id") == pair_id), None)
    if pair is None:
        raise HTTPException(404, "Пара не найдена")
    existing = text_llm_mod.get_text_llm_diff(session_id, pair_id)
    if existing is None:
        return {
            "status": "not_run",
            "left_md_path": (pair.get("left") or {}).get("md_path"),
            "right_md_path": (pair.get("right") or {}).get("md_path"),
        }
    return existing


class PreflightRequest(BaseModel):
    scope: str = Field("session", pattern="^(pair|session|selected)$")
    pair_id: Optional[str] = None
    pair_ids: Optional[list[str]] = None
    force: bool = False


@router.post("/sessions/{session_id}/text-llm-preflight")
async def text_llm_preflight_session_endpoint(session_id: str, req: PreflightRequest):
    """Pre-run оценка для batch: суммирует runnable пары, агрегирует skipped."""
    try:
        return text_llm_preflight_mod.estimate_session(
            session_id,
            scope=req.scope, pair_id=req.pair_id, pair_ids=req.pair_ids,
            force=bool(req.force),
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/sessions/{session_id}/text-llm-diff-jobs")
async def create_text_llm_job_endpoint(session_id: str, req: CreateTextLLMJobRequest):
    if not req.confirm:
        # Создаём rejected job, чтобы он остался в истории
        job = text_llm_jobs_mod.create_text_llm_job(
            session_id, scope=req.scope, pair_id=req.pair_id,
            pair_ids=req.pair_ids, confirm=False,
        )
        return job
    # Hard-limit enforcement: для batch (session/selected) проверяем агрегатную
    # оценку и блокируем запуск, если выше HARD_*. Per-pair scope разрешаем
    # после подтверждения — там пользователь явно знает, что запускает.
    if req.scope in ("session", "selected"):
        try:
            preflight = text_llm_preflight_mod.estimate_session(
                session_id, scope=req.scope, pair_id=req.pair_id,
                pair_ids=req.pair_ids, force=bool(req.force),
            )
        except KeyError as exc:
            raise HTTPException(404, str(exc)) from exc
        limits = preflight.get("limits") or {}
        if limits.get("cost_hard") or limits.get("duration_hard"):
            raise HTTPException(
                422,
                {
                    "code": "batch_limit_exceeded",
                    "message": "Оценка превышает безопасный лимит. Запустите анализ по отдельным парам.",
                    "preflight": preflight,
                },
            )
    try:
        job = text_llm_jobs_mod.create_text_llm_job(
            session_id, scope=req.scope, pair_id=req.pair_id,
            pair_ids=req.pair_ids, confirm=True,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    if job.get("status") == "queued":
        text_llm_jobs_mod.start_job_in_background(session_id, job["id"])
    return job


@router.get("/sessions/{session_id}/text-llm-diff-jobs/{job_id}")
async def get_text_llm_job_endpoint(session_id: str, job_id: str):
    job = text_llm_jobs_mod.get_job(session_id, job_id)
    if job is None:
        raise HTTPException(404, "Job не найден")
    return job


@router.post("/sessions/{session_id}/text-llm-diff-jobs/{job_id}/cancel")
async def cancel_text_llm_job_endpoint(session_id: str, job_id: str):
    job = text_llm_jobs_mod.cancel_job(session_id, job_id)
    if job is None:
        raise HTTPException(404, "Job не найден")
    return job


@router.get("/sessions/{session_id}/text-llm-diff-flat")
async def text_llm_diff_flat_endpoint(session_id: str):
    """Плоский список текстовых смысловых изменений по всей сессии.

    Агрегирует все `text_llm_diff.json` пар сессии, привязывает каждое
    изменение к PDF-странице/листу/alignment-slot через `text_location`.
    Чтения только — никаких LLM-вызовов, никаких записей на диск.
    """
    try:
        return text_llm_flat_mod.build_flat(session_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc


@router.get("/sessions/{session_id}/text-llm-config")
async def text_llm_config_endpoint(session_id: str):
    """Лёгкая инфо-ручка для UI: enabled / provider / model / availability."""
    cfg = text_llm_provider_mod.load_config()
    provider, _ = text_llm_provider_mod.resolve_provider(cfg)
    info = {
        "enabled": cfg.enabled,
        "provider": cfg.provider,
        "model": cfg.model,
        "max_chars": cfg.max_chars,
        "timeout_sec": cfg.timeout_sec,
        "available": False,
        "reason": None,
    }
    if provider is not None:
        ok, reason = provider.check_availability()
        info["available"] = bool(ok)
        info["reason"] = reason
    elif not cfg.enabled:
        info["reason"] = "disabled_via_env"
    else:
        info["reason"] = f"unknown_provider:{cfg.provider}"
    return info


# ─── MD enrichment (Qwen image descriptions для enriched MD) ─────────────




def _md_enrichment_pair_payload(
    session_id: str, pair_id: str, pair: dict,
) -> dict:
    left = pair.get("left") or {}
    right = pair.get("right") or {}
    left_summary = md_enrichment_mod.read_summary_only(session_id, pair_id, "left")
    right_summary = md_enrichment_mod.read_summary_only(session_id, pair_id, "right")
    left_summary["md_path"] = left.get("md_path")
    right_summary["md_path"] = right.get("md_path")

    def _side_ready(side_summary: dict) -> bool:
        if (side_summary.get("status") or "") != "done":
            return False
        if int(side_summary.get("errors") or 0) > 0:
            return False
        if int(side_summary.get("pending") or 0) > 0:
            return False
        enriched = side_summary.get("enriched_md_path")
        if not enriched:
            return False
        try:
            from pathlib import Path as _P
            return _P(enriched).exists()
        except Exception:  # noqa: BLE001
            return False

    return {
        "pair_id": pair_id,
        "left": left_summary,
        "right": right_summary,
        "ready_for_unified_analysis": _side_ready(left_summary) and _side_ready(right_summary),
    }


@router.get("/sessions/{session_id}/pairs/{pair_id}/md-enrichment")
async def get_md_enrichment_endpoint(session_id: str, pair_id: str):
    session = store.get_session(session_id)
    if session is None:
        raise HTTPException(404, "Сессия не найдена")
    pair = next((p for p in session.get("pairs") or [] if p.get("id") == pair_id), None)
    if pair is None:
        raise HTTPException(404, "Пара не найдена")
    return _md_enrichment_pair_payload(session_id, pair_id, pair)


_FAILED_BLOCK_STATUSES = {"error", "no_image", "render_failed"}


@router.get("/sessions/{session_id}/pairs/{pair_id}/failed-blocks")
async def get_failed_blocks_endpoint(session_id: str, pair_id: str):
    """Список упавших image-блоков пары (status ∈ error/no_image/render_failed).

    Используется поповером на цифре «упало» в таблице пар: оператор кликает по
    числу, видит какие именно блоки не распознались, и может перейти к блоку на
    вкладке «Связь блоков». Агрегаты (block_metrics) этот endpoint не меняет —
    только читает per-block items[] из <side>_image_descriptions.json.
    """
    session = store.get_session(session_id)
    if session is None:
        raise HTTPException(404, "Сессия не найдена")
    pair = next((p for p in session.get("pairs") or [] if p.get("id") == pair_id), None)
    if pair is None:
        raise HTTPException(404, "Пара не найдена")

    blocks: list[dict] = []
    for side in ("left", "right"):
        p = sc_paths_mod.text_enrichment_descriptions_path(session_id, pair_id, side)
        if not p.exists():
            continue
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError, ValueError):
            continue
        for it in data.get("items") or []:
            if not isinstance(it, dict):
                continue
            if (it.get("status") or "").lower() not in _FAILED_BLOCK_STATUSES:
                continue
            blocks.append({
                "side": side,
                "order": it.get("order"),
                "page": it.get("page"),
                "image_order_on_page": it.get("image_order_on_page"),
                "md_block_id": it.get("md_block_id"),
                "side_block_id": it.get("side_block_id"),
                "status": it.get("status"),
                "error": it.get("error"),
                "parse_error_detail": it.get("parse_error_detail"),
                "block_type": it.get("block_type"),
            })

    blocks.sort(key=lambda b: (
        0 if b["side"] == "left" else 1,
        b.get("page") or 0,
        b.get("order") or 0,
    ))
    return {"pair_id": pair_id, "blocks": blocks, "count": len(blocks)}


@router.get("/sessions/{session_id}/pairs/{pair_id}/enriched-md")
async def get_enriched_md_content_endpoint(
    session_id: str, pair_id: str,
    side: str = Query(..., pattern="^(left|right)$"),
):
    """Содержимое `<side>_enriched.md` для просмотра в UI.

    Используется переключателем PDF ↔ MD в двухпанельном вьюере: вместо
    рендеренных страниц PDF показывается текст enriched MD соответствующей
    стороны. Если enrichment ещё не запускался — `exists=false`, content пуст.
    """
    session = store.get_session(session_id)
    if session is None:
        raise HTTPException(404, "Сессия не найдена")
    pair = next((p for p in session.get("pairs") or [] if p.get("id") == pair_id), None)
    if pair is None:
        raise HTTPException(404, "Пара не найдена")
    md_path = sc_paths_mod.text_enrichment_md_path(session_id, pair_id, side)
    exists = md_path.exists()
    content = ""
    if exists:
        try:
            content = md_path.read_text(encoding="utf-8", errors="replace")
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(500, f"Не удалось прочитать enriched MD: {exc}") from exc
    return {
        "side": side,
        "path": str(md_path),
        "filename": md_path.name,
        "exists": exists,
        "char_count": len(content),
        "content": content,
    }








class ClearAnalysisRequest(BaseModel):
    pair_ids: list[str] = Field(default_factory=list)
    clear_findings: bool = True
    clear_review: bool = True
    clear_enrichment: bool = False


@router.post("/sessions/{session_id}/pairs/clear-analysis")
async def clear_pairs_analysis_endpoint(session_id: str, req: ClearAnalysisRequest):
    """Очистить найденные расхождения и/или ручные отметки проверки по
    ВЫБРАННЫМ парам (backup → удаление). Возвращает per-pair backup_paths /
    deleted_files / skipped.

    Пары с running/queued job пропускаются (warning «pair has running job,
    cancel first»). НЕ удаляются: исходные PDF, OCR result.json,
    page_enriched.json/tiles (large-sheet Qwen), left/right_enriched.md,
    Qwen image-cache.
    """
    if not req.pair_ids:
        raise HTTPException(400, "pair_ids required")
    try:
        result = await run_in_threadpool(
            clear_analysis_mod.clear_pairs_analysis,
            session_id, req.pair_ids,
            clear_findings=bool(req.clear_findings),
            clear_review=bool(req.clear_review),
            clear_enrichment=bool(req.clear_enrichment),
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    return result


class OpusOnlyRequest(BaseModel):
    pair_ids: list[str] = Field(default_factory=list)
    force: bool = True
    backup_existing: bool = True
    clear_comparison_result: bool = False


@router.post("/sessions/{session_id}/pairs/opus-only")
async def opus_only_endpoint(session_id: str, req: OpusOnlyRequest):
    """Запустить ТОЛЬКО Opus / unified-analysis по выбранным парам, без Qwen.

    Читает готовые `left_enriched.md` / `right_enriched.md` и (пере)создаёт
    `comparison_result.json`. Qwen / large-sheet / md-enrichment НЕ запускаются,
    enriched MD не пересобирается, page_enriched.json / OCR / PDF не трогаются.

    Пары без enriched MD пропускаются (`missing_enriched_md`); с running job —
    `running_job`; превышающие лимит Opus — `too_large` (для них используйте
    per-pair fallback-бейдж). При `backup_existing`/`clear_comparison_result`
    текущий `comparison_result.json` бэкапится (и опц. удаляется);
    `expert_review` / `v2_review_status` не трогаются.
    """
    if not req.pair_ids:
        raise HTTPException(400, "pair_ids required")
    try:
        prep = await run_in_threadpool(
            opus_only_mod.prepare_opus_only,
            session_id, req.pair_ids,
            backup_existing=bool(req.backup_existing),
            clear_comparison_result=bool(req.clear_comparison_result),
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    eligible = prep.get("eligible") or []
    if not eligible:
        return {"ok": True, "job_id": None, "started_pairs": [],
                "skipped": prep.get("skipped") or [], "backups": prep.get("backups") or {}}
    # Только Opus: force_enrichment=False (Qwen/enrichment не трогаем),
    # force_compare=True (пере-сравнить готовые enriched MD).
    job = unified_jobs_mod.create_unified_job(
        session_id, scope="selected", pair_ids=eligible,
        force_enrichment=False, force_compare=True, force_fallback=False,
        confirm=True, skip_ineligible=False)
    if job.get("status") == "queued":
        unified_jobs_mod.start_job_in_background(session_id, job["id"])
    return {"ok": True, "job_id": job.get("id"), "started_pairs": eligible,
            "skipped": prep.get("skipped") or [], "backups": prep.get("backups") or {}}






class UnifiedAnalysisPreflightRequest(BaseModel):
    force_enrichment: bool = False
    force_compare: bool = False


class UnifiedAnalysisRunRequest(BaseModel):
    force_enrichment: bool = False
    force_compare: bool = False
    confirm: bool = False
    # Профиль анализа: "default" (быстрый) | "rich_grsh" (глубокий ГРЩ). None →
    # env-профиль. rich_grsh = per-run override без правки .env; для эталонного
    # результата комбинировать с force_enrichment=true.
    analysis_profile: Optional[str] = None
    # Разрешить перезаписать сохранённый rich_grsh результат быстрым прогоном.
    allow_profile_downgrade: bool = False


class CreateUnifiedAnalysisJobRequest(BaseModel):
    scope: str = Field(..., pattern="^(pair|session|selected)$")
    pair_id: Optional[str] = None
    pair_ids: Optional[list[str]] = None
    force_enrichment: bool = False
    force_compare: bool = False
    confirm: bool = False
    # Pre-flight фильтр: пропустить пары, у которых enriched MD не готов /
    # суммарный размер превышает лимит / comparison уже done (если
    # force_compare=false). Используется session-level Opus batch с этапа
    # «Загрузка документации», чтобы не звать Qwen и не запускать too_large.
    skip_ineligible: bool = False
    # Явный per-pair override: too_large прогнать через evidence_first_s2_fallback
    # даже при выключенном глобальном флаге. UI-кнопка «запустить fallback».
    force_fallback: bool = False
    # Профиль анализа для batch. Default-массовый прогон НЕ должен включать
    # rich-флаги: default остаётся default. rich_grsh — только явно / selected.
    analysis_profile: Optional[str] = None
    allow_profile_downgrade: bool = False


class UnifiedAnalysisBatchPreflightRequest(BaseModel):
    scope: str = Field(..., pattern="^(session|selected)$")
    pair_ids: Optional[list[str]] = None
    force_compare: bool = False


@router.get("/sessions/{session_id}/pairs/{pair_id}/unified-analysis")
async def get_unified_pair_status_endpoint(session_id: str, pair_id: str):
    """Статус unified-анализа для пары: enriched MD / comparison_result / counts."""
    session = store.get_session(session_id)
    if session is None:
        raise HTTPException(404, "Сессия не найдена")
    pair = next((p for p in session.get("pairs") or [] if p.get("id") == pair_id), None)
    if pair is None:
        raise HTTPException(404, "Пара не найдена")

    enriched_status = enriched_compare_mod.enriched_md_status(session_id, pair_id)
    comp = enriched_compare_mod.get_comparison_result(session_id, pair_id)
    left_sum = md_enrichment_mod.read_summary_only(session_id, pair_id, "left")
    right_sum = md_enrichment_mod.read_summary_only(session_id, pair_id, "right")
    return {
        "pair_id": pair_id,
        "pair_label": f"{(pair.get('left') or {}).get('filename') or '—'} ↔ "
                       f"{(pair.get('right') or {}).get('filename') or '—'}",
        "enrichment": {
            "left": left_sum,
            "right": right_sum,
            "enriched_md_ready": enriched_status.get("ready"),
            "enriched_md_total_chars": enriched_status.get("total_chars"),
            "left_path": enriched_status.get("left", {}).get("path"),
            "right_path": enriched_status.get("right", {}).get("path"),
        },
        "comparison": {
            "status": (comp or {}).get("status") or "not_run",
            "changes_count": len((comp or {}).get("changes") or []),
            "updated_at": (comp or {}).get("updated_at"),
            "warnings": (comp or {}).get("warnings") or [],
            "error": (comp or {}).get("error"),
            "result_path": str(sc_paths_mod.enriched_comparison_result_path(session_id, pair_id))
                            if comp else None,
        },
    }


@router.post("/sessions/{session_id}/pairs/{pair_id}/unified-analysis/preflight")
async def unified_pair_preflight_endpoint(
    session_id: str, pair_id: str, req: UnifiedAnalysisPreflightRequest,
):
    """Pre-run оценка: image-блоки, cache hits, Qwen/Opus availability, длительность."""
    try:
        pre = unified_analysis_mod.preflight_pair(
            session_id, pair_id,
            force_enrichment=req.force_enrichment,
            force_compare=req.force_compare,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    return pre.as_dict()


@router.post("/sessions/{session_id}/pairs/{pair_id}/unified-analysis")
async def unified_pair_run_endpoint(
    session_id: str, pair_id: str, req: UnifiedAnalysisRunRequest,
):
    """Запустить unified-анализ для одной пары (синхронно).

    Без confirm=true возвращает preflight без запуска моделей. Это защита
    от случайного live-вызова Qwen/Opus.
    """
    if not req.confirm:
        try:
            pre = unified_analysis_mod.preflight_pair(
                session_id, pair_id,
                force_enrichment=req.force_enrichment,
                force_compare=req.force_compare,
            )
        except KeyError as exc:
            raise HTTPException(404, str(exc)) from exc
        return {
            "ok": False,
            "status": "rejected_no_confirm",
            "preflight": pre.as_dict(),
        }
    try:
        res = await unified_analysis_mod.run_pair(
            session_id, pair_id,
            force_enrichment=req.force_enrichment,
            force_compare=req.force_compare,
            analysis_profile=req.analysis_profile,
            allow_profile_downgrade=req.allow_profile_downgrade,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    # После одиночного запуска обновляем unified findings.
    try:
        unified_findings_mod.rebuild_unified_findings(session_id)
    except Exception:  # noqa: BLE001
        logger.exception("unified_findings rebuild failed")
    return {"ok": True, "result": res.as_dict()}


@router.post("/sessions/{session_id}/unified-analysis-jobs")
async def create_unified_job_endpoint(
    session_id: str, req: CreateUnifiedAnalysisJobRequest,
):
    """Batch unified job. Без confirm=true создаём rejected (для истории)."""
    try:
        job = unified_jobs_mod.create_unified_job(
            session_id,
            scope=req.scope, pair_id=req.pair_id, pair_ids=req.pair_ids,
            force_enrichment=req.force_enrichment,
            force_compare=req.force_compare,
            force_fallback=req.force_fallback,
            analysis_profile=req.analysis_profile,
            allow_profile_downgrade=req.allow_profile_downgrade,
            confirm=req.confirm,
            skip_ineligible=req.skip_ineligible,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    if job.get("status") == "queued":
        unified_jobs_mod.start_job_in_background(session_id, job["id"])
    return unified_jobs_mod.get_job_with_progress(session_id, job["id"]) or job


@router.post("/sessions/{session_id}/unified-analysis-jobs/preflight")
async def unified_batch_preflight_endpoint(
    session_id: str, req: UnifiedAnalysisBatchPreflightRequest,
):
    """Dry-run сводка перед запуском Opus batch (без запуска моделей)."""
    try:
        return unified_jobs_mod.preflight_session_for_batch(
            session_id,
            scope=req.scope,
            pair_ids=req.pair_ids,
            force_compare=bool(req.force_compare),
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc


@router.get("/sessions/{session_id}/unified-analysis-jobs/active")
async def get_active_unified_job_endpoint(session_id: str):
    """Самая релевантная unified-job сессии — для resume в UI.

    Возвращает {"job": <job_with_progress>} или {"job": null}.
    """
    session = store.get_session(session_id)
    if session is None:
        raise HTTPException(404, "Сессия не найдена")
    job = unified_jobs_mod.find_active_session_job(session_id)
    return {"job": job}


@router.get("/sessions/{session_id}/unified-analysis-jobs/{job_id}")
async def get_unified_job_endpoint(session_id: str, job_id: str):
    job = unified_jobs_mod.get_job_with_progress(session_id, job_id)
    if job is None:
        raise HTTPException(404, "Job не найден")
    return job


@router.post("/sessions/{session_id}/unified-analysis-jobs/{job_id}/cancel")
async def cancel_unified_job_endpoint(session_id: str, job_id: str):
    job = unified_jobs_mod.cancel_job(session_id, job_id)
    if job is None:
        raise HTTPException(404, "Job не найден")
    return unified_jobs_mod.get_job_with_progress(session_id, job_id) or job


@router.get("/sessions/{session_id}/unified-diff-flat")
async def unified_diff_flat_endpoint(session_id: str, pair_id: Optional[str] = None):
    """Единый плоский список unified findings по всей сессии.

    Read-only — никаких LLM/моделей. Использует существующие
    comparison_result.json по парам.

    `pair_id`: query-параметр. Если задан, summary и items считаются только
    по этой паре. UI вкладки «Расхождения» вызывает endpoint с pair_id
    активной PDF-пары, чтобы не подмешивать findings других пар. Без
    параметра — поведение по всей сессии (legacy / «Показать все пары»).
    """
    try:
        return unified_findings_mod.build_unified_flat(session_id, pair_id=pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc


_SC_UNIFIED_SOURCE_LABELS = {
    "text": "Текст",
    "image_enrichment": "Описание изобр.",
    "scheme_analysis": "Схема",
    "table": "Таблица",
    "stamp": "Штамп",
    "mixed": "Текст + изобр.",
}

# Метки источника/направления для grouped-отчёта — зеркало фронтенда
# (вкладка «Отчёт»): источник схлопнут до «текст» / «изображение».
_SC_REPORT_SOURCE_LABELS = {
    "text": "текст",
    "image_enrichment": "изображение",
    "scheme_analysis": "изображение",
    "table": "текст",
    "stamp": "текст",
    "mixed": "изображение",
}
_SC_REPORT_DIRECTION_LABELS = {
    "complication": "усложнение",
    "simplification": "упрощение",
    "neutral": "нейтрально",
    "unknown": "—",
}
_SC_REPORT_SEV_RANK = {"high": 0, "medium": 1, "low": 2, "unknown": 3}


def _sc_page_str(page) -> str:
    """Нормализует page (int | list | None) в строку для ячейки «Место»."""
    if isinstance(page, list):
        return ", ".join(str(p) for p in page if p is not None)
    if page is None:
        return ""
    return str(page)


def _sc_split_lines(val) -> str:
    """Зеркало фронтендового scUnifiedLines: режет значение расхождения по ';'
    на отдельные строки (точка с запятой сохраняется, кроме последней)."""
    if val is None:
        return ""
    text = str(val).strip()
    if not text:
        return ""
    if ";" not in text:
        return text
    parts = [s.strip() for s in text.split(";") if s.strip()]
    if not parts:
        return text
    return "\n".join(
        (s + ";") if i < len(parts) - 1 else s for i, s in enumerate(parts)
    )


def _sc_report_page_sort_key(item: dict) -> int:
    page = item.get("page")
    if isinstance(page, list):
        return (page[0] if page and page[0] is not None else 0) or 0
    return (page if page is not None else 0) or 0


def _build_grouped_comparison_workbook(items: list[dict], pair_order: Optional[list[str]] = None):
    """Собирает XLSX в виде, повторяющем вкладку «Отчёт»: каждая PDF-пара —
    сворачиваемый раздел (Excel-группировка, открывается «плюсиком»), внутри —
    таблица сравнения с колонками №, Место, Изменение, Было, Стало, Влияние.

    `pair_order` — порядок PDF-пар как в сессии (так же, как вкладка «Отчёт»
    обходит `scSession.pairs`). Разделы выводятся в этом порядке; пары вне
    списка идут следом по порядку первого появления в `items`.

    Возвращает openpyxl Workbook.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Сравнение стадий"
    # «+» (toggle) должен стоять у строки-заголовка пары, а она ВЫШЕ группы
    # деталей — поэтому summary-строка сверху.
    ws.sheet_properties.outlinePr.summaryBelow = False

    headers = ["№", "Место", "Изменение", "Было", "Стало", "Влияние"]
    widths = [16, 26, 46, 34, 34, 30]
    n_cols = len(headers)
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sev_fill = {
        "high":   PatternFill("solid", fgColor="FEE2E2"),
        "medium": PatternFill("solid", fgColor="FEF3C7"),
        "low":    PatternFill("solid", fgColor="DBEAFE"),
    }
    was_fill = PatternFill("solid", fgColor="FEF2F2")
    became_fill = PatternFill("solid", fgColor="ECFDF5")
    impact_fill = PatternFill("solid", fgColor="FEFCE8")
    pair_fill = PatternFill("solid", fgColor="EFF6FF")
    subhdr_fill = PatternFill("solid", fgColor="F3F4F6")

    # Группируем по паре. Порядок разделов = порядок пар сессии (как вкладка
    # «Отчёт»); пары вне `pair_order` — по первому появлению в items.
    appearance: list[str] = []
    groups: dict[str, dict] = {}
    for it in items:
        pid = str(it.get("pair_id") or "")
        if pid not in groups:
            groups[pid] = {"label": (it.get("pair_label") or pid), "items": []}
            appearance.append(pid)
        groups[pid]["items"].append(it)

    order: list[str] = []
    seen: set[str] = set()
    for pid in (pair_order or []):
        pid = str(pid)
        if pid in groups and pid not in seen:
            order.append(pid)
            seen.add(pid)
    for pid in appearance:
        if pid not in seen:
            order.append(pid)
            seen.add(pid)

    r = 0
    for pid in order:
        g = groups[pid]
        rows = sorted(
            g["items"],
            key=lambda x: (
                _SC_REPORT_SEV_RANK.get(str(x.get("severity") or "").lower(), 3),
                _sc_report_page_sort_key(x),
            ),
        )

        # ── строка-заголовок пары (summary/toggle, outline level 0) ──
        r += 1
        hdr_row = r
        head_cell = ws.cell(
            row=hdr_row, column=1,
            value=f"{g['label']}    —    согласовано: {len(rows)}",
        )
        ws.merge_cells(start_row=hdr_row, start_column=1, end_row=hdr_row, end_column=n_cols)
        head_cell.font = Font(bold=True, size=11, color="1E3A8A")
        head_cell.alignment = Alignment(vertical="center", horizontal="left")
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=hdr_row, column=c)
            cell.fill = pair_fill
            cell.border = border
        # Стартуем со свёрнутого раздела — раскрывается «плюсиком».
        ws.row_dimensions[hdr_row].collapsed = True

        # ── под-заголовок таблицы (outline level 1) ──
        r += 1
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = subhdr_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = border
        ws.row_dimensions[r].outline_level = 1
        ws.row_dimensions[r].hidden = True

        # ── строки сравнения (outline level 1) ──
        for i, it in enumerate(rows, start=1):
            r += 1
            sev = str(it.get("severity") or "unknown").lower()

            no_lines = [f"№{i}", sev or "—"]
            direction = str(it.get("change_direction") or "").lower()
            if direction and direction != "unknown":
                no_lines.append(_SC_REPORT_DIRECTION_LABELS.get(direction, direction))
            src = str(it.get("source_layer") or "")
            if src:
                no_lines.append(_SC_REPORT_SOURCE_LABELS.get(src, src))
            no_text = "\n".join(no_lines)

            place_parts = []
            if it.get("sheet"):
                place_parts.append(str(it.get("sheet")))
            pstr = _sc_page_str(it.get("page"))
            if pstr:
                place_parts.append(f"стр. PDF: {pstr}")
            place_text = "\n".join(place_parts)

            change_parts = []
            if it.get("title"):
                change_parts.append(str(it.get("title")))
            if it.get("summary"):
                change_parts.append(str(it.get("summary")))
            change_text = "\n".join(change_parts) or "—"

            old_value = it.get("old_value") or ((it.get("evidence_left") or {}).get("quote")) or ""
            new_value = it.get("new_value") or ((it.get("evidence_right") or {}).get("quote")) or ""
            was_text = _sc_split_lines(old_value) or "—"
            became_text = _sc_split_lines(new_value) or "—"
            impact_text = str(it.get("construction_impact") or "").strip() or "—"

            values = [no_text, place_text, change_text, was_text, became_text, impact_text]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border
                cell.font = Font(size=10)
            if sev in sev_fill:
                ws.cell(row=r, column=1).fill = sev_fill[sev]
            ws.cell(row=r, column=4).fill = was_fill
            ws.cell(row=r, column=5).fill = became_fill
            ws.cell(row=r, column=6).fill = impact_fill
            ws.row_dimensions[r].outline_level = 1
            ws.row_dimensions[r].hidden = True

        # пустая строка-разделитель между парами (level 0)
        r += 1

    return wb


@router.get("/sessions/{session_id}/unified-diff-flat/export.xlsx")
async def unified_diff_flat_export_xlsx(
    session_id: str,
    pair_id: Optional[str] = None,
    accepted_only: bool = False,
    grouped: bool = False,
):
    """Экспорт таблицы расхождений в Excel (xlsx).

    Формат соответствует UI-таблице на вкладке «Расхождения»:
    №, Место (лист/стр.PDF/PDF-пара), Важность, Изменение (title + summary),
    Было, Стало, Влияние, Стоимость, Источник, На ручную проверку.

    Если задан `pair_id` — выгружаем только эту PDF-пару (соответствует
    текущему scope в UI). Без параметра — все пары сессии.

    `accepted_only=true` — оставить только расхождения, согласованные экспертом
    (`expert_review.json`, decision=accepted). Используется вкладкой «Отчёт»:
    один XLSX со всеми согласованными изменениями по всем парам сразу.

    `grouped=true` — выгрузка в виде вкладки «Отчёт»: каждая PDF-пара —
    сворачиваемый раздел (Excel-группировка, открывается «плюсиком»), внутри —
    таблица сравнения с колонками №, Место, Изменение, Было, Стало, Влияние.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    try:
        flat = unified_findings_mod.build_unified_flat(session_id, pair_id=pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc

    items = flat.get("items") or []

    if accepted_only:
        decisions = (expert_review_mod.load(session_id) or {}).get("decisions") or {}
        accepted_keys = {
            str(key)
            for key, entry in decisions.items()
            if isinstance(entry, dict) and (entry.get("decision") or "").lower() == "accepted"
        }

        def _it_accepted(it: dict) -> bool:
            pid = str(it.get("pair_id") or "")
            # Согласовано в классическом виде «Расхождения» (V1)…
            if expert_review_mod.make_key(pid, str(it.get("id") or "")) in accepted_keys:
                return True
            # …или в виде V2 (под двойником v2_<sha1(pid::raw_id)>). Учитываем оба,
            # чтобы экспорт совпадал с экраном «Отчёт» и не терял V2-only находки.
            v2_key = expert_review_mod.make_key(pid, v2_review_mod.make_v2_id(pid, it))
            return v2_key in accepted_keys

        items = [it for it in items if isinstance(it, dict) and _it_accepted(it)]

    if grouped:
        # Порядок разделов как на вкладке «Отчёт» — по порядку пар сессии
        # (scSession.pairs), а не по алфавиту pair_label из build_unified_flat.
        pair_order: list[str] = []
        try:
            _session = store.get_session(session_id)
            for _p in ((_session or {}).get("pairs") or []):
                if not isinstance(_p, dict) or _p.get("status") == "disabled":
                    continue
                _pid = str(_p.get("id") or "")
                if _pid:
                    pair_order.append(_pid)
        except Exception:  # noqa: BLE001
            pair_order = []
        wb = _build_grouped_comparison_workbook(items, pair_order=pair_order)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_sid = re.sub(r"[^A-Za-z0-9_.-]", "_", session_id)[:64] or "session"
        scope = "report" if accepted_only else ("pair" if pair_id else "all")
        fname = f"stage_comparison_{safe_sid}_{scope}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Расхождения"

    header = [
        "№", "Лист", "Стр. PDF", "PDF-пара", "Важность",
        "Изменение", "Описание",
        "Было", "Стало",
        "Влияние", "Стоимость",
        "Источник", "На ручную проверку",
    ]
    ws.append(header)
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border

    sev_fill = {
        "high":   PatternFill("solid", fgColor="FEE2E2"),
        "medium": PatternFill("solid", fgColor="FEF3C7"),
        "low":    PatternFill("solid", fgColor="DBEAFE"),
    }
    was_fill = PatternFill("solid", fgColor="FEF2F2")
    became_fill = PatternFill("solid", fgColor="ECFDF5")
    impact_fill = PatternFill("solid", fgColor="FEFCE8")

    for i, it in enumerate(items, start=1):
        page = it.get("page")
        if isinstance(page, list):
            page_str = ", ".join(str(p) for p in page if p is not None)
        elif page is None:
            page_str = ""
        else:
            page_str = str(page)

        old_value = it.get("old_value") or ((it.get("evidence_left") or {}).get("quote")) or ""
        new_value = it.get("new_value") or ((it.get("evidence_right") or {}).get("quote")) or ""
        cost = it.get("cost_impact")
        cost_str = "" if (not cost or cost == "none") else cost
        source_layer = it.get("source_layer") or ""
        source_label = _SC_UNIFIED_SOURCE_LABELS.get(source_layer, source_layer)

        ws.append([
            i,
            it.get("sheet") or "",
            page_str,
            it.get("pair_label") or "",
            it.get("severity") or "",
            it.get("title") or "",
            it.get("summary") or "",
            old_value,
            new_value,
            it.get("construction_impact") or "",
            cost_str,
            source_label,
            "да" if it.get("requires_human_review") else "",
        ])
        row = ws[ws.max_row]
        sev = (it.get("severity") or "").lower()
        # №-ячейка подкрашена под severity
        if sev in sev_fill:
            row[0].fill = sev_fill[sev]
            row[4].fill = sev_fill[sev]
        row[7].fill = was_fill         # Было
        row[8].fill = became_fill      # Стало
        row[9].fill = impact_fill      # Влияние
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            cell.font = Font(size=10)

    widths = [5, 14, 10, 30, 11, 38, 50, 38, 38, 32, 12, 18, 11]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_sid = re.sub(r"[^A-Za-z0-9_.-]", "_", session_id)[:64] or "session"
    scope = "pair" if pair_id else "all"
    if accepted_only:
        scope += "_accepted"
    fname = f"stage_comparison_{safe_sid}_{scope}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── V2 review (pair-scoped manual verification) ─────────────────────────


def _v2_require_pair(session_id: str, pair_id: str) -> dict:
    """Найти пару в сессии или поднять 404. V2 строго pair-scoped."""
    session = store.get_session(session_id)
    if session is None:
        raise HTTPException(404, "Сессия не найдена")
    pair = next(
        (p for p in (session.get("pairs") or [])
         if isinstance(p, dict) and str(p.get("id") or "") == str(pair_id)),
        None,
    )
    if pair is None:
        raise HTTPException(404, "PDF-пара не найдена")
    return pair


@router.get("/sessions/{session_id}/pairs/{pair_id}/v2/changes")
async def v2_pair_changes(
    session_id: str,
    pair_id: str,
    include_excluded: bool = False,
):
    """V2-список расхождений ТОЛЬКО текущей PDF-пары.

    Read-only по отношению к comparison_result.json — никаких Qwen/Opus/
    unified-analysis. Накладывает ручные статусы из v2_review_status.json.

    По умолчанию (`include_excluded=false`) возвращает только инженерно
    значимые изменения; административные / только-оформление / косметика-шум
    скрыты (их аудит-снимок пишется в v2_excluded_changes.json). При
    `include_excluded=true` возвращаются все изменения, и у каждого есть
    `impact_class`, `excluded_from_main`, `exclusion_reason`.
    """
    _v2_require_pair(session_id, pair_id)
    try:
        return v2_review_mod.build_pair_v2_changes(
            session_id, pair_id, include_excluded=include_excluded,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc


@router.get("/sessions/{session_id}/pairs/{pair_id}/v2/summary")
async def v2_pair_summary(session_id: str, pair_id: str):
    """Сводка по V2-расхождениям текущей PDF-пары."""
    _v2_require_pair(session_id, pair_id)
    try:
        built = v2_review_mod.build_pair_v2_changes(session_id, pair_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    return {
        "session_id": session_id,
        "pair_id": pair_id,
        "summary": built.get("summary") or {},
    }


@router.patch("/sessions/{session_id}/pairs/{pair_id}/v2/changes/{change_id}")
async def v2_patch_change(session_id: str, pair_id: str, change_id: str, body: V2ReviewPatch):
    """Обновить ручной статус одного V2-изменения текущей пары."""
    _v2_require_pair(session_id, pair_id)
    patch = body.model_dump(exclude_none=True)
    try:
        entry = v2_review_mod.patch_change(session_id, pair_id, change_id, patch)
    except KeyError:
        raise HTTPException(404, "Изменение не найдено в текущей PDF-паре")
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {"ok": True, "id": change_id, "entry": entry}


@router.patch("/sessions/{session_id}/pairs/{pair_id}/v2/changes")
async def v2_bulk_patch_changes(session_id: str, pair_id: str, body: V2ReviewBulkPatch):
    """Пакетное обновление статусов — строго в рамках текущей пары."""
    _v2_require_pair(session_id, pair_id)
    patch = body.patch.model_dump(exclude_none=True)
    try:
        result = v2_review_mod.bulk_patch(session_id, pair_id, body.ids, patch)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {"ok": True, **result}


_V2_EXPORT_COLUMNS = [
    "№", "Лист", "Источник", "Тип", "Категория", "Важность",
    "Изменение", "Описание", "Было", "Стало", "Влияние", "Стоимость",
    "Evidence A", "Evidence B", "Quality label",
    "Статус проверки", "Комментарий", "Impact class",
]

_V2_IMPACT_CLASS_LABELS = {
    "construction_cost_impact": "влияет на стоимость",
    "construction_technical_impact": "влияет на строительство",
    "procurement_impact": "влияет на закупку",
    "schedule_or_risk_impact": "сроки / риски",
    "design_solution_impact": "проектное решение",
    "engineering_system_impact": "инж. система",
    "manual_review_required": "ручная проверка",
    "admin_only": "административное",
    "documentation_only": "только оформление",
    "cosmetic_or_noise": "косметика / шум",
    "unknown": "не классифицировано",
}


def _v2_fill_sheet(ws, items: list[dict]):
    """Заполнить лист книги V2-таблицей (общие колонки для всех листов)."""
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    ws.append(_V2_EXPORT_COLUMNS)
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border

    sev_fill = {
        "high":   PatternFill("solid", fgColor="FEE2E2"),
        "medium": PatternFill("solid", fgColor="FEF3C7"),
        "low":    PatternFill("solid", fgColor="DBEAFE"),
    }
    for i, it in enumerate(items, start=1):
        cost = str(it.get("cost_impact") or "")
        cost_str = "" if cost in ("", "none", "unknown") else cost
        source_layer = it.get("source_layer") or ""
        source_label = _SC_UNIFIED_SOURCE_LABELS.get(source_layer, source_layer)
        ws.append([
            i,
            it.get("sheet") or "",
            source_label,
            it.get("type") or "",
            it.get("category") or "",
            it.get("severity") or "",
            it.get("title") or "",
            it.get("summary") or "",
            it.get("old_value") or "",
            it.get("new_value") or "",
            it.get("construction_impact") or "",
            cost_str,
            it.get("evidence_left") or "",
            it.get("evidence_right") or "",
            it.get("quality_label") or "",
            it.get("review_status") or "not_reviewed",
            it.get("review_comment") or "",
            _V2_IMPACT_CLASS_LABELS.get(str(it.get("impact_class") or ""), it.get("impact_class") or ""),
        ])
        row = ws[ws.max_row]
        sev = (it.get("severity") or "").lower()
        if sev in sev_fill:
            row[0].fill = sev_fill[sev]
            row[5].fill = sev_fill[sev]
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            cell.font = Font(size=10)

    widths = [5, 16, 14, 16, 16, 11, 38, 46, 34, 34, 30, 12, 30, 30, 16, 18, 28, 18]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"


@router.get("/sessions/{session_id}/pairs/{pair_id}/v2/export.xlsx")
async def v2_export_xlsx(session_id: str, pair_id: str, include_excluded: bool = False):
    """Экспорт V2-таблицы ТОЛЬКО текущей PDF-пары.

    По умолчанию выгружаются только инженерно значимые изменения
    (административные / только-оформление / косметика-шум исключены; их
    количество показано в Summary как `excluded_count`). При
    `include_excluded=true` добавляется отдельный лист
    «Excluded admin-doc-noise» с исключёнными изменениями.

    Листы: Summary · All V2 changes · Confirmed · Needs clarification ·
    Rejected · Cost impact · Not reviewed [· Excluded admin-doc-noise].
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    _v2_require_pair(session_id, pair_id)
    try:
        # Берём ВСЕ изменения (с impact_class), сами разводим на kept/excluded.
        built = v2_review_mod.build_pair_v2_changes(session_id, pair_id, include_excluded=True)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc

    all_items = built.get("items") or []
    summary = built.get("summary") or {}
    # Основная инженерная ведомость — без исключённых.
    items = [it for it in all_items if not it.get("excluded_from_main")]
    excluded_items = [it for it in all_items if it.get("excluded_from_main")]

    wb = Workbook()
    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Метрика", "Значение"])
    for cell in ws_sum[1]:
        cell.font = Font(bold=True, size=10)
    eng_summary = v2_review_mod.compute_summary(items)
    _sum_rows = [
        ("Всего инженерных изменений", summary.get("engineering_total", len(items))),
        ("Высокая важность", eng_summary.get("high", 0)),
        ("Средняя важность", eng_summary.get("medium", 0)),
        ("Низкая важность", eng_summary.get("low", 0)),
        ("good", eng_summary.get("good", 0)),
        ("needs_human_review", eng_summary.get("needs_human_review", 0)),
        ("questionable", eng_summary.get("questionable", 0)),
        ("Подтверждено", eng_summary.get("confirmed", 0)),
        ("Отклонено", eng_summary.get("rejected", 0)),
        ("Не проверено", eng_summary.get("not_reviewed", 0)),
        ("Исключено всего", summary.get("excluded_total", len(excluded_items))),
        ("— административные", summary.get("excluded_admin_only", 0)),
        ("— только оформление", summary.get("excluded_documentation_only", 0)),
        ("— косметика / шум", summary.get("excluded_cosmetic_or_noise", 0)),
    ]
    for name, val in _sum_rows:
        ws_sum.append([name, val])
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 12

    def _by_status(status: str) -> list[dict]:
        return [it for it in items if str(it.get("review_status") or "") == status]

    cost_items = [
        it for it in items
        if str(it.get("review_status") or "") == "cost_impact"
        or str(it.get("cost_impact") or "") in ("possible", "likely")
    ]

    _v2_fill_sheet(wb.create_sheet("All V2 changes"), items)
    _v2_fill_sheet(wb.create_sheet("Confirmed"), _by_status("confirmed"))
    _v2_fill_sheet(wb.create_sheet("Needs clarification"), _by_status("needs_clarification"))
    _v2_fill_sheet(wb.create_sheet("Rejected"), _by_status("rejected"))
    _v2_fill_sheet(wb.create_sheet("Cost impact"), cost_items)
    _v2_fill_sheet(wb.create_sheet("Not reviewed"), _by_status("not_reviewed"))
    if include_excluded:
        _v2_fill_sheet(wb.create_sheet("Excluded admin-doc-noise"), excluded_items)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_sid = re.sub(r"[^A-Za-z0-9_.-]", "_", session_id)[:48] or "session"
    safe_pid = re.sub(r"[^A-Za-z0-9_.-]", "_", pair_id)[:32] or "pair"
    fname = f"stage_comparison_v2_{safe_sid}_{safe_pid}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── Unified GROUPED (deterministic post-processing) ─────────────────────


class RegroupRequest(BaseModel):
    force: bool = True


@router.get("/sessions/{session_id}/unified-grouped")
async def unified_grouped_endpoint(
    session_id: str,
    pair_id: Optional[str] = None,
    include_formal: bool = False,
    force_rebuild: bool = False,
    significance: Optional[str] = None,
    theme: Optional[str] = None,
):
    """Сгруппированный реестр значимых отличий (без LLM).

    Lazy-build: если `unified_findings_grouped.json` отсутствует, собираем
    его из `unified_findings.json` (или live flat) и сохраняем на диск.
    `force_rebuild=true` всегда пересобирает.

    Query params:
        pair_id          фильтр по конкретной паре
        include_formal   true → отдать hidden_formal_groups; default false
        significance     high|medium|low|formal — фильтр groups
        theme            фильтр по теме (см. THEMES в unified_grouping.py)
        force_rebuild    пересборка из flat findings
    """
    try:
        return unified_grouping_mod.get_unified_grouped(
            session_id,
            pair_id=pair_id,
            include_formal=include_formal,
            force_rebuild=force_rebuild,
            significance=significance,
            theme=theme,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc


class ExpertDecisionItem(BaseModel):
    item_id: str
    decision: str = Field(..., pattern="^(accepted|rejected)$")
    rejection_reason: Optional[str] = None


class ExpertReviewSubmission(BaseModel):
    decisions: list[ExpertDecisionItem] = []
    removed_ids: list[str] = []
    reviewer: str = ""


@router.get("/sessions/{session_id}/expert-review")
async def get_expert_review_endpoint(session_id: str, include_pairs: bool = False):
    """Решения эксперта по расхождениям сессии.

    Ключ хранения — стабильный raw `id` расхождения из `unified_findings.json`.
    Группированный вид агрегирует решения по `source_finding_ids` на фронте,
    поэтому регруппировка ничего не теряет.

    `include_pairs=true` добавляет `per_pair` — для каждой PDF-пары количество
    расхождений и сколько из них размечено, плюс флаг `fully_verified`. Это
    питает колонку «Проверено экспертом» на этапе «Загрузка документации».
    """
    return expert_review_mod.get_with_summary(session_id, include_pairs=include_pairs)


@router.post("/sessions/{session_id}/expert-review")
async def post_expert_review_endpoint(session_id: str, req: ExpertReviewSubmission):
    """Применить пачку решений эксперта (apply + removed)."""
    return expert_review_mod.apply_batch(
        session_id,
        decisions=[d.model_dump() for d in req.decisions],
        removed_ids=req.removed_ids,
        reviewer=req.reviewer or "",
    )


@router.post("/sessions/{session_id}/expert-review/prune-orphans")
async def prune_expert_review_orphans_endpoint(session_id: str, dry_run: bool = False):
    """Удалить «осиротевшие» экспертные решения по исчезнувшим `raw_id`.

    После регенерации сравнения id'шники расхождений (`chg_…`) меняются, а
    старые решения остаются в `expert_review.json` — у них нет строки в UI,
    снять их галочкой нельзя, и они накручивают счётчик «Принято/Отклонено».
    Чистит только пары, которые сейчас done, перечислимы и ЧАСТИЧНО совпадают
    по id (zero-overlap guard не даёт стереть пару целиком при регенерации).
    ЯВНОЕ действие: чтение раздела (`get_with_summary`) ничего не мутирует —
    счётчик чинится скоупингом на фронте, а реальная чистка диска только здесь.
    `dry_run=true` — посчитать без записи (рекомендуется сначала dry-run).
    """
    return expert_review_mod.prune_orphans(session_id, dry_run=dry_run)


@router.post("/sessions/{session_id}/v2-review/transfer")
async def v2_review_transfer_endpoint(session_id: str, req: Optional[V2ReviewTransferRequest] = None):
    """Перенести решения из классических «Расхождений» в V2 по всей сессии.

    Точные совпадения по `raw_id` переносятся детерминированно; остаток
    (находки, переименованные/слитые при перепрогоне Opus) сопоставляется
    Claude по смыслу. Конфликты помечаются, не перезаписываются; неуверенные
    совпадения переносятся с флагом «проверить». Возвращает отчёт.
    """
    req = req or V2ReviewTransferRequest()
    try:
        report = await asyncio.to_thread(
            review_transfer_mod.transfer_session,
            session_id,
            use_claude=req.use_claude,
        )
    except KeyError as exc:
        raise HTTPException(404, f"session not found: {exc}") from exc
    return report


@router.post("/sessions/{session_id}/regroup")
async def regroup_endpoint(session_id: str, req: RegroupRequest):
    """Принудительно пересобрать `unified_findings_grouped.json`.

    Не запускает Qwen/Opus. Только пересборка из существующего
    `unified_findings.json` / `unified-diff-flat`.
    """
    try:
        payload = unified_grouping_mod.build_unified_grouped(
            session_id, force=bool(req.force), persist=True,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    return {
        "ok": True,
        "session_id": session_id,
        "summary": payload.get("summary", {}),
        "groups_count": len(payload.get("groups") or []),
        "hidden_formal_count": len(payload.get("hidden_formal_groups") or []),
    }


@router.get("/enriched-compare-config")
async def enriched_compare_config_endpoint():
    """Инфо-ручка для UI: включён ли Opus pipeline, какая модель, доступен ли CLI."""
    cfg = enriched_compare_mod.load_config()
    provider, _ = enriched_compare_mod.resolve_provider(cfg)
    info = {
        "enabled": cfg.enabled,
        "provider": cfg.provider,
        "model": cfg.model,
        "max_chars": cfg.max_chars,
        "timeout_sec": cfg.timeout_sec,
        "available": False,
        "reason": None,
    }
    if provider is not None:
        ok, reason = provider.check_availability()
        info["available"] = bool(ok)
        info["reason"] = reason
    elif not cfg.enabled:
        info["reason"] = "disabled_via_env"
    else:
        info["reason"] = f"unknown_provider:{cfg.provider}"
    return info


# ─── Graphic summary ─────────────────────────────────────────────────────


@router.get("/sessions/{session_id}/pairs/{pair_id}/graphic-summary")
async def graphic_summary_endpoint(session_id: str, pair_id: str):
    summary = store.compute_graphic_summary(session_id, pair_id)
    if summary is None:
        raise HTTPException(404, "Сессия/пара не найдены")
    return summary


# ─── Graphic diff (paid LLM, gated) ──────────────────────────────────────


def _png_to_data_url(path: Path) -> str:
    raw = Path(path).read_bytes()
    return "data:image/png;base64," + base64.b64encode(raw).decode("ascii")


class _LLMResultError(RuntimeError):
    """LLM-вызов вернул is_error=True. is_paid_blocked=True для
    paid_api_blocked, иначе обычная ошибка."""

    def __init__(self, reason: str, is_paid_blocked: bool = False):
        super().__init__(reason)
        self.reason = reason
        self.is_paid_blocked = is_paid_blocked


async def _call_graphic_llm(
    *,
    model: str,
    left_png: Path,
    right_png: Path,
    project_id: str,
    job_id: str,
) -> tuple[str, Optional[float], Optional[str]]:
    """Вызов LLM (Gemini / GPT через OpenRouter) с двумя изображениями.

    Возвращает (summary, cost_usd, raw_response).
    Бросает _LLMResultError если result.is_error (включая paid_api_blocked).
    """
    from backend.app.services.llm.llm_runner import run_llm

    messages = [
        {
            "role": "system",
            "content": "Ты — эксперт-аудитор проектной документации. Отвечай кратко, структурированным списком на русском языке.",
        },
        {
            "role": "user",
            "content": [
                {"type": "text", "text": GRAPHIC_DIFF_PROMPT},
                {"type": "text", "text": "Первое изображение (предыдущая стадия):"},
                {"type": "image_url", "image_url": {"url": _png_to_data_url(left_png)}},
                {"type": "text", "text": "Второе изображение (новая стадия):"},
                {"type": "image_url", "image_url": {"url": _png_to_data_url(right_png)}},
            ],
        },
    ]

    # Используем существующий run_llm: он сам прогоняет paid_api_guard и
    # возвращает LLMResult(is_error=True, error_message="paid_api_blocked:...").
    result = await run_llm(
        stage="stage_comparison_graphic_diff",
        messages=messages,
        response_format=None,                # plain text
        temperature=0.2,
        timeout=300,
        model_override=model,
        project_id=project_id or "stage_comparison",
        job_id=job_id,
        source="stage_comparison.graphic_diff",
    )
    if getattr(result, "is_error", False):
        msg = getattr(result, "error_message", "") or "llm_error"
        raise _LLMResultError(msg, is_paid_blocked=msg.startswith("paid_api_blocked"))
    text = (result.text or "").strip()
    cost = getattr(result, "cost_usd", None)
    return text, cost, result.text




@router.post("/sessions/{session_id}/pairs/{pair_id}/graphic-diff")
async def graphic_diff_endpoint(
    session_id: str,
    pair_id: str,
    req: GraphicDiffRequest,
):
    """Подготовить crop'ы блоков; опционально запустить LLM-сравнение.

    Платный вызов разрешён только при run_paid=true и проходит через
    paid_api_guard. По умолчанию (run_paid=false) возвращает только URL'ы
    crop'ов для подготовки.

    Поведение синхронизировано с batch job (jobs.run_job):
      • paid_api_blocked → 200 OK, status='blocked', stored entry.status='blocked'
      • LLM is_error      → 200 OK, status='error',   stored entry.status='error'
      • success           → status='done'
    """
    # 1. Сгенерировать crop'ы (всегда). Если упадёт — сразу 4xx.
    try:
        left_png = store.render_block_crop(session_id, pair_id, "left", req.left_block_id)
        right_png = store.render_block_crop(session_id, pair_id, "right", req.right_block_id)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc

    base_url = (
        f"/api/stage-comparison/sessions/{session_id}/pairs/{pair_id}/block-image"
    )
    left_url = f"{base_url}?side=left&block_id={req.left_block_id}"
    right_url = f"{base_url}?side=right&block_id={req.right_block_id}"

    if not req.run_paid:
        return {
            "status": "prepared",
            "left_block_id": req.left_block_id,
            "right_block_id": req.right_block_id,
            "left_image_url": left_url,
            "right_image_url": right_url,
            "prompt": GRAPHIC_DIFF_PROMPT,
            "provider": "existing",
            "note": "Crop'ы подготовлены. Для запуска сравнения вызовите этот endpoint с run_paid=true.",
        }

    # 2. Платный путь: OpenRouter/Gemini через paid_api_guard
    model = req.model or "google/gemini-3.1-pro-preview"

    try:
        summary, cost, raw_text = await _call_graphic_llm(
            model=model,
            left_png=left_png,
            right_png=right_png,
            project_id=f"stage_comparison/{session_id}",
            job_id=f"{pair_id}:{req.left_block_id}->{req.right_block_id}",
        )
    except _LLMResultError as exc:
        # Унифицированная обработка: paid_api_blocked → 'blocked', иначе 'error'.
        # Возвращаем 200 OK, а status указываем в теле — так клиенту проще
        # отрисовать одинаковый UI, как для batch job.
        entry_status = "blocked" if exc.is_paid_blocked else "error"
        entry = store.add_graphic_diff_result(
            session_id, pair_id, req.left_block_id, req.right_block_id,
            status=entry_status,
            summary="",
            raw_response=None,
            model=model,
            cost_usd=None,
            error=exc.reason,
        )
        return {
            "status": entry_status,
            "left_block_id": req.left_block_id,
            "right_block_id": req.right_block_id,
            "left_image_url": left_url,
            "right_image_url": right_url,
            "error": exc.reason,
            "is_paid_blocked": exc.is_paid_blocked,
            "model": model,
            "entry": entry,
        }
    except Exception as exc:  # noqa: BLE001
        logger.exception("graphic-diff LLM call failed unexpectedly")
        entry = store.add_graphic_diff_result(
            session_id, pair_id, req.left_block_id, req.right_block_id,
            status="error",
            summary="",
            raw_response=None,
            model=model,
            cost_usd=None,
            error=str(exc)[:500],
        )
        return {
            "status": "error",
            "left_block_id": req.left_block_id,
            "right_block_id": req.right_block_id,
            "left_image_url": left_url,
            "right_image_url": right_url,
            "error": str(exc)[:500],
            "is_paid_blocked": False,
            "model": model,
            "entry": entry,
        }

    entry = store.add_graphic_diff_result(
        session_id, pair_id, req.left_block_id, req.right_block_id,
        status="done",
        summary=summary,
        raw_response=raw_text,
        model=model,
        cost_usd=cost,
        error=None,
    )
    return {
        "status": "done",
        "left_block_id": req.left_block_id,
        "right_block_id": req.right_block_id,
        "left_image_url": left_url,
        "right_image_url": right_url,
        "summary": summary,
        "model": model,
        "cost_usd": cost,
        "entry": entry,
    }


# ─── Graphic LLM config endpoint ─────────────────────────────────────────


@router.get("/graphic-llm-config")
async def graphic_llm_config_endpoint():
    """Инфо-ручка для UI / диагностики: какой provider активен.

    Локальные LLM-мощности удалены с платформы — остался единственный внешний
    provider (OpenRouter/Gemini). Никаких credentials не возвращает.
    """
    return {
        "provider": "existing",
        "base_url_present": False,
        "model": "",
        "fallback_model": "",
        "auth": "",
        "auth_configured": True,
        "model_load_enabled": False,
        "available": True,
        "reason": None,
    }


# ─── Batch graphic LLM jobs ──────────────────────────────────────────────


@router.post("/sessions/{session_id}/graphic-diff-jobs")
async def create_graphic_diff_job_endpoint(session_id: str, req: CreateGraphicDiffJobRequest):
    """Создать пакетный job сравнения графики через LLM.

    run_paid + confirm_paid должны быть true одновременно. Если нет — job
    создаётся в статусе rejected_no_confirm и LLM не запускается.
    """
    try:
        items = [it.model_dump() for it in (req.items or [])]
        job = jobs_mod.create_graphic_llm_job(
            session_id,
            scope=req.scope, pair_id=req.pair_id, items=items,
            run_paid=req.run_paid, confirm_paid=req.confirm_paid, model=req.model,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc

    # Если status == "queued" — запускаем в фоне (Задача 1).
    # HTTP-запрос возвращается сразу, прогресс читается через GET /graphic-diff-jobs/{id}.
    if job.get("status") == "queued":
        try:
            start_status = jobs_mod.start_job_in_background(session_id, job["id"])
        except KeyError as exc:
            raise HTTPException(404, str(exc)) from exc
        return {
            "ok": True,
            "job_id": job["id"],
            "status": "queued" if start_status == "started" else job.get("status"),
            "start_status": start_status,
            "scope": job.get("scope"),
            "model": job.get("model"),
            "progress": job.get("progress"),
        }
    # rejected_no_confirm / другой статус → возвращаем как есть
    return {
        "ok": False,
        "job_id": job["id"],
        "status": job.get("status"),
        "warnings": job.get("warnings"),
        "progress": job.get("progress"),
    }


@router.get("/sessions/{session_id}/graphic-diff-jobs")
async def list_graphic_diff_jobs(session_id: str):
    return {"jobs": jobs_mod.list_jobs(session_id)}


@router.get("/sessions/{session_id}/graphic-diff-jobs/{job_id}")
async def get_graphic_diff_job(session_id: str, job_id: str):
    job = jobs_mod.get_job(session_id, job_id)
    if job is None:
        raise HTTPException(404, "job_not_found")
    return job


@router.post("/sessions/{session_id}/graphic-diff-jobs/{job_id}/cancel")
async def cancel_graphic_diff_job(session_id: str, job_id: str):
    job = jobs_mod.cancel_job(session_id, job_id)
    if job is None:
        raise HTTPException(404, "job_not_found")
    return job


# ─── Graphic LLM batch jobs остаются выше; findings/warnings/reports
#     эндпоинты удалены вместе с переездом вкладки «Отчёт» на read-only
#     сводку согласованных расхождений (см. unified-diff-flat/export.xlsx
#     ?accepted_only=true). ───────────────────────────────────────────────


# ─── Pipeline V2 (controlled integration): read-only UI payload ────────────
#
# Endpoint НИЧЕГО не запускает (ни Pipeline V2, ни Qwen/Opus/LLM, ни jobs)
# и НИЧЕГО не пишет: отдаёт готовый pipeline_v2_ui_payload.json или собирает
# payload из готовых артефактов dry-run. Отсутствие артефактов — обычный
# JSON-ответ {"status": "not_found", ...}, не 404 (контракт для портала).
# Дисковое I/O уведено в threadpool (sync-тяжёлый handler в event loop
# блокирует /api/info и провоцирует watchdog-restart).


@router.get("/pipeline-v2/{session_id}/ui-payload")
async def get_pipeline_v2_ui_payload_endpoint(session_id: str,
                                              pair_id: Optional[str] = None):
    try:
        return await run_in_threadpool(
            pipeline_v2_payload_mod.discover_pipeline_v2_payload,
            session_id, pair_id)
    except ValueError as exc:
        # невалидный session_id/pair_id (path traversal и т.п.)
        raise HTTPException(400, str(exc)) from exc


@router.get("/pipeline-v2/{session_id}/graphic-vision-grounding")
async def get_pipeline_v2_grounding_detail_endpoint(
        session_id: str, pair_id: Optional[str] = None,
        kind: str = "all", status: str = "all",
        item_id: Optional[str] = None, limit: int = 100, offset: int = 0):
    """Read-only детализация graphic_vision_grounding_report.json.

    Отдаёт конкретные grounded/weakly_grounded/ungrounded/rejected_* сущности и
    изменения карточками (value/status/reason/anchor/source/page/fact_level).
    НИЧЕГО не запускает и не пишет; отсутствие отчёта — обычный JSON
    {"status":"not_found"}, битый — {"status":"error"}, не 500.
    """
    try:
        return await run_in_threadpool(
            pipeline_v2_payload_mod.discover_graphic_vision_grounding_detail,
            session_id, pair_id, kind=kind, status=status, item_id=item_id,
            limit=limit, offset=offset)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


# ─── Pipeline V2: controlled operator-triggered run («Запустить V2») ──────
# State-changing: запускает существующий dry-run runner в фоновом job'е.
# read-only ui-payload сервис этим НЕ затрагивается.

class PipelineV2RunRequest(BaseModel):
    """Тело POST .../pairs/{pair_id}/run."""
    mode: str = Field(default="dry_run")
    confirm: bool = Field(default=False)
    confirm_session_id: Optional[str] = Field(default=None)
    confirm_pair_id: Optional[str] = Field(default=None)
    rerun_existing: bool = Field(default=False)
    create_backup: bool = Field(default=True)
    operator_note: Optional[str] = Field(default=None)


def _pipeline_v2_run_payload(job: dict) -> dict:
    """Компактный accepted-ответ для UI."""
    return {
        "ok": True,
        "job_id": job.get("id"),
        "session_id": job.get("session_id"),
        "pair_id": job.get("pair_id"),
        "status": job.get("status"),
        "status_url": pipeline_v2_run_jobs_mod.status_url(
            job.get("session_id"), job.get("pair_id"), job.get("id")),
        "message": "Pipeline V2 run accepted",
    }


@router.post("/pipeline-v2/{session_id}/pairs/{pair_id}/run")
async def post_pipeline_v2_run_endpoint(
        session_id: str, pair_id: str, req: PipelineV2RunRequest):
    """Запустить controlled Pipeline V2 run для пары (operator-triggered).

    Запускает СУЩЕСТВУЮЩИЙ ``run_pipeline_v2_dry_run`` в фоновом job'е
    (offline: ``llm_runner=None``/``vision_runner=None`` → модели НЕ
    задействуются). Гейты: confirm + confirm_session_id/pair_id (422);
    сессия/пара существуют (404); артефакты уже есть без ``rerun_existing``
    (409); уже идёт run на эту пару (409). При rerun создаётся backup
    ``pipeline_v2_backup_before_ui_run_<TS>``. Пишет ТОЛЬКО артефакты
    pipeline_v2 этой пары + job-статус + manifest. ui-payload остаётся
    read-only.
    """
    payload = req.model_dump()
    try:
        job = await run_in_threadpool(
            pipeline_v2_run_jobs_mod.create_run_job, session_id, pair_id, payload)
    except pipeline_v2_run_jobs_mod.PipelineV2RunConfirmError as exc:
        raise HTTPException(422, str(exc)) from exc
    except pipeline_v2_run_jobs_mod.PipelineV2RunNotFound as exc:
        raise HTTPException(404, str(exc)) from exc
    except pipeline_v2_run_jobs_mod.PipelineV2RunConflict as exc:
        raise HTTPException(409, str(exc)) from exc
    except pipeline_v2_run_jobs_mod.PipelineV2RunError as exc:
        raise HTTPException(400, str(exc)) from exc
    except ValueError as exc:  # невалидный session_id/pair_id (_safe_id отверг)
        raise HTTPException(400, str(exc)) from exc
    pipeline_v2_run_jobs_mod.start_job_in_background(session_id, job["id"])
    return _pipeline_v2_run_payload(job)


@router.get("/pipeline-v2/{session_id}/pairs/{pair_id}/run-status/{job_id}")
async def get_pipeline_v2_run_status_endpoint(
        session_id: str, pair_id: str, job_id: str):
    """Статус controlled Pipeline V2 run job'а (для polling'а UI)."""
    try:
        job = await run_in_threadpool(
            pipeline_v2_run_jobs_mod.get_job, session_id, job_id)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    if job is None or job.get("pair_id") != pair_id:
        raise HTTPException(404, "run_job_not_found")
    return job


@router.get("/pipeline-v2/{session_id}/pairs/{pair_id}/run-active")
async def get_pipeline_v2_run_active_endpoint(
        session_id: str, pair_id: str):
    """Активный (running/queued) run job по паре — для восстановления UI."""
    try:
        job = await run_in_threadpool(
            pipeline_v2_run_jobs_mod.find_active_pair_job, session_id, pair_id)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {"job": job}
