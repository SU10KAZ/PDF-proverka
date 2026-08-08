"""Контракт семантической эквивалентности локального и удалённого аудита.

**Что здесь доказывается.** Не «файлы совпали побайтово» — они и не могут:
имена каталогов, идентификаторы прогона и отметки времени различаются по
построению. Доказывается другое и более сильное: **аудит, выполненный на чужой
машине, дал ТОТ ЖЕ инженерный результат**, что аудит, выполненный на центре.

**Почему нормализатор опасен и потому написан узко.** Любое «уберём поле, оно
шумит» — это способ спрятать расхождение. Правило `ключ на _s волатилен`
однажды уже вычищало из сравнения расчётный расход в л/с; правило «вырезать
meta целиком» уносило атрибуцию замечаний по детекторам. Поэтому здесь:

  * список волатильных ключей ЗАКРЫТ и перечислен поимённо;
  * отметки времени распознаются по ЗНАЧЕНИЮ (ISO-8601), а ключ остаётся —
    исчезнувшее поле обязано остаться видимым;
  * сортируются только те коллекции, которые семантически являются МНОЖЕСТВОМ;
  * поля дисциплины, замечаний, норм и ссылок не вычищаются НИКОГДА — они и
    есть предмет сравнения;
  * отсутствие обязательного артефакта — расхождение, а не «нечего сравнивать».

Модуль лежит в коде платформы, а не в тестах, потому что им пользуются оба:
автоматические тесты контракта и живой стенд.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Iterable, Optional

#: Артефакты pre-norm части — то, что делает воркер.
WORKER_ARTIFACTS: tuple[str, ...] = (
    "03_findings.json",
    "01_blocks_analysis.json",
    "02_text_analysis.json",
    "03_findings_review.json",
    "optimization.json",
    "optimization_review.json",
)

#: Артефакты ЦЕНТРАЛЬНОГО хвоста. Их отсутствие в удалённом случае означало бы,
#: что хвост не выполнялся, — то есть ровно тот дефект, который закрывает этап.
CENTRAL_ARTIFACTS: tuple[str, ...] = (
    "norm_checks.json",
    "03a_norms_verified.json",
)

#: Артефакты, которые ОБЯЗАНЫ существовать на обеих сторонах.
REQUIRED_ARTIFACTS: tuple[str, ...] = WORKER_ARTIFACTS + CENTRAL_ARTIFACTS + (
    "pipeline_log.json",
)

#: Ключи, различие которых допустимо (§17.2 задания). Список ЗАКРЫТ.
VOLATILE_KEYS: frozenset[str] = frozenset(
    {
        # время и длительность
        "generated_at", "created_at", "finished_at", "started_at", "completed_at",
        "updated_at", "timestamp", "date", "duration_sec", "duration_ms",
        "elapsed", "elapsed_sec", "wall_clock_sec",
        # идентификаторы прогона и транспорта
        "job_id", "attempt_id", "run_id", "package_id", "worker_id",
        "correlation_id", "instance_id", "pid", "process_group_id",
        "hostname", "host", "machine",
        # пути рантайма
        "project_dir", "output_dir", "artifacts_dir", "runtime_plan_path",
        "version_dir", "run_dir", "log_path", "spec_path", "staging",
        "journal", "path", "file_path", "crop_path",
        # расход: сравнивается ОТДЕЛЬНО и агрегатом, а не по каждому вызову
        "cost_usd", "cost_usd_notional", "tokens", "input_tokens",
        "output_tokens", "cache_creation_tokens", "cache_read_tokens",
        "api_calls", "model_calls",
    }
)

#: Ключи, которые НЕ вычищаются ни при каких обстоятельствах. Отдельный список
#: нужен как рубеж против будущего расширения `VOLATILE_KEYS`: пересечение
#: проверяется машинно (`assert_contract_is_sane`).
PROTECTED_KEYS: frozenset[str] = frozenset(
    {
        "discipline", "discipline_id", "section", "profile",
        "findings", "text_findings", "block_analyses", "problem", "description",
        "recommendation", "solution", "risk", "norm", "norm_quote", "norms",
        "references", "related_block_ids", "severity", "category", "sheet",
        "page", "status", "id", "schema_version", "stages", "summary",
        "value_found", "highlight_regions", "meta", "_meta",
    }
)

#: ISO-8601 по ЗНАЧЕНИЮ. `detected_at` слова о времени в имени не содержит, а
#: `12,5` в поле расхода содержит цифры, но временем не является.
_ISO_TS = re.compile(
    r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}(\.\d+)?(Z|[+-]\d{2}:?\d{2})?$"
)

#: Коллекции, порядок которых семантически не значим. Всё остальное
#: сравнивается В ПОРЯДКЕ: перестановка замечаний — это расхождение.
UNORDERED_LIST_KEYS: frozenset[str] = frozenset(
    {"completed_stages", "forbidden_stages_not_run", "required_artifacts",
     "generated_artifacts", "related_block_ids", "tags", "sources"}
)


class SemanticContractError(RuntimeError):
    """Контракт эквивалентности внутренне противоречив."""


def assert_contract_is_sane() -> None:
    """Проверить сам контракт: волатильное и защищённое не пересекаются.

    Проверка дешёвая и стоит здесь ровно потому, что расширение списка
    волатильных ключей — самый простой способ незаметно «починить» красный
    diff.
    """
    overlap = VOLATILE_KEYS & PROTECTED_KEYS
    if overlap:
        raise SemanticContractError(
            "Ключи одновременно волатильны и защищены: " + ", ".join(sorted(overlap))
        )


def project(value: Any, *, key_name: str = "") -> Any:
    """Семантическая проекция значения.

    Убирает ТОЛЬКО перечисленное. Абсолютные пути схлопываются в метку по
    значению: они попадают в артефакты в разных полях, и перечислять их
    поимённо значило бы догонять список вечно.
    """
    if isinstance(value, dict):
        out = {}
        for key, item in sorted(value.items()):
            if key in VOLATILE_KEYS and key not in PROTECTED_KEYS:
                continue
            out[key] = project(item, key_name=str(key))
        return out
    if isinstance(value, list):
        projected = [project(item, key_name=key_name) for item in value]
        if key_name in UNORDERED_LIST_KEYS:
            return sorted(projected, key=_stable_key)
        return projected
    if isinstance(value, str):
        if _ISO_TS.match(value):
            return "<timestamp>"
        if value.startswith("/") and len(value) > 1 and "\n" not in value:
            return "<path>"
        return value
    if isinstance(value, float):
        return round(value, 6)
    return value


def _stable_key(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


# ─── Сбор проекции результата ────────────────────────────────────────────────
def _read_json(path: Path) -> Optional[Any]:
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None


def _find(name: str, directories: Iterable[Path]) -> Optional[Path]:
    for directory in directories:
        candidate = Path(directory) / name
        if candidate.is_file():
            return candidate
    return None


def findings_signature(payload: Any) -> list[dict[str, Any]]:
    """Инженерная подпись набора замечаний.

    Сравнивается ИМЕННО она, а не сырой JSON: у замечания есть поля, которые
    зависят от прогона (порядковый идентификатор, отметки времени), и поля,
    которые составляют смысл. Ниже — вторые.
    """
    items = payload
    if isinstance(payload, dict):
        for key in ("findings", "items", "results"):
            if isinstance(payload.get(key), list):
                items = payload[key]
                break
    if not isinstance(items, list):
        return []
    out: list[dict[str, Any]] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        out.append(
            {
                "severity": item.get("severity") or item.get("category"),
                "category": item.get("category"),
                "sheet": item.get("sheet"),
                "page": item.get("page"),
                "problem": _norm_text(item.get("problem") or item.get("finding")),
                "description": _norm_text(item.get("description")),
                "recommendation": _norm_text(
                    item.get("recommendation") or item.get("solution")
                ),
                "norm": _norm_text(item.get("norm")),
                "norm_quote": _norm_text(item.get("norm_quote")),
                "references": sorted(
                    str(x) for x in (item.get("references") or []) if x
                ),
                "related_block_ids": sorted(
                    str(x) for x in (item.get("related_block_ids") or []) if x
                ),
            }
        )
    out.sort(key=_stable_key)
    return out


def _norm_text(value: Any) -> str:
    """Нормализация текста: пробелы и регистр краёв, но не содержание."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def stage_completion_map(pipeline_log: Any) -> dict[str, str]:
    if not isinstance(pipeline_log, dict):
        return {}
    stages = pipeline_log.get("stages")
    if not isinstance(stages, dict):
        return {}
    return {
        str(name): str((entry or {}).get("status") or "")
        for name, entry in sorted(stages.items())
        if isinstance(entry, dict)
    }


def excel_projection(version_dir: Path) -> dict[str, Any]:
    """Структурная проекция Excel: имена листов и число строк.

    Байтовое сравнение xlsx бессмысленно (в файле лежат отметки времени и
    порядок zip-записей), а отсутствие проверки означало бы «финальный артефакт
    не сравнивается вовсе».
    """
    candidates = sorted(Path(version_dir).rglob("*.xlsx"))
    if not candidates:
        return {"present": False}
    try:
        from openpyxl import load_workbook
    except Exception:                              # noqa: BLE001 — библиотека опциональна
        return {"present": True, "files": [p.name for p in candidates],
                "structure": "openpyxl_unavailable"}
    sheets: dict[str, int] = {}
    for path in candidates:
        try:
            book = load_workbook(path, read_only=True)
        except Exception:                          # noqa: BLE001
            continue
        for sheet in book.worksheets:
            sheets[f"{path.name}:{sheet.title}"] = int(sheet.max_row or 0)
        book.close()
    return {"present": True, "files": sorted(p.name for p in candidates),
            "sheets": dict(sorted(sheets.items()))}


def markdown_projection(version_dir: Path) -> dict[str, Any]:
    """Обязательный Markdown отчёта: наличие и нормализованный текст."""
    out: dict[str, Any] = {}
    for path in sorted(Path(version_dir).rglob("*.md")):
        rel = path.relative_to(version_dir).as_posix()
        if rel.startswith(("01_input/", "02_work/")):
            continue                               # исходники, не результат
        try:
            out[rel] = _norm_text(path.read_text(encoding="utf-8"))
        except OSError:
            out[rel] = "<unreadable>"
    return out


def usage_totals(usage_report: Any) -> dict[str, Any]:
    """Агрегат расхода поддельных провайдеров: суммы, а не отдельные вызовы."""
    entries = (usage_report or {}).get("entries") if isinstance(usage_report, dict) else None
    if not isinstance(entries, list):
        return {}
    totals = {"calls": 0, "input_tokens": 0, "output_tokens": 0, "stages": []}
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        totals["calls"] += int(entry.get("calls") or 0)
        totals["input_tokens"] += int(entry.get("input_tokens") or 0)
        totals["output_tokens"] += int(entry.get("output_tokens") or 0)
        totals["stages"].append(str(entry.get("stage") or ""))
    totals["stages"] = sorted(totals["stages"])
    return totals


def collect_projection(
    *,
    version_dir: Path,
    artifact_dirs: Optional[list[Path]] = None,
    final_status: Optional[str] = None,
    discipline_id: Optional[str] = None,
    discipline_profile_hash: Optional[str] = None,
    source_tree_hash: Optional[str] = None,
    usage_report: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    """Собрать полную семантическую проекцию РЕЗУЛЬТАТА аудита."""
    assert_contract_is_sane()
    version_dir = Path(version_dir)
    dirs = [Path(d) for d in (artifact_dirs or [])]
    dirs += [version_dir / "03_analysis" / "latest"]
    runs = sorted((version_dir / "03_analysis" / "runs").glob("*"))
    dirs += list(reversed(runs))

    artifacts: dict[str, Any] = {}
    missing: list[str] = []
    for name in REQUIRED_ARTIFACTS:
        path = _find(name, dirs)
        if path is None:
            artifacts[name] = None
            missing.append(name)
            continue
        artifacts[name] = project(_read_json(path))

    findings_path = _find("03_findings.json", dirs)
    findings_raw = _read_json(findings_path) if findings_path else None
    log_path = _find("pipeline_log.json", dirs)

    return {
        "contract_version": 1,
        "discipline_id": discipline_id,
        "discipline_profile_hash": discipline_profile_hash,
        "final_status": final_status,
        "source_tree_hash": source_tree_hash,
        "missing_artifacts": sorted(missing),
        "stage_completion": stage_completion_map(_read_json(log_path) if log_path else None),
        "findings_count": len(findings_signature(findings_raw)),
        "findings": findings_signature(findings_raw),
        "artifacts": artifacts,
        "excel": excel_projection(version_dir),
        "markdown": markdown_projection(version_dir),
        "usage_totals": usage_totals(usage_report),
    }


# ─── Сравнение ───────────────────────────────────────────────────────────────
def semantic_diff(left: Any, right: Any, *, path: str = "") -> list[str]:
    """Читаемый список расхождений. Пустой список = проекции совпали."""
    out: list[str] = []
    if type(left) is not type(right) and not (
        isinstance(left, (int, float)) and isinstance(right, (int, float))
    ):
        out.append(f"{path or '<корень>'}: тип {type(left).__name__} ≠ {type(right).__name__}")
        return out
    if isinstance(left, dict):
        for key in sorted(set(left) | set(right)):
            where = f"{path}.{key}" if path else str(key)
            if key not in left:
                out.append(f"{where}: есть только справа ({_short(right[key])})")
                continue
            if key not in right:
                out.append(f"{where}: есть только слева ({_short(left[key])})")
                continue
            out.extend(semantic_diff(left[key], right[key], path=where))
        return out
    if isinstance(left, list):
        if len(left) != len(right):
            out.append(f"{path}: длина {len(left)} ≠ {len(right)}")
        for index, (a, b) in enumerate(zip(left, right)):
            out.extend(semantic_diff(a, b, path=f"{path}[{index}]"))
        return out
    if left != right:
        out.append(f"{path or '<корень>'}: {_short(left)} ≠ {_short(right)}")
    return out


def _short(value: Any, limit: int = 160) -> str:
    text = json.dumps(value, ensure_ascii=False, sort_keys=True) if not isinstance(
        value, str
    ) else value
    return text if len(text) <= limit else text[: limit - 1] + "…"
