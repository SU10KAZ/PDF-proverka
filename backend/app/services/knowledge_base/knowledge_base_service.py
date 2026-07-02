"""
Сервис базы знаний — сбор экспертных решений, хранение, анализ паттернов.
"""
import json
import os
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from backend.app.core.config import KNOWLEDGE_BASE_DIR, DECISIONS_LOG_FILE, PATTERNS_FILE
from backend.app.models.expert_review import (
    ExpertDecision, KnowledgeBaseEntry, PatternSuggestion,
)
from backend.app.services.common import version_service
from backend.app.services.common.atomic_json import load_modify_save
from backend.app.services.common.project_service import resolve_project_dir
from backend.app.services.storage.projects_v2_source_resolver import (
    is_projects_v2_version_dir,
    load_version_project_info,
)


def _version_dir(project_id: str, *, must_exist: bool = False) -> Path:
    """Активная версия проекта (из ContextVar bound_version_id, fallback на latest).

    must_exist=True: для writer-ов — `resolve_project_dir` бросит
    `ProjectNotResolvedError`, если project_id не резолвится в реальную папку
    (вместо возврата несуществующего пути на корне объекта → orphan)."""
    if _v2_primary_enabled():
        try:
            ctx = version_service.resolve_project_version_context(
                project_id,
                version_service.get_bound_version_id(),
            )
            if ctx.get("storage_layout") == "projects_v2":
                return Path(ctx["version_dir"])
        except Exception:
            if must_exist:
                raise
    project_dir = resolve_project_dir(project_id, must_exist=must_exist)
    vid = version_service.get_bound_version_id()
    try:
        return version_service.get_version_dir(project_dir, project_id, vid)
    except version_service.VersionNotFoundError:
        return project_dir


def _output_dir(project_id: str, *, must_exist: bool = False) -> Path:
    return _version_dir(project_id, must_exist=must_exist) / "_output"


def _v2_primary_enabled() -> bool:
    try:
        from backend.app.services.storage.storage_write_facade import v2_is_primary
        return bool(v2_is_primary())
    except Exception:
        return False


def _analysis_dirs_for_version_dir(version_dir: Path) -> list[Path]:
    """Analysis artifacts inside ONE version dir: v2 03_analysis/latest, then _output."""
    if is_projects_v2_version_dir(version_dir):
        return [
            version_dir / "03_analysis" / "latest",
            version_dir / "_output",
        ]
    return [version_dir / "_output"]


def _analysis_dirs(project_id: str, *, must_exist: bool = False) -> list[Path]:
    """Analysis artifacts for enrichment: v2 latest first, legacy _output as before."""
    return _analysis_dirs_for_version_dir(
        _version_dir(project_id, must_exist=must_exist)
    )


_VERSION_DIR_RE = re.compile(r"v\d+$")


def _document_version_dirs(version_dir: Path) -> list[Path]:
    """Все версии документа, новейшая первой.

    Для projects_v2 перечисляет сестринские `versions/vNNN` (числовая сортировка =
    хронологическая, т.к. id zero-padded). Для legacy-раскладки или если перечислить
    не удалось — возвращает `[version_dir]`. Нужна для version-aware fallback
    хайдрейтинга БЗ: решение по item_id из старой версии находит источник там."""
    if not is_projects_v2_version_dir(version_dir):
        return [version_dir]
    versions_root = version_dir.parent  # .../versions
    try:
        siblings = [
            p for p in versions_root.iterdir()
            if p.is_dir() and _VERSION_DIR_RE.match(p.name)
        ]
    except OSError:
        return [version_dir]
    if not siblings:
        return [version_dir]
    siblings.sort(key=lambda p: p.name, reverse=True)  # vNNN … v001
    return siblings


def _review_paths(project_id: str, *, must_exist: bool = False) -> list[Path]:
    """Review storage priority: v2 canonical 04_review, then read-compatible fallbacks."""
    version_dir = _version_dir(project_id, must_exist=must_exist)
    if is_projects_v2_version_dir(version_dir):
        return [
            version_dir / "04_review" / "expert_review.json",
            version_dir / "03_analysis" / "latest" / "expert_review.json",
            version_dir / "_output" / "expert_review.json",
        ]
    return [version_dir / "_output" / "expert_review.json"]


def _review_path(project_id: str, *, must_exist: bool = False) -> Path:
    return _review_paths(project_id, must_exist=must_exist)[0]


def _decision_key(decision: dict) -> tuple[str, str]:
    return (
        str(decision.get("item_type") or ""),
        str(decision.get("item_id") or ""),
    )


def _merge_review_decisions(
    existing_payloads: list[dict],
    decisions: list[ExpertDecision],
    removed_ids: list[str] | None,
) -> list[dict]:
    """Merge idempotently by (item_type, item_id); canonical payloads win."""
    existing_by_key: dict[tuple[str, str], dict] = {}
    for payload in reversed(existing_payloads):
        for item in payload.get("decisions") or []:
            if not isinstance(item, dict):
                continue
            key = _decision_key(item)
            if not key[1]:
                continue
            existing_by_key[key] = item

    new_items = [d.model_dump() for d in decisions]
    new_keys = {_decision_key(item) for item in new_items}
    removed = set(removed_ids or [])

    merged = [
        item
        for key, item in existing_by_key.items()
        if key not in new_keys and key[1] not in removed
    ]
    merged.extend(new_items)
    return merged


def _ensure_kb_dir():
    """Создать папку knowledge_base/ если не существует."""
    KNOWLEDGE_BASE_DIR.mkdir(parents=True, exist_ok=True)


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")


# ═══════════════════════════════════════════════════════════════════════════
# Чтение / запись JSON
# ═══════════════════════════════════════════════════════════════════════════

def _load_json(path: Path) -> dict | list | None:
    if not path.exists():
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, UnicodeDecodeError):
        return None


def _save_json(path: Path, data):
    # Атомарная потокобезопасная запись: decisions_log/patterns — shared-стораж
    # на 140 проектов; plain open('w') рвался при крахе/гонке (reserc.md #7/#81/#87).
    _ensure_kb_dir()
    from backend.app.services.common.atomic_json import atomic_write_json
    atomic_write_json(path, data)


# ═══════════════════════════════════════════════════════════════════════════
# Экспертная оценка (per-project)
# ═══════════════════════════════════════════════════════════════════════════

def save_expert_review(project_id: str, decisions: list[ExpertDecision], reviewer: str = "", removed_ids: list[str] | None = None, *, stamp_schedule: bool = True) -> dict:
    """Сохранить решения эксперта по проекту.

    1. Записывает expert_review.json в canonical review storage проекта
    2. Обогащает решения контекстом из findings/optimization
    3. Добавляет записи в глобальный decisions_log.json

    stamp_schedule: если False — НЕ штамповать «день завершения» проекта в графике.
        Используется авто-переносом вердиктов (decision carryover): авто-решения
        не должны фиксировать день завершения — его ставит только ручная разметка.
    """
    # must_exist=True: не создаём orphan `_output` на несуществующем пути —
    # если project_id не резолвится в реальный проект/контейнер, поднимется
    # ProjectNotResolvedError (роутер вернёт 404).
    review_path = _review_path(project_id, must_exist=True)
    review_paths = _review_paths(project_id, must_exist=True)
    fallback_paths = [p for p in review_paths if Path(p) != review_path]

    # 1. Сохранить per-project файл (merge с существующими решениями)
    # под единым read-modify-write локом canonical review_path.
    def _mutate_review(current):
        existing_payloads = []
        if isinstance(current, dict):
            existing_payloads.append(current)
        existing_payloads.extend(
            payload
            for payload in (_load_json(path) for path in fallback_paths)
            if isinstance(payload, dict)
        )
        merged = _merge_review_decisions(existing_payloads, decisions, removed_ids)
        return {
            "project_id": project_id,
            "reviewer": reviewer,
            "reviewed_at": _now_iso(),
            "decisions": merged,
        }

    load_modify_save(
        review_path,
        _mutate_review,
        default={"project_id": project_id, "decisions": []},
    )

    # 2. Обогатить и добавить в глобальный лог
    enriched = _enrich_decisions(project_id, decisions, reviewer)
    _append_to_decisions_log(enriched)

    # 2b. Зафиксировать день завершения проекта для графика, если разметка стала
    # полной (все замечания + все оптимизации). Идемпотентно, fail-soft.
    # Авто-перенос вердиктов вызывает с stamp_schedule=False — авто-решения не
    # должны фиксировать день завершения проекта.
    if stamp_schedule:
        _stamp_schedule_completion_if_complete(project_id, reviewer)

    # Step 9/10 dual-write canary: shadow-зеркало проекта в v2 после сохранения
    # expert_review.json (no-op в legacy, fail-soft). decisions_log остаётся
    # общим shared-файлом (его v2-плечо здесь намеренно НЕ форкается).
    try:
        from backend.app.services.storage import storage_write_facade as _swf
        _swf.shadow_mirror_project_id_safe(project_id)
    except Exception:  # noqa: BLE001 — fail-soft, но #91: не молчим (наблюдаемость)
        import logging
        logging.getLogger(__name__).debug(
            "shadow_mirror_project_id_safe failed for %s", project_id, exc_info=True
        )

    return {
        "saved": len(decisions),
        "accepted": sum(1 for d in decisions if d.decision == "accepted"),
        "rejected": sum(1 for d in decisions if d.decision == "rejected"),
    }


def load_expert_review(project_id: str) -> Optional[dict]:
    """Загрузить сохранённые решения эксперта для проекта."""
    for path in _review_paths(project_id):
        data = _load_json(path)
        if data is not None:
            return data
    return None


def _load_item_maps_from_analysis_dirs(
    analysis_dirs: list[Path],
) -> tuple[dict[str, dict], dict[str, dict]]:
    """findings/opt карты ОДНОЙ версии по её analysis-папкам (latest-first внутри версии)."""
    findings_map: dict[str, dict] = {}
    for output_dir in analysis_dirs:
        for fname in ["03a_norms_verified.json", "03_findings.json"]:
            fdata = _load_json(output_dir / fname)
            if fdata:
                for item in fdata.get("findings", fdata.get("items", [])):
                    if isinstance(item, dict) and item.get("id"):
                        findings_map[str(item.get("id"))] = item
                break
        if findings_map:
            break

    opt_map: dict[str, dict] = {}
    for output_dir in analysis_dirs:
        opt_data = _load_json(output_dir / "optimization.json")
        if opt_data:
            for item in opt_data.get("items", []):
                if isinstance(item, dict) and item.get("id"):
                    opt_map[str(item.get("id"))] = item
            break

    return findings_map, opt_map


def _load_source_item_maps(project_id: str) -> tuple[dict[str, dict], dict[str, dict]]:
    """Source findings/optimizations активной (latest) версии, latest-first.

    Только текущая версия — для write-time enrich и расчёта дня завершения, где
    «источник» = именно активные findings. Для KB-хайдрейтинга (которому нужен
    fallback по версиям) используйте `_load_source_item_maps_versioned`."""
    return _load_item_maps_from_analysis_dirs(_analysis_dirs(project_id))


def _load_source_item_maps_versioned(
    project_id: str,
) -> tuple[dict[str, dict], dict[str, dict]]:
    """Version-aware источник для хайдрейтинга БЗ.

    Берём активную (latest) версию, затем добираем из более ранних версий те item_id,
    которых в latest нет. Новейшая версия выигрывает (`setdefault`), старые лишь
    заполняют пробелы. Это чинит пустые «Суть»/«Критичн.» у орфанных записей БЗ:
    решение принято по item_id (`F-NNN`) из старой версии, которого в latest уже нет
    после переаудита/перенумерации. Для item_id, присутствующих в latest, поведение
    НЕ меняется. Недеструктивно — только чтение."""
    version_dir = _version_dir(project_id)
    findings_map: dict[str, dict] = {}
    opt_map: dict[str, dict] = {}
    for vdir in _document_version_dirs(version_dir):  # newest-first
        f_map, o_map = _load_item_maps_from_analysis_dirs(
            _analysis_dirs_for_version_dir(vdir)
        )
        for k, v in f_map.items():
            findings_map.setdefault(k, v)
        for k, v in o_map.items():
            opt_map.setdefault(k, v)
    return findings_map, opt_map


def _source_summary(source: dict, item_type: str) -> str:
    for key in ("problem", "description", "summary", "title"):
        value = source.get(key)
        if value:
            return str(value).strip()
    if item_type != "optimization":
        return ""

    section = str(source.get("section") or "").strip()
    current = str(source.get("current") or "").strip()
    proposed = str(source.get("proposed") or "").strip()
    opt_type = str(source.get("type") or "").strip()
    if current and proposed:
        prefix = f"{section}: " if section else ""
        return f"{prefix}{current} → {proposed}"
    if proposed:
        return proposed
    if current:
        return current
    return " / ".join(part for part in (section, opt_type) if part)


def _norm_refs_from_source(source: dict) -> list[str]:
    norm = source.get("norm", source.get("norm_ref", ""))
    if not norm:
        return []
    return [norm] if isinstance(norm, str) else norm


def _primary_block_ids_from_source(source: dict) -> list[str]:
    """Первичные block_id замечания: source_block_ids + evidence-блоки, fallback related."""
    ids: list[str] = []
    seen: set[str] = set()

    def _add(value) -> None:
        if not value:
            return
        bid = str(value).strip()
        if bid and bid not in seen:
            seen.add(bid)
            ids.append(bid)

    for bid in source.get("source_block_ids") or []:
        _add(bid)
    for e in source.get("evidence") or []:
        if isinstance(e, dict) and e.get("source") != "grounding_service":
            _add(e.get("block_id") or e.get("id"))
    if not ids:
        for bid in source.get("related_block_ids") or []:
            _add(bid)
    return ids


def _evidence_types_from_source(source: dict) -> list[str]:
    """Уникальные типы evidence (image/text/...) у замечания."""
    types: list[str] = []
    seen: set[str] = set()
    for e in source.get("evidence") or []:
        if not isinstance(e, dict):
            continue
        etype = str(e.get("type") or "").strip()
        if etype and etype not in seen:
            seen.add(etype)
            types.append(etype)
    return types


def _source_for_entry(entry: dict, source_cache: dict[str, tuple[dict[str, dict], dict[str, dict]]]) -> dict:
    project_id = str(entry.get("source_project") or "").strip()
    item_id = str(entry.get("item_id") or "").strip()
    if not project_id or not item_id:
        return {}
    if project_id not in source_cache:
        try:
            # Version-aware: добираем item_id из старых версий, если в latest нет
            # (орфаны БЗ после переаудита/перенумерации F-NNN).
            source_cache[project_id] = _load_source_item_maps_versioned(project_id)
        except Exception:
            source_cache[project_id] = ({}, {})
    findings_map, opt_map = source_cache[project_id]
    if entry.get("item_type") == "optimization":
        return opt_map.get(item_id) or {}
    return findings_map.get(item_id) or opt_map.get(item_id) or {}


def _hydrate_kb_entry_from_source(entry: dict, source_cache: dict[str, tuple[dict[str, dict], dict[str, dict]]]) -> dict:
    """Fill display fields from source artifacts without writing decisions_log."""
    source = _source_for_entry(entry, source_cache)
    if not source:
        return entry

    item_type = str(entry.get("item_type") or "")
    if not entry.get("severity") and source.get("severity"):
        entry["severity"] = source.get("severity")
    if not entry.get("category"):
        entry["category"] = source.get("category") or source.get("type") or ""
    if not entry.get("summary"):
        entry["summary"] = _source_summary(source, item_type)
    if not entry.get("norm_refs"):
        entry["norm_refs"] = _norm_refs_from_source(source)
    if not entry.get("sheet") and source.get("sheet"):
        entry["sheet"] = str(source.get("sheet") or "")
    if entry.get("page") in (None, "") and source.get("page") not in (None, ""):
        entry["page"] = source.get("page")
    return entry


def _resolve_object_id() -> str:
    """Текущий объект (здание/комплекс): из bound-контекста, иначе выбранный.

    Единый резолвер, чтобы object_id в decisions_log и в штампе дня завершения
    графика были одинаковыми (иначе ключ (object_id, source_project) не сойдётся).
    """
    try:
        from backend.app.services.common import object_service, project_service
        return project_service._get_bound_object_id() or object_service.get_current_id() or ""
    except Exception:
        return ""


def _project_completion_day(project_id: str) -> Optional[str]:
    """День завершения проекта или None, если разметка ещё не полная.

    «Полная» = у КАЖДОГО исходного замечания и КАЖДОЙ исходной оптимизации есть
    решение эксперта (item_id из 03_findings.json / optimization.json ⊆ решённых
    в expert_review.json). День завершения = последний день среди решений (момент,
    когда разметка стала полной); fallback — сегодня. Если в проекте вообще нет
    исходных пунктов → None (нечего фиксировать).

    Чистая функция (без object_id и без записи) — переиспользуется runtime-штампом
    и backfill-скриптом.
    """
    review = load_expert_review(project_id) or {}
    decided_find: set[str] = set()
    decided_opt: set[str] = set()
    days: list[str] = []
    for d in (review.get("decisions") or []):
        if not isinstance(d, dict):
            continue
        iid = str(d.get("item_id") or "").strip()
        if not iid:
            continue
        if str(d.get("item_type") or "").strip() == "optimization":
            decided_opt.add(iid)
        else:
            decided_find.add(iid)
        ts = str(d.get("timestamp") or "")[:10]
        if len(ts) == 10:
            days.append(ts)

    findings_map, opt_map = _load_source_item_maps(project_id)
    src_find = {str(k) for k in findings_map.keys()}
    src_opt = {str(k) for k in opt_map.keys()}
    if not src_find and not src_opt:
        return None  # нечего размечать
    if not (src_find <= decided_find and src_opt <= decided_opt):
        return None  # ещё размечено не всё
    return max(days) if days else _now_iso()[:10]


def _stamp_schedule_completion_if_complete(project_id: str, reviewer: str) -> None:
    """Зафиксировать день завершения проекта для графика — если разметка полная.

    Ставится ОДИН раз и не меняется при правках (см.
    schedule_service.set_completion_once). Fail-soft — ошибки не мешают сохранению
    экспертной разметки.
    """
    try:
        comp_day = _project_completion_day(project_id)
        if not comp_day:
            return
        import backend.app.services.common.schedule_service as schedule_service
        schedule_service.set_completion_once(
            object_id=_resolve_object_id(),
            source_project=project_id,
            date=comp_day,
            reviewer=reviewer,
        )
    except Exception:
        import logging
        logging.getLogger(__name__).debug(
            "schedule completion stamp failed for %s", project_id, exc_info=True
        )


def _enrich_decisions(project_id: str, decisions: list[ExpertDecision], reviewer: str) -> list[KnowledgeBaseEntry]:
    """Обогатить решения контекстом из findings/optimization JSON."""
    findings_map, opt_map = _load_source_item_maps(project_id)

    # Загрузить project_info для section (из активной версии, fallback на корень)
    version_dir = _version_dir(project_id)
    info = load_version_project_info(version_dir) or {}
    if not info:
        info = _load_json(resolve_project_dir(project_id) / "project_info.json") or {}
    section = info.get("section", "")

    # Объект (здание/комплекс): из bound-контекста, иначе текущий выбранный.
    object_id = _resolve_object_id()

    # Следующий ID (монотонный max+1, НЕ len+1 — см. _next_decision_num)
    existing_log = _load_decisions_log()
    next_num = _next_decision_num(existing_log)

    entries = []
    for dec in decisions:
        # Записи без вердикта (pending-пометки авто-переноса, decision="") —
        # это НЕ решения эксперта, в глобальный decisions_log их не заносим.
        if not (dec.decision or "").strip():
            continue
        source = findings_map.get(dec.item_id) or opt_map.get(dec.item_id) or {}

        norm_refs = _norm_refs_from_source(source)

        entry = KnowledgeBaseEntry(
            id=f"DEC-{next_num:04d}",
            object_id=object_id,
            source_project=project_id,
            section=section,
            item_id=dec.item_id,
            item_type=dec.item_type,
            severity=source.get("severity", ""),
            category=source.get("category", source.get("type", "")),
            summary=_source_summary(source, dec.item_type),
            norm_refs=norm_refs,
            sheet=str(source.get("sheet", "")),
            page=source.get("page"),
            grounding_level=str(source.get("grounding_level") or ""),
            primary_block_ids=_primary_block_ids_from_source(source),
            evidence_types=_evidence_types_from_source(source),
            expert_decision=dec.decision,
            expert_reason=dec.rejection_reason or "",
            expert_reviewer=dec.reviewer or reviewer,
            expert_date=dec.timestamp or _now_iso(),
            customer_response=(source.get("external_register") or {}).get("customer_response", ""),
            # Авто-перенос вердикта из предыдущей версии (decision carryover).
            # current_version_id берём из bound-контекста (авто-этап bind-ит версию
            # перед записью) — нужен для кросс-версионного guard в decisions_log.
            carried_over=bool(getattr(dec, "carried_over", False)),
            carried_from_version=getattr(dec, "carried_from_version", "") or "",
            current_version_id=version_service.get_bound_version_id() or "",
        )
        entries.append(entry)
        next_num += 1

    return entries


# ═══════════════════════════════════════════════════════════════════════════
# Глобальный лог решений (knowledge_base/decisions_log.json)
# ═══════════════════════════════════════════════════════════════════════════

def _load_decisions_log() -> list[dict]:
    data = _load_json(DECISIONS_LOG_FILE)
    if isinstance(data, dict):
        return data.get("entries", [])
    if isinstance(data, list):
        return data
    return []


def _save_decisions_log(entries: list[dict]):
    _save_json(DECISIONS_LOG_FILE, {"entries": entries})


def _decisions_entries(data) -> list[dict]:
    if isinstance(data, dict):
        raw = data.get("entries", [])
    elif isinstance(data, list):
        raw = data
    else:
        raw = []
    return [e for e in raw if isinstance(e, dict)]


def _decisions_payload(entries: list[dict]) -> dict:
    return {"entries": entries}


def _load_modify_decisions_log(mutate_fn):
    return load_modify_save(
        DECISIONS_LOG_FILE,
        lambda data: _decisions_payload(mutate_fn(_decisions_entries(data))),
        default={"entries": []},
    )


def _next_decision_num(existing_log: list[dict]) -> int:
    """Следующий монотонный номер для DEC-NNNN: max(существующих) + 1.

    Раньше было len(log)+1 → после revoke номер переиспользовался, и один id
    указывал на десятки решений (audit: 1208 дублей id, 3122 записи). max+1
    гарантирует глобальную уникальность НОВЫХ id даже при пробелах от revoke
    (reserc.md #79/#100). Существующие коллизии лечит мигратор #82 (шаг 28).
    """
    import re
    mx = 0
    for e in existing_log:
        m = re.match(r"DEC-(\d+)$", str(e.get("id") or ""))
        if m:
            mx = max(mx, int(m.group(1)))
    return mx + 1


def _append_to_decisions_log(new_entries: list[KnowledgeBaseEntry]):
    """Добавить записи в глобальный лог (дедупликация по project+item_id)."""
    incoming = [entry.model_dump() for entry in new_entries]

    def _mutate(existing: list[dict]) -> list[dict]:
        updated_map = {(e.get("source_project"), e.get("item_id")): e for e in existing}
        used_ids = {str(e.get("id")) for e in existing if e.get("id")}
        next_num = _next_decision_num(existing)
        for item in incoming:
            key = (item.get("source_project"), item.get("item_id"))
            previous = updated_map.get(key)
            # Кросс-версионный guard для авто-переноса: ключ (source_project, item_id)
            # не версионный, а source_project = базовый pid для всех версий, поэтому
            # F-001 из V1 и V2 дают один ключ. Авто-перенос НЕ должен затирать запись,
            # относящуюся к другой версии. Ключи БЗ целиком не переделываем (это
            # отдельный трек docs/stable_finding_id.md).
            if item.get("carried_over") and previous is not None:
                prev_ver = str(previous.get("current_version_id") or "")
                cur_ver = str(item.get("current_version_id") or "")
                if prev_ver and cur_ver and prev_ver != cur_ver:
                    continue  # чужая версия — не трогаем
            if previous:
                if previous.get("id"):
                    item["id"] = previous.get("id")
                for field in ("customer_confirmed", "customer_date", "customer_note"):
                    if field in previous:
                        item[field] = previous.get(field)
            elif item.get("id") in used_ids:
                while f"DEC-{next_num:04d}" in used_ids:
                    next_num += 1
                item["id"] = f"DEC-{next_num:04d}"
                next_num += 1
            if item.get("id"):
                used_ids.add(str(item.get("id")))
            updated_map[key] = item
        return list(updated_map.values())

    _load_modify_decisions_log(_mutate)


def get_knowledge_base(
    status: Optional[str] = None,
    section: Optional[str] = None,
    item_type: Optional[str] = None,
    search: Optional[str] = None,
    object_id: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
) -> dict:
    """Получить записи базы знаний с фильтрацией."""
    entries = _load_decisions_log()

    # Вычислить status для каждой записи
    for e in entries:
        e["status"] = _entry_status(e)

    # Фильтрация
    if object_id:
        entries = [e for e in entries if e.get("object_id") == object_id]
    if status:
        entries = [e for e in entries if e.get("status") == status]
    if section:
        entries = [e for e in entries if e.get("section", "").upper() == section.upper()]
    if item_type:
        entries = [e for e in entries if e.get("item_type") == item_type]
    source_cache: dict[str, tuple[dict[str, dict], dict[str, dict]]] = {}
    if search:
        for e in entries:
            _hydrate_kb_entry_from_source(e, source_cache)
        s = search.lower()
        entries = [e for e in entries if s in json.dumps(e, ensure_ascii=False).lower()]

    total = len(entries)

    # Сортировка по дате (новые первые)
    entries.sort(key=lambda e: e.get("expert_date", ""), reverse=True)

    # Пагинация
    paginated = entries[offset:offset + limit]
    for e in paginated:
        _hydrate_kb_entry_from_source(e, source_cache)

    return {
        "total": total,
        "offset": offset,
        "limit": limit,
        "entries": paginated,
    }


def _entry_status(e: dict) -> str:
    """Статус записи для вкладок KB.

    customer_confirmed + ответ заказчика «Внесено» → fixed_by_customer
    (заказчик внёс изменения в РД); остальные согласованные → customer_confirmed.
    """
    if e.get("customer_confirmed"):
        if (e.get("customer_response") or "").strip() == "Внесено":
            return "fixed_by_customer"
        return "customer_confirmed"
    return e.get("expert_decision", "")


def get_kb_stats(object_id: Optional[str] = None) -> dict:
    """Счётчики по вкладкам (опционально в рамках одного объекта)."""
    entries = _load_decisions_log()
    if object_id:
        entries = [e for e in entries if e.get("object_id") == object_id]
    stats = {"rejected": 0, "accepted": 0, "customer_confirmed": 0, "fixed_by_customer": 0}
    for e in entries:
        st = _entry_status(e)
        if st in stats:
            stats[st] += 1
    stats["total"] = sum(stats.values())
    return stats


def mark_customer_confirmed(entry_ids: list[str], note: str = "") -> int:
    """Отметить записи как согласованные заказчиком."""
    target_ids = set(entry_ids)
    count = 0
    now = _now_iso()

    def _mutate(entries: list[dict]) -> list[dict]:
        nonlocal count
        for e in entries:
            if e.get("id") in target_ids and e.get("expert_decision") == "accepted":
                e["customer_confirmed"] = True
                e["customer_date"] = now
                if note:
                    e["customer_note"] = note
                count += 1
        return entries

    _load_modify_decisions_log(_mutate)
    return count


def unmark_customer_confirmed(entry_ids: list[str]) -> int:
    """Снять отметку согласования заказчиком."""
    target_ids = set(entry_ids)
    count = 0

    def _mutate(entries: list[dict]) -> list[dict]:
        nonlocal count
        for e in entries:
            if e.get("id") in target_ids and e.get("customer_confirmed"):
                e["customer_confirmed"] = False
                e["customer_date"] = None
                e["customer_note"] = None
                count += 1
        return entries

    _load_modify_decisions_log(_mutate)
    return count


def _find_project_dir(project_id: str) -> Optional[Path]:
    """Найти папку проекта — пробует resolve_project_dir, fallback через iter_project_dirs."""
    try:
        return resolve_project_dir(project_id)
    except Exception:
        pass
    # Fallback: поиск по имени
    try:
        from backend.app.services.common.project_service import iter_project_dirs
        for pid, path in iter_project_dirs():
            if pid == project_id or pid.endswith("/" + project_id):
                return path
    except Exception:
        pass
    return None


def revoke_decision(entry_id: str, project_id: str, item_id: str) -> int:
    # Удалить решение из глобального лога и из expert_review проекта.
    import logging
    removed = 0

    def _mutate_log(entries: list[dict]) -> list[dict]:
        nonlocal removed
        if project_id and item_id:
            def _match(e: dict) -> bool:
                return e.get("source_project") == project_id and e.get("item_id") == item_id
        else:
            id_hits = [e for e in entries if e.get("id") == entry_id]
            if len(id_hits) != 1:
                logging.getLogger(__name__).warning(
                    "revoke_decision: id=%r неуникален (%d записей) и нет составного "
                    "ключа — отказ, чтобы не удалить чужие решения",
                    entry_id, len(id_hits),
                )
                return entries

            def _match(e: dict) -> bool:
                return e.get("id") == entry_id

        kept = [e for e in entries if not _match(e)]
        matched = len(entries) - len(kept)
        if matched > 1:
            logging.getLogger(__name__).warning(
                "revoke_decision: совпало %d записей по (%s,%s)/id=%s — ожидалась ≤1; "
                "пропуск без удаления", matched, project_id, item_id, entry_id,
            )
            return entries
        removed = matched
        return kept

    _load_modify_decisions_log(_mutate_log)

    # Удалить из expert_review.json проекта (активная версия).
    if removed and project_id and item_id:
        project_dir = _find_project_dir(project_id)
        if project_dir:
            review_path = _review_path(project_id)
            if review_path.exists():
                def _mutate_review(review_data):
                    if not isinstance(review_data, dict):
                        return review_data
                    decisions = review_data.get("decisions")
                    if isinstance(decisions, list):
                        review_data["decisions"] = [
                            d for d in decisions
                            if not (isinstance(d, dict) and d.get("item_id") == item_id)
                        ]
                    return review_data

                load_modify_save(review_path, _mutate_review, default={"decisions": []})

    return removed


# ═══════════════════════════════════════════════════════════════════════════
# Детекция паттернов из отклонённых решений
# ═══════════════════════════════════════════════════════════════════════════

def _load_patterns() -> list[dict]:
    data = _load_json(PATTERNS_FILE)
    if isinstance(data, dict):
        return data.get("patterns", [])
    if isinstance(data, list):
        return data
    return []


def _save_patterns(patterns: list[dict]):
    _save_json(PATTERNS_FILE, {"patterns": patterns})


def detect_patterns(min_frequency: int = 3) -> list[dict]:
    """Найти повторяющиеся паттерны среди отклонённых решений.

    Группирует по (section, category, norm_prefix) и ищет кластеры с >= min_frequency.
    """
    entries = _load_decisions_log()
    rejected = [e for e in entries if e.get("expert_decision") == "rejected"]

    if not rejected:
        return []

    # Группировка по (section, category)
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for e in rejected:
        section = e.get("section", "").upper()
        category = e.get("category", "").lower()
        # Извлечь prefix нормы (до пункта)
        norm_prefix = ""
        norms = e.get("norm_refs", [])
        if norms:
            norm_prefix = norms[0].split(",")[0].split("п.")[0].strip()
        key = (section, category, norm_prefix)
        groups[key].append(e)

    # Фильтр по частоте
    existing_patterns = _load_patterns()
    existing_ids = {p.get("pattern_id") for p in existing_patterns}
    next_num = len(existing_patterns) + 1

    new_patterns = []
    for (section, category, norm_prefix), items in groups.items():
        if len(items) < min_frequency:
            continue

        # Собрать уникальные причины отклонения
        reasons = [e.get("expert_reason", "") for e in items if e.get("expert_reason")]
        common_reason = reasons[0] if reasons else "Повторяющееся отклонение"

        # Уникальные проекты
        projects = list({e.get("source_project", "") for e in items})
        example_ids = [e.get("id", "") for e in items[:5]]

        pattern_id = f"PAT-{next_num:03d}"
        # Проверить что не дублирует существующий
        desc = f"[{section}] Категория '{category}'"
        if norm_prefix:
            desc += f", норма {norm_prefix}"
        desc += f" — {len(items)} отклонений"

        # Пропустить если уже есть паттерн с такой же description
        if any(p.get("description") == desc for p in existing_patterns):
            continue

        suggested_fix = f"Не генерировать замечания типа '{category}'"
        if norm_prefix:
            suggested_fix += f" со ссылкой на {norm_prefix}"
        suggested_fix += f". Причина: {common_reason}"

        target_file = f"disciplines/{section}/checklist.md" if section else ""

        new_patterns.append({
            "pattern_id": pattern_id,
            "section": section,
            "description": desc,
            "frequency": len(items),
            "projects_affected": projects,
            "example_ids": example_ids,
            "suggested_fix": suggested_fix,
            "target_file": target_file,
            "status": "pending",
            "proposed_at": _now_iso(),
            "decided_by": None,
            "decided_at": None,
        })
        next_num += 1

    # Сохранить новые + существующие
    if new_patterns:
        all_patterns = existing_patterns + new_patterns
        _save_patterns(all_patterns)

    return _load_patterns()


def get_patterns() -> list[dict]:
    """Получить все паттерны."""
    return _load_patterns()


def update_pattern_status(pattern_id: str, status: str, edited_fix: Optional[str] = None, decided_by: str = "") -> bool:
    """Обновить статус паттерна (approve/dismiss/edit)."""
    patterns = _load_patterns()
    for p in patterns:
        if p.get("pattern_id") == pattern_id:
            p["status"] = status
            p["decided_at"] = _now_iso()
            p["decided_by"] = decided_by
            if edited_fix is not None:
                p["suggested_fix"] = edited_fix
            _save_patterns(patterns)
            return True
    return False


# ═══════════════════════════════════════════════════════════════════════════
# Импорт решений из Excel
# ═══════════════════════════════════════════════════════════════════════════

def _resolve_project_id_from_sheet(ws_title: str) -> str:
    """Найти полный project_id по сокращённому имени листа Excel.

    Имя листа может быть '133_23-ГК-ГРЩ', а нужен 'EOM/133_23-ГК-ГРЩ'.
    """
    from backend.app.services.common.project_service import iter_project_dirs

    # Убрать префикс "ОПТ " если есть
    name = ws_title
    if name.startswith("ОПТ "):
        name = name[4:]

    for pid, path in iter_project_dirs():
        # pid может быть "133_23-ГК-ГРЩ" или "EOM/133_23-ГК-ГРЩ"
        if pid == name or pid.endswith("/" + name) or pid.replace("/", "-") == name:
            return pid
    return name  # fallback — вернуть как есть


def import_decisions_from_excel(
    file_path: str,
    default_project_id: Optional[str] = None,
    reviewer: str = "",
) -> dict:
    """Импортировать решения из Excel-файла с колонками 'Решение эксперта' и 'Причина отклонения'.

    Возвращает {project_id: {saved, accepted, rejected}} для каждого обнаруженного проекта.

    `default_project_id` — fallback из UI-контекста, когда в Excel ни скрытая
    ячейка, ни имя листа не дают валидный project_id (старые экспорты для V2
    клали в скрытую ячейку "v2" вместо реального ID).

    `reviewer` — автор импорта (ФИО сотрудника). Резолвится роутером из
    портал-сессии; раньше не передавался → импортированные решения теряли
    автора (expert_reviewer="") и не показывались в графике работ.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    results = {}

    for ws in wb.worksheets:
        if ws.title.upper() in ("ИНСТРУКЦИЯ", "СВОДКА", "SUMMARY"):
            continue

        # Найти колонки по заголовку (строка 1)
        headers = {}
        header_row = list(next(ws.iter_rows(min_row=1, max_row=1), []))
        for col_idx, c in enumerate(header_row):
            val = str(c.value or "").strip().lower()
            if val == "решение эксперта" or ("решение" in val and "эксперт" in val):
                headers["decision"] = col_idx
            elif "причина" in val and "отклон" in val:
                headers["reason"] = col_idx
            elif val == "id":
                headers["id"] = col_idx
            elif val == "№" and "id" not in headers:
                headers["num"] = col_idx  # fallback: если нет ID, используем №
            elif val == "project_id":
                headers["project_id"] = col_idx
            elif "тип" in val:
                headers["type"] = col_idx

        # Если нет колонки ID, но есть № — не подходит (в № порядковый номер, а не F-001)
        # Для листов оптимизации колонка ID есть всегда

        if "decision" not in headers or "id" not in headers:
            continue

        # Определить project_id: скрытый столбец (строка 2) → имя листа → fallback.
        # Защита: если в скрытой ячейке или в имени листа лежит просто метка
        # версии ("v1"/"v2"/…) — это не project_id (старые экспорты для V2
        # клали туда basename папки = "v2"). В таких случаях используем
        # default_project_id, переданный из UI-контекста (currentProjectId).
        import re as _re
        def _looks_like_version(s: str) -> bool:
            return bool(s and _re.fullmatch(r"v\d+", s.strip(), flags=_re.IGNORECASE))

        def _pid_resolves(pid: str) -> bool:
            """project_id указывает на реальную папку проекта/контейнер."""
            if not pid:
                return False
            try:
                from backend.app.services.common.project_service import (
                    resolve_project_dir,
                )
                resolve_project_dir(pid, must_exist=True)
                return True
            except Exception:
                return False

        def _strip_version_label(pid: str) -> str:
            """Снять хвостовую метку версии ("<id> V1"/"<id>_v2"/…).

            Версия в импорте передаётся отдельно (`version_id`), поэтому в самом
            project_id метки "V1" быть не должно. Старые/внешние экспорты иногда
            запекали в скрытую ячейку композит вида "KM/1232-ЧМ-КМ-1 V1" —
            он не резолвится (реальная папка называется "KM/1232-ЧМ-КМ-1").
            """
            return _re.sub(r"[ _-]?[vV]\d+$", "", pid).strip() if pid else pid

        # Кандидаты в порядке приоритета. Берём ПЕРВЫЙ, который реально
        # резолвится в папку проекта — это и чинит 500 «Project directory not
        # resolved» на стейл-ячейках, и сохраняет per-sheet идентичность для
        # многопроектных отчётов (валидная скрытая ячейка по-прежнему выигрывает).
        candidates: list[str] = []
        if "project_id" in headers:
            row2 = list(next(ws.iter_rows(min_row=2, max_row=2), []))
            if row2 and headers["project_id"] < len(row2):
                pid_val = str(row2[headers["project_id"]].value or "").strip()
                if pid_val and not _looks_like_version(pid_val):
                    candidates.append(pid_val)
        # Снимаем префикс "ОПТ " вручную, чтобы проверить «голое» имя
        bare_title = ws.title[4:] if ws.title.startswith("ОПТ ") else ws.title
        if not _looks_like_version(bare_title):
            resolved = _resolve_project_id_from_sheet(ws.title)
            if resolved and not _looks_like_version(resolved):
                candidates.append(resolved)
        if default_project_id:
            candidates.append(default_project_id)

        project_id = None
        # Pass 1: первый кандидат, который резолвится как есть.
        for cand in candidates:
            if _pid_resolves(cand):
                project_id = cand
                break
        # Pass 2: ни один не резолвится → пробуем снять хвостовую метку версии
        # (стейл-экспорты "<id> V1"). Принимаем только если stripped резолвится.
        if not project_id:
            for cand in candidates:
                stripped = _strip_version_label(cand)
                if stripped and stripped != cand and _pid_resolves(stripped):
                    project_id = stripped
                    break

        decisions = []
        for row in ws.iter_rows(min_row=3, values_only=False):
            cells = list(row)
            item_id = str(cells[headers["id"]].value or "").strip()
            decision_raw = str(cells[headers["decision"]].value or "").strip().lower()
            reason = ""
            if "reason" in headers:
                reason = str(cells[headers["reason"]].value or "").strip()

            if not item_id or not decision_raw:
                continue

            # Нормализация
            if decision_raw in ("принято", "accepted", "да", "yes", "+"):
                decision = "accepted"
            elif decision_raw in ("отклонено", "отклонить", "rejected", "нет", "no", "-"):
                decision = "rejected"
            else:
                continue

            # Определить тип
            item_type = "finding"
            if item_id.upper().startswith("OPT"):
                item_type = "optimization"
            if "type" in headers:
                type_val = str(cells[headers["type"]].value or "").strip().lower()
                if "opt" in type_val:
                    item_type = "optimization"

            decisions.append(ExpertDecision(
                item_id=item_id,
                item_type=item_type,
                decision=decision,
                rejection_reason=reason if decision == "rejected" else None,
                timestamp=_now_iso(),
            ))

        import logging
        _log = logging.getLogger(__name__)
        _log.warning(
            "[import-excel] sheet=%r project_id=%r decisions=%d headers=%s bound_vid=%s",
            ws.title, project_id, len(decisions), list(headers.keys()),
            version_service.get_bound_version_id(),
        )

        if decisions and project_id:
            result = save_expert_review(project_id, decisions, reviewer)
            results[project_id] = result
        elif decisions and not project_id:
            _log.warning("[import-excel] decisions=%d but project_id not resolved for sheet=%r", len(decisions), ws.title)

    wb.close()
    return results
