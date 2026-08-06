# -*- coding: utf-8 -*-
"""
generate_excel_report.py
========================
Читает 03_findings.json из каждого проекта в папке projects/
и создаёт Excel-отчёт: каждый проект — отдельный лист.

Использование:
    python generate_excel_report.py                         # все проекты
    python generate_excel_report.py projects/133-23-GK-EM1 # один проект
    python generate_excel_report.py --out my_report.xlsx    # имя файла
    python generate_excel_report.py --no-summary            # без листа СВОДКА
"""

import os
import sys
import json
import argparse
from datetime import datetime
from backend.app.services.common import audit_scope

# Фикс кодировки Windows (cp1251 -> utf-8 в консоли)
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not found. Run: pip install openpyxl")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════
#  СИСТЕМА КАТЕГОРИЙ (5 уровней)
# ═══════════════════════════════════════════════════════════════════════

SEVERITY_CONFIG = {
    "КРИТИЧЕСКОЕ": {
        "bg":      "FFCCCC",   # светло-красный
        "bg_alt":  "FFD9D9",
        "fg":      "C00000",
        "icon":    "🔴",
        "desc":    "Нельзя строить — безопасность, нарушения ПУЭ / ГОСТ / СП",
    },
    "СУЩЕСТВЕННОЕ": {
        "bg":      "F4CCCC",
        "bg_alt":  "F9D9D9",
        "fg":      "990000",
        "icon":    "🔴",
        "desc":    "Серьёзное нарушение норм, существенное влияние на проект",
    },
    "ЭКОНОМИЧЕСКОЕ": {
        "bg":      "FCE4D6",
        "bg_alt":  "FDEBD9",
        "fg":      "C55A11",
        "icon":    "🟠",
        "desc":    "Деньги / объёмы / пересортица / риск ошибочной поставки",
    },
    "ЭКСПЛУАТАЦИОННОЕ": {
        "bg":      "FFFF99",
        "bg_alt":  "FFFFB3",
        "fg":      "7F6000",
        "icon":    "🟡",
        "desc":    "Будущие проблемы при эксплуатации (падение U, нет запасных, нет байпаса)",
    },
    "РЕКОМЕНДАТЕЛЬНОЕ": {
        "bg":      "DDEEFF",
        "bg_alt":  "E8F4FF",
        "fg":      "2E75B6",
        "icon":    "🔵",
        "desc":    "Опечатки, мелкие несоответствия, отсутствие пояснений",
    },
    "ПРОВЕРИТЬ ПО СМЕЖНЫМ": {
        "bg":      "F2F2F2",
        "bg_alt":  "F8F8F8",
        "fg":      "595959",
        "icon":    "⚪",
        "desc":    "Требует информации из ГП3, ГП4, ГП6, ГП8, АР и других смежных разделов",
    },
    "СНЯТО": {
        "bg":      "E0E0E0",
        "bg_alt":  "EBEBEB",
        "fg":      "808080",
        "icon":    "⬜",
        "desc":    "Замечание снято (неактуально или исправлено)",
    },
}

# Нормализация старых/английских имён → каноничные
SEV_NORMALIZE = {
    # старые русские
    "КРИТИЧНО":       "КРИТИЧЕСКОЕ",
    "СУЩЕСТВЕННО":    "СУЩЕСТВЕННОЕ",
    "РЕКОМЕНДАЦИЯ":   "РЕКОМЕНДАТЕЛЬНОЕ",
    "ПРОВЕРИТЬ":      "ПРОВЕРИТЬ ПО СМЕЖНЫМ",
    # английские (из пакетного анализа тайлов)
    "CRITICAL":       "КРИТИЧЕСКОЕ",
    "SUBSTANTIAL":    "СУЩЕСТВЕННОЕ",
    "ECONOMIC":       "ЭКОНОМИЧЕСКОЕ",
    "OPERATIONAL":    "ЭКСПЛУАТАЦИОННОЕ",
    "RECOMMENDATION": "РЕКОМЕНДАТЕЛЬНОЕ",
    "INFORMATIONAL":  "РЕКОМЕНДАТЕЛЬНОЕ",
    "CHECK_RELATED":  "ПРОВЕРИТЬ ПО СМЕЖНЫМ",
}

# Порядок сортировки категорий в сводке (по убыванию важности)
SEV_ORDER = [
    "КРИТИЧЕСКОЕ",
    "СУЩЕСТВЕННОЕ",
    "ЭКОНОМИЧЕСКОЕ",
    "ЭКСПЛУАТАЦИОННОЕ",
    "РЕКОМЕНДАТЕЛЬНОЕ",
    "ПРОВЕРИТЬ ПО СМЕЖНЫМ",
    "СНЯТО",
]

# ── Цвета шапок ──────────────────────────────────────────────────────
HEADER_BG   = "1F497D"   # тёмно-синий
HEADER_FG   = "FFFFFF"
PROJ_HDR_BG = "2E75B6"   # синий — строка с именем проекта
PROJ_HDR_FG = "FFFFFF"
TOTAL_BG    = "2E3F50"   # тёмный для итоговой строки

# ── Структура столбцов листа проекта ─────────────────────────────────
# (ключ,          заголовок,          ширина)
PROJ_COLUMNS = [
    ("num",              "№",                    5),
    ("finding_id",       "ID",                  12),
    ("sheet",            "Лист/Раздел",         24),
    ("problem",          "Проблема",            28),
    ("description",      "Описание",            52),
    ("solution",         "Решение",             48),
    ("severity",         "Категория",           22),
    ("risk",             "Чем грозит",          32),
    ("expert_decision",  "Решение эксперта",    16),
    ("rejection_reason", "Причина отклонения",  32),
]

# Столбцы листа СВОДКА (ширины)
SUMMARY_COL_WIDTHS = [5, 30, 38, 14, 14, 16, 16, 16, 10]


# ═══════════════════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ═══════════════════════════════════════════════════════════════════════

def normalize_sev(sev: str) -> str:
    """Нормализует имя категории: старые → новые."""
    return SEV_NORMALIZE.get(sev, sev)


def get_sev_cfg(sev: str) -> dict:
    sev = normalize_sev(sev)
    return SEVERITY_CONFIG.get(sev, SEVERITY_CONFIG["ПРОВЕРИТЬ ПО СМЕЖНЫМ"])


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def cell(ws, row: int, col: int, value,
         bg: str = None, fg: str = "000000",
         bold: bool = False, italic: bool = False,
         wrap: bool = True,
         align_h: str = "left", align_v: str = "top",
         font_size: int = 10, border: bool = True):
    """Записывает значение в ячейку с форматированием."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, italic=italic, color=fg,
                  size=font_size, name="Calibri")
    if bg:
        c.fill = make_fill(bg)
    c.alignment = Alignment(wrap_text=wrap,
                            horizontal=align_h,
                            vertical=align_v)
    if border:
        c.border = thin_border()
    return c


# ═══════════════════════════════════════════════════════════════════════
#  ИЗВЛЕЧЕНИЕ ПОЛЕЙ ИЗ FINDING
# ═══════════════════════════════════════════════════════════════════════

def f_num(f, idx):         return idx
def f_finding_id(f, _):    return f.get("id", "")
def f_sheet(f, _):
    v = f.get("sheet") or f.get("location") or ""
    if isinstance(v, list):
        v = ", ".join(str(x) for x in v)
    if not v:
        # Штампованного листа нет (страница-продолжение / штамп не распознан) —
        # показываем хотя бы страницу PDF, чтобы столбец не был пустым.
        page = f.get("page")
        if isinstance(page, list) and page:
            v = "стр. PDF " + ", ".join(str(x) for x in page)
        elif isinstance(page, int):
            v = f"стр. PDF {page}"
    return v
def f_problem(f, _):
    """Короткое название проблемы (≤ 80 символов)."""
    s = (f.get("problem") or f.get("title") or f.get("finding") or "")
    return s

def f_description(f, _):
    """Подробное описание + норма."""
    parts = []
    desc = f.get("description") or f.get("finding") or ""
    if desc:
        parts.append(desc)
    norm = f.get("norm") or ""
    if norm:
        parts.append(f"Норма: {norm}")
    # MD/PDF расхождение (если есть)
    disc = f.get("md_pdf_discrepancy")
    if disc and isinstance(disc, dict) and disc.get("verdict"):
        v = disc["verdict"]
        if "ошибка в самом проекте" not in v:
            parts.append(f"[MD↔PDF: {v}]")
    return "\n".join(parts)

def f_solution(f, _):      return f.get("solution") or f.get("recommendation") or ""
def f_risk(f, _):          return f.get("risk") or f.get("consequence") or ""

def f_severity(f, _):
    sev = normalize_sev(f.get("severity") or "ПРОВЕРИТЬ ПО СМЕЖНЫМ")
    cfg = get_sev_cfg(sev)
    return f"{cfg['icon']} {sev}"


# Экспертные решения — подгружаются из expert_review.json
_expert_review_cache = {}  # folder_path -> {item_id: {decision, rejection_reason}}

def _expert_review_candidates(folder: str) -> list:
    """Кандидаты для expert_review.json (folder = version_dir).

    В projects_v2 файл лежит в `<version_dir>/04_review/`, в legacy — в
    `<version_dir>/_output/`. Порядок совпадает с каноническим резолвером
    knowledge_base_service._review_paths, иначе решения/причины «Возможный
    повтор» не попадают в Excel пакета выгрузки для V2-проектов.
    """
    return [
        os.path.join(folder, "04_review", "expert_review.json"),
        os.path.join(folder, "03_analysis", "latest", "expert_review.json"),
        os.path.join(folder, "_output", "expert_review.json"),
    ]


def _load_expert_review(folder: str) -> dict:
    """Загрузить решения эксперта для проекта (кешировано)."""
    if folder in _expert_review_cache:
        return _expert_review_cache[folder]
    result = {}
    path = next((p for p in _expert_review_candidates(folder) if os.path.isfile(p)), None)
    if path:
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                data = json.load(f)
            for d in data.get("decisions", []):
                result[d.get("item_id", "")] = {
                    "decision": d.get("decision", ""),
                    "rejection_reason": d.get("rejection_reason", ""),
                }
        except Exception:
            pass
    _expert_review_cache[folder] = result
    return result

def f_expert_decision(f, _, folder=""):
    review = _load_expert_review(folder)
    item_id = f.get("id", "")
    d = review.get(item_id, {})
    dec = d.get("decision", "")
    if dec == "accepted":
        return "Принято"
    elif dec == "rejected":
        return "Отклонено"
    return ""

def f_rejection_reason(f, _, folder=""):
    review = _load_expert_review(folder)
    item_id = f.get("id", "")
    d = review.get(item_id, {})
    return d.get("rejection_reason", "") or ""


FIELD_FUNCS = {
    "num":              f_num,
    "finding_id":       f_finding_id,
    "sheet":            f_sheet,
    "problem":          f_problem,
    "description":      f_description,
    "solution":         f_solution,
    "severity":         f_severity,
    "risk":             f_risk,
    "expert_decision":  f_expert_decision,
    "rejection_reason": f_rejection_reason,
}


# ═══════════════════════════════════════════════════════════════════════
#  СИСТЕМА ТИПОВ ОПТИМИЗАЦИИ
# ═══════════════════════════════════════════════════════════════════════

OPT_TYPE_CONFIG = {
    "cheaper_analog":  {"icon": "💰", "label": "Дешевле аналог",     "bg": "D5F5E3", "bg_alt": "E8F8EE", "fg": "1E8449"},
    "faster_install":  {"icon": "⚡", "label": "Быстрее монтаж",     "bg": "D6EAF8", "bg_alt": "E8F4FC", "fg": "1F618D"},
    "simpler_design":  {"icon": "🔧", "label": "Проще конструкция",  "bg": "FDEBD0", "bg_alt": "FEF5E7", "fg": "B9770E"},
    "lifecycle":       {"icon": "🔄", "label": "Жизн. цикл",        "bg": "E8DAEF", "bg_alt": "F2E8F7", "fg": "6C3483"},
}

OPT_COLUMNS = [
    ("num",              "№",                    5),
    ("id",               "ID",                  10),
    ("section",          "Раздел/Лист",         24),
    ("current",          "Текущее решение",     42),
    ("proposed",         "Предложение",         42),
    ("type",             "Тип",                 20),
    ("savings",          "Экономия",            12),
    ("timeline",         "Сроки",               14),
    ("risks",            "Риски",               32),
    ("expert_decision",  "Решение эксперта",    16),
    ("rejection_reason", "Причина отклонения",  32),
]

def opt_type_label(t):
    return OPT_TYPE_CONFIG.get(t, {}).get("label", t)

def opt_type_cfg(t):
    return OPT_TYPE_CONFIG.get(t, {"icon": "❓", "label": t, "bg": "F2F2F2", "bg_alt": "F8F8F8", "fg": "595959"})


# ═══════════════════════════════════════════════════════════════════════
#  ПОИСК И ЗАГРУЗКА ПРОЕКТОВ
# ═══════════════════════════════════════════════════════════════════════

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
# Централизованный источник пути к проектам (см. backend.app.core.config).
# Прежний BASE_DIR/"projects" указывал на несуществующий stages/report/projects.
try:
    from backend.app.core.config import PROJECTS_DIR as _CONFIG_PROJECTS_DIR
    PROJECTS_DIR = str(_CONFIG_PROJECTS_DIR)
except Exception:
    # standalone-fallback: корень репозитория (../../../../.. от report/)
    PROJECTS_DIR = os.path.abspath(
        os.path.join(BASE_DIR, "..", "..", "..", "..", "..", "projects")
    )
REPORTS_DIR  = os.path.join(BASE_DIR, "отчет")


def _iter_project_dirs(root=None):
    """Version-aware обход проектов (reserc.md #41).

    Делегирует каноническому project_service.iter_project_dirs, который корректно
    обрабатывает контейнеры версий `<база>(main)/` — отдаёт ОДНУ primary-версию
    (а не дубли всех версий, как делала прежняя ручная рекурсия) и берёт
    стабильный basename как project_id, читая `_output` нужной версии.

    Аргумент `root` сохранён для совместимости сигнатуры, но игнорируется:
    канонический обход берёт PROJECTS_DIR из конфига.
    """
    from backend.app.services.common.project_service import iter_project_dirs
    return [(pid, str(path)) for pid, path in iter_project_dirs()]


def _specific_context(raw_path: str) -> tuple[str, str, str]:
    d = os.path.abspath(raw_path)
    env_version_dir = audit_scope.get_version_dir()
    env_output_dir = audit_scope.get_output_dir()

    output_dir = None
    version_dir = d
    if env_output_dir and os.path.abspath(env_output_dir) == d:
        output_dir = d
        version_dir = os.path.abspath(env_version_dir) if env_version_dir else os.path.dirname(d)
    elif os.path.isfile(os.path.join(d, "03_findings.json")) or os.path.isfile(os.path.join(d, "optimization.json")):
        output_dir = d
        version_dir = os.path.abspath(env_version_dir) if env_version_dir else os.path.dirname(d)
    else:
        output_dir = os.path.join(d, "_output")
        if env_output_dir and os.path.isdir(env_output_dir):
            output_dir = os.path.abspath(env_output_dir)

    return os.path.basename(version_dir), version_dir, output_dir


def _project_info_path(version_dir: str) -> str:
    try:
        from backend.app.services.storage.projects_v2_source_resolver import resolve_project_info_path

        resolved = resolve_project_info_path(version_dir)
        if resolved is not None:
            return str(resolved)
    except Exception:
        pass
    return os.path.join(version_dir, "project_info.json")


def find_projects(specific_paths=None) -> list:
    results = []
    if specific_paths:
        dirs = [_specific_context(p) for p in specific_paths]
    else:
        if not os.path.isdir(PROJECTS_DIR):
            print(f"[ERR] Папка projects/ не найдена: {PROJECTS_DIR}")
            return results
        dirs = [(pid, path, os.path.join(path, "_output")) for pid, path in _iter_project_dirs(PROJECTS_DIR)]
    for pid, d, output_dir in dirs:
        fp  = os.path.join(output_dir, "03_findings.json")
        op  = os.path.join(output_dir, "optimization.json")
        ip  = _project_info_path(d)
        # Если передан путь к версионной папке (`_versions/v2`) или v2 output-dir,
        # basename даёт техническое имя. Читаем настоящий project_id из project_info.
        if os.path.isfile(ip):
            try:
                with open(ip, "r", encoding="utf-8-sig") as _f:
                    _pi = json.load(_f)
                _real_pid = (_pi.get("project_id") or _pi.get("document_code") or "").strip()
                if _real_pid:
                    pid = _real_pid
            except Exception:
                pass
        # Имя Excel-листа (≤ 31 символ, без спецсимволов)
        sheet_name = pid.replace("/", "-").replace("\\", "-")
        for ch in r'*?[]':
            sheet_name = sheet_name.replace(ch, "")
        sheet_name = sheet_name[:31]
        results.append({
            "project_id":       pid,
            "folder":           d,
            "findings_path":    fp,
            "optimization_path": op,
            "info_path":        ip,
            "has_findings":     os.path.isfile(fp),
            "has_optimization": os.path.isfile(op),
            "sheet_name":       sheet_name,
        })
    return results


def load_json(path: str) -> dict:
    # utf-8-sig обрабатывает и чистый UTF-8, и UTF-8 с BOM (PowerShell)
    with open(path, "r", encoding="utf-8-sig") as f:
        return json.load(f)


def normalize_findings_sheets(findings: list, findings_path: str) -> int:
    """Перекрыть sheet=НАЗВАНИЕ листа номером из document_graph.json.

    Старые прогоны stage 02 читали граф по v1-ключу sheet_no (в v2 он всегда
    None) и клали в поле sheet НАЗВАНИЕ листа из штампа. Такое значение
    непустое, поэтому ни бэкфилл в findings_merge, ни фолбэк «стр. PDF N» в
    f_sheet не срабатывали — в столбце «Лист/Раздел» оказывалось название без
    номера и без страницы. Правим display-side, файл аудита не трогаем.
    """
    if not findings:
        return 0
    try:
        from backend.app.pipeline.stages.prepare.graph_builder import (
            build_block_page_index, build_md_line_page_index, build_page_sheet_map,
            looks_like_sheet_ref, resolve_document_markdown,
            resolve_finding_sheet_label,
        )
    except Exception:
        return 0

    graph_path = os.path.join(os.path.dirname(findings_path), "document_graph.json")
    if not os.path.isfile(graph_path):
        return 0
    try:
        graph = load_json(graph_path)
    except Exception:
        return 0

    page_to_sheet = build_page_sheet_map(graph)
    if not page_to_sheet:
        return 0
    block_to_page = build_block_page_index(graph)
    md_line_index = []
    md_path = resolve_document_markdown(os.path.dirname(findings_path))
    if md_path is not None:
        try:
            md_line_index = build_md_line_page_index(md_path.read_text(encoding="utf-8"))
        except OSError:
            pass

    fixed = 0
    for f in findings:
        if not isinstance(f, dict):
            continue
        sheet = f.get("sheet")
        if sheet and looks_like_sheet_ref(sheet):
            continue
        label = resolve_finding_sheet_label(f, page_to_sheet, block_to_page, md_line_index)
        if not label:
            continue
        if sheet:
            f.setdefault("sheet_title", str(sheet).strip())
        f["sheet"] = label
        fixed += 1
    return fixed


# ═══════════════════════════════════════════════════════════════════════
#  ЛИСТ СВОДКА
# ═══════════════════════════════════════════════════════════════════════

def build_summary_sheet(wb, projects_data: list):
    ws = wb.active
    ws.title = "СВОДКА"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85

    # Ширины столбцов (A…I)
    for i, w in enumerate(SUMMARY_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Строка 1: шапка ───────────────────────────────────────────────
    row = 1
    headers = [
        "№", "Проект (ID)", "Объект / Раздел",
        "🔴 Крит.", "🟠 Эконом.", "🟡 Эксплуат.",
        "🔵 Рекомен.", "⚪ По смежным", "Итого"
    ]
    for col, h in enumerate(headers, 1):
        cell(ws, row, col, h,
             bg=HEADER_BG, fg=HEADER_FG, bold=True,
             align_h="center", align_v="center",
             font_size=10)
    ws.row_dimensions[row].height = 24

    # ── Строка 2: заголовок отчёта ────────────────────────────────────
    row = 2
    ws.merge_cells(f"A{row}:I{row}")
    hdr = ws[f"A{row}"]
    hdr.value = (
        f"СВОДНЫЙ ОТЧЁТ АУДИТА   |   "
        f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}   |   "
        f"Аудит проектной документации"
    )
    hdr.font  = Font(bold=True, size=11, color=PROJ_HDR_FG, name="Calibri")
    hdr.fill  = make_fill(PROJ_HDR_BG)
    hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

    # AutoFilter по строке 1
    ws.auto_filter.ref = f"A1:I1"

    # ── Строки проектов ───────────────────────────────────────────────
    total_by_sev = {s: 0 for s in SEV_ORDER}

    for idx, pd in enumerate(projects_data, 1):
        row += 1
        ws.row_dimensions[row].height = 28

        meta     = pd.get("meta_json", {})
        pinfo    = pd.get("project_info", {})
        findings = pd.get("findings_json", {}).get("findings", [])

        # Подсчёт по категориям
        by_sev = {}
        for f in findings:
            sev = normalize_sev(f.get("severity") or "ПРОВЕРИТЬ ПО СМЕЖНЫМ")
            by_sev[sev] = by_sev.get(sev, 0) + 1

        obj_name = (pinfo.get("object") or pinfo.get("description") or "—")
        row_bg   = "F7F9FC" if idx % 2 == 0 else "FFFFFF"

        # Номер
        cell(ws, row, 1, idx, bg=row_bg, align_h="center")

        # ID проекта — гиперссылка на лист
        pid_cell = cell(ws, row, 2, pd["project_id"], bg=row_bg, bold=True)
        try:
            sn = pd.get("sheet_name", pd["project_id"])
            pid_cell.hyperlink = f"#'{sn}'!A1"
            pid_cell.font = Font(bold=True, color="1F497D", underline="single",
                                 size=10, name="Calibri")
        except Exception:
            pass

        cell(ws, row, 3, obj_name, bg=row_bg)

        if pd["has_findings"]:
            vals = [
                by_sev.get("КРИТИЧЕСКОЕ", 0),
                by_sev.get("ЭКОНОМИЧЕСКОЕ", 0),
                by_sev.get("ЭКСПЛУАТАЦИОННОЕ", 0),
                by_sev.get("РЕКОМЕНДАТЕЛЬНОЕ", 0),
                by_sev.get("ПРОВЕРИТЬ ПО СМЕЖНЫМ", 0),
            ]
            bgs = ["FFCCCC", "FCE4D6", "FFFACD", "DDEEFF", "F2F2F2"]
            for ci, (v, bg_sev) in enumerate(zip(vals, bgs), 4):
                cell(ws, row, ci,
                     v if v else "—",
                     bg=bg_sev if v else row_bg,
                     align_h="center")
            total = sum(v for v in vals if isinstance(v, int))
            cell(ws, row, 9, total, bg=row_bg, bold=True, align_h="center")

            for sev, cnt in by_sev.items():
                if sev in total_by_sev:
                    total_by_sev[sev] += cnt
        else:
            for c_i in range(4, 10):
                cell(ws, row, c_i, "нет аудита",
                     bg="EEEEEE", fg="999999",
                     align_h="center", italic=True, font_size=9)

    # ── Итоговая строка ───────────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:C{row}")
    cell(ws, row, 1, "ИТОГО",
         bg=TOTAL_BG, fg="FFFFFF", bold=True, align_h="right")
    sev_bgs = ["FFCCCC", "FCE4D6", "FFFACD", "DDEEFF", "F2F2F2"]
    for ci, (sev, bg_sev) in enumerate(zip(SEV_ORDER, sev_bgs), 4):
        v = total_by_sev.get(sev, 0)
        cell(ws, row, ci, v if v else "—",
             bg=bg_sev, bold=True, align_h="center")
    grand = sum(total_by_sev.values())
    cell(ws, row, 9, grand,
         bg="D9D9D9", bold=True, align_h="center")

    # ── Легенда ───────────────────────────────────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:I{row}")
    leg_hdr = ws[f"A{row}"]
    leg_hdr.value = "ЛЕГЕНДА КАТЕГОРИЙ"
    leg_hdr.font  = Font(bold=True, color=HEADER_FG, size=10, name="Calibri")
    leg_hdr.fill  = make_fill(HEADER_BG)
    leg_hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18

    for sev in SEV_ORDER:
        cfg = SEVERITY_CONFIG[sev]
        row += 1
        # цветная плашка
        ws.merge_cells(f"A{row}:B{row}")
        cell(ws, row, 1,
             f"{cfg['icon']}  {sev}",
             bg=cfg["bg"], fg=cfg["fg"],
             bold=True, align_h="center", font_size=10)
        ws.merge_cells(f"C{row}:I{row}")
        cell(ws, row, 3,
             cfg.get("desc", ""),
             bg="FAFAFA", fg="333333",
             align_h="left", font_size=9)
        ws.row_dimensions[row].height = 18

    # ── Дата формирования отчёта ────────────────────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:I{row}")
    dt_cell = ws[f"A{row}"]
    dt_cell.value = f"Отчёт сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    dt_cell.font  = Font(italic=True, size=9, color="666666", name="Calibri")
    dt_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 16

    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════
#  ЛИСТ ПРОЕКТА
# ═══════════════════════════════════════════════════════════════════════

def build_project_sheet(wb, pd_entry: dict):
    project_id = pd_entry["project_id"]
    pinfo      = pd_entry.get("project_info", {})
    data       = pd_entry.get("findings_json", {})
    findings   = data.get("findings", [])
    meta       = data.get("meta", {})
    sheet_name = pd_entry["sheet_name"]

    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    last_col_letter = get_column_letter(len(PROJ_COLUMNS))

    # ── Установка ширин столбцов ──────────────────────────────────────
    for i, (_, _, width) in enumerate(PROJ_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Строка 1: заголовки столбцов ─────────────────────────────────
    row = 1
    for col, (key, header, _) in enumerate(PROJ_COLUMNS, 1):
        cell(ws, row, col, header,
             bg=HEADER_BG, fg=HEADER_FG, bold=True,
             align_h="center" if key == "num" else "left",
             align_v="center", font_size=10)
    ws.row_dimensions[row].height = 24

    # AutoFilter
    ws.auto_filter.ref = f"A{row}:{last_col_letter}{row}"

    # ── Строка 2: имя и дата аудита ───────────────────────────────────
    row = 2
    audit_dt = meta.get("audit_completed", "")
    if audit_dt:
        try:
            audit_dt = datetime.fromisoformat(audit_dt).strftime("%d.%m.%Y")
        except Exception:
            pass

    obj = pinfo.get("object") or pinfo.get("description") or ""
    total_cnt = meta.get("total_findings", len(findings))
    report_dt = datetime.now().strftime("%d.%m.%Y %H:%M")
    proj_label = project_id
    if audit_dt:
        proj_label += f"  |  аудит от {audit_dt}"
    if obj:
        proj_label += f"  |  {obj}"
    if total_cnt:
        proj_label += f"  |  замечаний: {total_cnt}"
    proj_label += f"  |  отчёт: {report_dt}"

    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    hdr2 = ws[f"A{row}"]
    hdr2.value     = proj_label
    hdr2.font      = Font(bold=True, size=11, color=PROJ_HDR_FG, name="Calibri")
    hdr2.fill      = make_fill(PROJ_HDR_BG)
    hdr2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

    # Скрытый столбец: полный project_id (для обратной загрузки решений)
    pid_col = len(PROJ_COLUMNS) + 1
    pid_letter = get_column_letter(pid_col)
    cell(ws, 1, pid_col, "project_id", bg=HEADER_BG, fg=HEADER_FG, bold=True, font_size=8)
    cell(ws, 2, pid_col, project_id, bg=PROJ_HDR_BG, fg=PROJ_HDR_FG, font_size=8)
    ws.column_dimensions[pid_letter].hidden = True
    ws.column_dimensions[pid_letter].width = 1

    # ── Нет данных ────────────────────────────────────────────────────
    if not findings:
        row += 1
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        c = ws[f"A{row}"]
        c.value = "Аудит не завершён — файл 03_findings.json отсутствует или не содержит замечаний."
        c.font  = Font(italic=True, color="888888", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A3"
        return

    # ── Строки замечаний ──────────────────────────────────────────────
    for idx, finding in enumerate(findings, 1):
        row += 1
        try:
            sev     = normalize_sev(finding.get("severity") or "ПРОВЕРИТЬ ПО СМЕЖНЫМ")
            cfg     = get_sev_cfg(sev)
            row_bg  = cfg["bg"] if idx % 2 != 0 else cfg["bg_alt"]

            for col, (key, _, _) in enumerate(PROJ_COLUMNS, 1):
                func = FIELD_FUNCS[key]
                if key in ("expert_decision", "rejection_reason"):
                    val = func(finding, idx, folder=pd_entry.get("folder", ""))
                else:
                    val = func(finding, idx)

                is_num = (key == "num")
                is_sev = (key == "severity")
                is_sm  = key in ("description", "norm")

                cell(ws, row, col, val,
                     bg=row_bg,
                     fg=cfg["fg"] if is_sev else "000000",
                     bold=is_num or is_sev,
                     align_h="center" if is_num else "left",
                     align_v="top",
                     font_size=9 if is_sm else 10)

            # Авто-высота по длине текста
            max_len = max(
                len(str(finding.get("description") or finding.get("finding") or "")),
                len(str(finding.get("solution")    or finding.get("recommendation") or "")),
                len(str(finding.get("norm") or ""))
            )
            ws.row_dimensions[row].height = max(35, min(130, max_len // 2))
        except Exception as e:
            # Один битый finding не должен ломать весь Excel
            cell(ws, row, 1, idx, bg="FFFFFF")
            cell(ws, row, 2, f"[Ошибка чтения замечания: {e}]", bg="FFEEEE")
            print(f"  [WARN] Пропущено замечание #{idx}: {e}")

    # ── Итоговая мини-строка ──────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 18
    by_sev_cnt = {}
    for f in findings:
        s = normalize_sev(f.get("severity") or "ПРОВЕРИТЬ ПО СМЕЖНЫМ")
        by_sev_cnt[s] = by_sev_cnt.get(s, 0) + 1

    summary_parts = []
    for sev in SEV_ORDER:
        cnt = by_sev_cnt.get(sev, 0)
        if cnt:
            cfg = SEVERITY_CONFIG[sev]
            summary_parts.append(f"{cfg['icon']} {sev}: {cnt}")

    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    summary_cell = ws[f"A{row}"]
    summary_cell.value = "  |  ".join(summary_parts) if summary_parts else ""
    summary_cell.font  = Font(italic=True, size=9, color="444444", name="Calibri")
    summary_cell.fill  = make_fill("EEF2F7")
    summary_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_cell.border = thin_border()

    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════
#  ЛИСТ ОПТИМИЗАЦИИ
# ═══════════════════════════════════════════════════════════════════════

def build_optimization_summary_sheet(wb, projects_data: list):
    """Сводный лист оптимизации (аналог СВОДКА для замечаний)."""
    ws = wb.active if wb.active.title == "Sheet" else wb.create_sheet()
    ws.title = "СВОДКА"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85

    col_widths = [5, 30, 38, 14, 14, 14, 14, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    headers = [
        "№", "Проект (ID)", "Объект / Раздел",
        "💰 Аналог", "⚡ Монтаж", "🔧 Конструкция",
        "🔄 Жизн.цикл", "Итого"
    ]
    for col, h in enumerate(headers, 1):
        cell(ws, row, col, h, bg=HEADER_BG, fg=HEADER_FG, bold=True,
             align_h="center", align_v="center", font_size=10)
    ws.row_dimensions[row].height = 24

    row = 2
    ws.merge_cells(f"A{row}:H{row}")
    hdr = ws[f"A{row}"]
    hdr.value = (
        f"СВОДНЫЙ ОТЧЁТ ОПТИМИЗАЦИИ   |   "
        f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}   |   "
        f"Сценарии оптимизации проектных решений"
    )
    hdr.font = Font(bold=True, size=11, color=PROJ_HDR_FG, name="Calibri")
    hdr.fill = make_fill("1B7F4B")
    hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

    ws.auto_filter.ref = "A1:H1"
    total_by_type = {"cheaper_analog": 0, "faster_install": 0, "simpler_design": 0, "lifecycle": 0}

    for idx, pd in enumerate(projects_data, 1):
        row += 1
        ws.row_dimensions[row].height = 28
        pinfo = pd.get("project_info", {})
        opt_data = pd.get("optimization_json", {})
        items = opt_data.get("items", [])
        meta = opt_data.get("meta", {})
        by_type = meta.get("by_type", {})

        obj_name = pinfo.get("object") or pinfo.get("description") or "—"
        row_bg = "F7F9FC" if idx % 2 == 0 else "FFFFFF"

        cell(ws, row, 1, idx, bg=row_bg, align_h="center")
        pid_cell = cell(ws, row, 2, pd["project_id"], bg=row_bg, bold=True)
        try:
            sn = "ОПТ " + pd.get("sheet_name", pd["project_id"])[:27]
            pid_cell.hyperlink = f"#'{sn}'!A1"
            pid_cell.font = Font(bold=True, color="1B7F4B", underline="single", size=10, name="Calibri")
        except Exception:
            pass
        cell(ws, row, 3, obj_name, bg=row_bg)

        if pd["has_optimization"] and items:
            type_keys = ["cheaper_analog", "faster_install", "simpler_design", "lifecycle"]
            type_bgs = ["D5F5E3", "D6EAF8", "FDEBD0", "E8DAEF"]
            for ci, (tk, bg_t) in enumerate(zip(type_keys, type_bgs), 4):
                v = by_type.get(tk, 0)
                cell(ws, row, ci, v if v else "—", bg=bg_t if v else row_bg, align_h="center")
                total_by_type[tk] = total_by_type.get(tk, 0) + v
            total = sum(by_type.get(k, 0) for k in type_keys)
            cell(ws, row, 8, total, bg=row_bg, bold=True, align_h="center")
        else:
            for c_i in range(4, 9):
                cell(ws, row, c_i, "нет данных", bg="EEEEEE", fg="999999",
                     align_h="center", italic=True, font_size=9)

    # Итого
    row += 1
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:C{row}")
    cell(ws, row, 1, "ИТОГО", bg=TOTAL_BG, fg="FFFFFF", bold=True, align_h="right")
    type_bgs = ["D5F5E3", "D6EAF8", "FDEBD0", "E8DAEF"]
    type_keys = ["cheaper_analog", "faster_install", "simpler_design", "lifecycle"]
    for ci, (tk, bg_t) in enumerate(zip(type_keys, type_bgs), 4):
        v = total_by_type.get(tk, 0)
        cell(ws, row, ci, v if v else "—", bg=bg_t, bold=True, align_h="center")
    cell(ws, row, 8, sum(total_by_type.values()), bg="D9D9D9", bold=True, align_h="center")

    # Легенда
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    leg_hdr = ws[f"A{row}"]
    leg_hdr.value = "ЛЕГЕНДА ТИПОВ ОПТИМИЗАЦИИ"
    leg_hdr.font = Font(bold=True, color=HEADER_FG, size=10, name="Calibri")
    leg_hdr.fill = make_fill(HEADER_BG)
    leg_hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18

    for tk in type_keys:
        cfg = OPT_TYPE_CONFIG[tk]
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        cell(ws, row, 1, f"{cfg['icon']}  {cfg['label']}", bg=cfg["bg"], fg=cfg["fg"],
             bold=True, align_h="center", font_size=10)
        ws.merge_cells(f"C{row}:H{row}")
        descs = {
            "cheaper_analog": "Замена на более дешёвый аналог без потери качества",
            "faster_install": "Упрощение монтажа, сокращение сроков",
            "simpler_design": "Упрощение конструктивных решений",
            "lifecycle": "Оптимизация стоимости жизненного цикла",
        }
        cell(ws, row, 3, descs.get(tk, ""), bg="FAFAFA", fg="333333", align_h="left", font_size=9)
        ws.row_dimensions[row].height = 18

    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    dt_cell = ws[f"A{row}"]
    dt_cell.value = f"Отчёт сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    dt_cell.font = Font(italic=True, size=9, color="666666", name="Calibri")
    dt_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 16

    ws.freeze_panes = "A3"


def build_optimization_project_sheet(wb, pd_entry: dict):
    """Лист оптимизации одного проекта."""
    project_id = pd_entry["project_id"]
    pinfo = pd_entry.get("project_info", {})
    opt_data = pd_entry.get("optimization_json", {})
    items = opt_data.get("items", [])
    meta = opt_data.get("meta", {})

    sheet_name = "ОПТ " + pd_entry["sheet_name"][:27]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    last_col_letter = get_column_letter(len(OPT_COLUMNS))

    for i, (_, _, width) in enumerate(OPT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Строка 1: заголовки
    row = 1
    for col, (key, header, _) in enumerate(OPT_COLUMNS, 1):
        cell(ws, row, col, header, bg=HEADER_BG, fg=HEADER_FG, bold=True,
             align_h="center" if key in ("num", "savings") else "left",
             align_v="center", font_size=10)
    ws.row_dimensions[row].height = 24
    ws.auto_filter.ref = f"A{row}:{last_col_letter}{row}"

    # Строка 2: заголовок проекта
    row = 2
    total_cnt = meta.get("total_items", len(items))
    savings_pct = meta.get("estimated_savings_pct", 0)
    report_dt = datetime.now().strftime("%d.%m.%Y %H:%M")
    proj_label = f"{project_id}  |  оптимизаций: {total_cnt}"
    if savings_pct:
        proj_label += f"  |  экономия: −{savings_pct}%"
    proj_label += f"  |  отчёт: {report_dt}"

    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    hdr2 = ws[f"A{row}"]
    hdr2.value = proj_label
    hdr2.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    hdr2.fill = make_fill("1B7F4B")
    hdr2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

    # Скрытый столбец: полный project_id (для обратной загрузки решений)
    pid_col = len(OPT_COLUMNS) + 1
    pid_letter = get_column_letter(pid_col)
    cell(ws, 1, pid_col, "project_id", bg=HEADER_BG, fg=HEADER_FG, bold=True, font_size=8)
    cell(ws, 2, pid_col, project_id, bg="1B7F4B", fg="FFFFFF", font_size=8)
    ws.column_dimensions[pid_letter].hidden = True
    ws.column_dimensions[pid_letter].width = 1

    if not items:
        row += 1
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        c = ws[f"A{row}"]
        c.value = "Оптимизация не выполнена — файл optimization.json отсутствует или пуст."
        c.font = Font(italic=True, color="888888", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A3"
        return

    # Строки данных
    for idx, item in enumerate(items, 1):
        row += 1
        try:
            opt_type = item.get("type", "") if isinstance(item, dict) else ""
            cfg = opt_type_cfg(opt_type)
            row_bg = cfg["bg"] if idx % 2 != 0 else cfg["bg_alt"]

            # Загрузить экспертные решения для оптимизаций
            review = _load_expert_review(pd_entry.get("folder", ""))
            item_review = review.get(item.get("id", ""), {})
            dec = item_review.get("decision", "")
            dec_label = "Принято" if dec == "accepted" else ("Отклонено" if dec == "rejected" else "")
            rej_reason = item_review.get("rejection_reason", "") or ""

            vals = [
                idx,
                item.get("id", ""),
                item.get("section", ""),
                item.get("current", ""),
                item.get("proposed", ""),
                f"{cfg['icon']} {cfg['label']}",
                f"{item.get('savings_pct', 0)}%" if item.get("savings_pct") else "—",
                item.get("timeline_impact", ""),
                item.get("risks", ""),
                dec_label,
                rej_reason,
            ]
        except Exception as e:
            cell(ws, row, 1, idx, bg="FFFFFF")
            cell(ws, row, 2, f"[Ошибка чтения оптимизации: {e}]", bg="FFEEEE")
            print(f"  [WARN] Пропущена оптимизация #{idx}: {e}")
            continue

        for col, (key, _, _) in enumerate(OPT_COLUMNS, 1):
            val = vals[col - 1]
            is_num = (key == "num")
            is_type = (key == "type")
            cell(ws, row, col, val, bg=row_bg,
                 fg=cfg["fg"] if is_type else "000000",
                 bold=is_num or is_type,
                 align_h="center" if key in ("num", "savings") else "left",
                 align_v="top", font_size=10)

        max_len = max(len(str(item.get("current", ""))), len(str(item.get("proposed", ""))))
        ws.row_dimensions[row].height = max(35, min(130, max_len // 2))

    # Итоговая строка
    row += 1
    ws.row_dimensions[row].height = 18
    by_type = {}
    for item in items:
        t = item.get("type", "unknown")
        by_type[t] = by_type.get(t, 0) + 1
    parts = []
    for tk in ["cheaper_analog", "faster_install", "simpler_design", "lifecycle"]:
        cnt = by_type.get(tk, 0)
        if cnt:
            cfg = OPT_TYPE_CONFIG.get(tk, {})
            parts.append(f"{cfg.get('icon', '')} {cfg.get('label', tk)}: {cnt}")
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    summary_cell = ws[f"A{row}"]
    summary_cell.value = "  |  ".join(parts) if parts else ""
    summary_cell.font = Font(italic=True, size=9, color="444444", name="Calibri")
    summary_cell.fill = make_fill("E8F5E9")
    summary_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_cell.border = thin_border()

    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════
#  ГЛАВНАЯ ФУНКЦИЯ
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Генерация Excel-отчёта по результатам аудита проектной документации"
    )
    parser.add_argument(
        "projects", nargs="*",
        help="Папки конкретных проектов (по умолчанию — все в projects/)"
    )
    parser.add_argument(
        "--out", "-o", default=None,
        help="Имя/путь выходного файла (по умолч.: audit_report_YYYYMMDD_HHMM.xlsx)"
    )
    parser.add_argument(
        "--no-summary", action="store_true",
        help="Не создавать лист СВОДКА"
    )
    parser.add_argument(
        "--type", choices=["findings", "optimization", "all"], default="all",
        help="Тип отчёта: findings (замечания), optimization (оптимизации), all (всё)"
    )
    args = parser.parse_args()

    # ── Найти проекты ─────────────────────────────────────────────────
    projects = find_projects(args.projects if args.projects else None)
    if not projects:
        print("[ERR] Проекты не найдены.")
        sys.exit(1)

    print(f"\n{'='*62}")
    print(f"  Генерация Excel-отчёта  |  проектов: {len(projects)}")
    print(f"{'='*62}")

    report_type = args.type  # findings | optimization | all

    # ── Загрузить данные ──────────────────────────────────────────────
    for p in projects:
        p["project_info"] = {}
        p["findings_json"] = {}
        p["optimization_json"] = {}
        p["meta_json"] = {}

        if os.path.isfile(p["info_path"]):
            try:
                p["project_info"] = load_json(p["info_path"])
            except Exception:
                pass

        if report_type in ("findings", "all") and p["has_findings"]:
            try:
                fj = load_json(p["findings_path"])
                p["findings_json"] = fj
                p["meta_json"]     = fj.get("meta", {})
                cnt = len(fj.get("findings", []))
                fixed = normalize_findings_sheets(fj.get("findings", []), p["findings_path"])
                suffix = f"  (лист восстановлен у {fixed})" if fixed else ""
                print(f"  [OK]  {p['project_id']:32s}  {cnt} замечаний{suffix}")
            except Exception as e:
                print(f"  [!!]  {p['project_id']:32s}  Ошибка: {e}")

        if report_type in ("optimization", "all") and p["has_optimization"]:
            try:
                oj = load_json(p["optimization_path"])
                p["optimization_json"] = oj
                cnt = len(oj.get("items", []))
                print(f"  [OK]  {p['project_id']:32s}  {cnt} оптимизаций")
            except Exception as e:
                print(f"  [!!]  {p['project_id']:32s}  Ошибка оптимизации: {e}")

        if not p["has_findings"] and report_type in ("findings", "all"):
            print(f"  [--]  {p['project_id']:32s}  нет 03_findings.json")
        if not p["has_optimization"] and report_type in ("optimization", "all"):
            print(f"  [--]  {p['project_id']:32s}  нет optimization.json")

    # ── Создать книгу ─────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    if report_type == "findings":
        if not args.no_summary:
            build_summary_sheet(wb, projects)
        for p in projects:
            build_project_sheet(wb, p)
    elif report_type == "optimization":
        build_optimization_summary_sheet(wb, projects)
        for p in projects:
            build_optimization_project_sheet(wb, p)
    else:  # all
        if not args.no_summary:
            build_summary_sheet(wb, projects)
        for p in projects:
            build_project_sheet(wb, p)
        for p in projects:
            build_optimization_project_sheet(wb, p)

    # ── Убрать пустой дефолтный лист openpyxl ────────────────────────
    # `Workbook()` всегда создаёт активный лист "Sheet". Его занимает под себя
    # СВОДКА, но с `--no-summary` (так вызывает audit-package) summary не
    # строится, и в книге остаётся пустая вкладка "Sheet".
    for _ws in list(wb.worksheets):
        if _ws.title == "Sheet" and _ws.max_row == 1 and _ws.max_column == 1 \
                and _ws["A1"].value is None and len(wb.worksheets) > 1:
            wb.remove(_ws)

    # ── Сохранить ─────────────────────────────────────────────────────
    if args.out:
        out_path = (args.out if os.path.isabs(args.out)
                    else os.path.join(BASE_DIR, args.out))
    else:
        ts = datetime.now().strftime("%d.%m.%Y")
        os.makedirs(REPORTS_DIR, exist_ok=True)
        out_path = os.path.join(REPORTS_DIR, f"audit_report_{ts}.xlsx")

    try:
        wb.save(out_path)
        print(f"\n  Файл сохранён: {out_path}")
        print(f"{'='*62}\n")
    except PermissionError as e:
        print(f"\n  [ERR] Не удалось сохранить файл (занят?): {e}")
        print(f"{'='*62}\n")
        sys.exit(1)
    except Exception as e:
        print(f"\n  [ERR] Ошибка при сохранении Excel: {e}")
        print(f"{'='*62}\n")
        sys.exit(1)

    # Автооткрытие только при ручном запуске (не из webapp pipeline)
    if os.environ.get("AUDIT_NO_OPEN") != "1":
        try:
            os.startfile(out_path)
            print("  Excel открыт автоматически.")
        except AttributeError:
            pass
        except Exception as e:
            print(f"  (Автооткрытие: {e})")


if __name__ == "__main__":
    main()
