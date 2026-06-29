"""Выгрузка 58 подтверждённых кандидатов в классическую Excel-таблицу замечаний."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent / "results" / "audit_alia"

SEV_FILL = {
    "КРИТИЧЕСКОЕ": "FFC7CE",
    "ЭКСПЛУАТАЦИОННОЕ": "FFEB9C",
    "ЭКОНОМИЧЕСКОЕ": "DDEBF7",
    "РЕКОМЕНДАТЕЛЬНОЕ": "E2EFDA",
    "ПРОВЕРИТЬ ПО СМЕЖНЫМ": "EDEDED",
}

COLS = [
    ("№", 5),
    ("Дисциплина", 11),
    ("Документ / лист", 22),
    ("ID", 8),
    ("Категория", 16),
    ("Суть замечания", 46),
    ("Обоснование эксперта (почему отклонил)", 46),
    ("Что найдено против эксперта", 46),
    ("ВЫВОД: почему эксперт мог ошибиться", 52),
]


def _clean_finding(s: str) -> str:
    return (s or "").replace("прочитано с чертежа:", "С чертежа:").replace("; вывод:", "\n→ ").replace("довод:", "Довод:").replace("; пояснение:", "\n→ ")


def main() -> int:
    data = json.loads((OUT_DIR / "arbiter_sonnet_result.json").read_text(encoding="utf-8"))
    rows = data["survivors"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Замечания — перепроверка"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Заголовок-титул
    ncol = len(COLS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(1, 1, "Аудит проектной документации «Алия (ASTERUS)» — кандидаты «эксперт мог ошибиться» "
                      f"(подтверждено арбитром Claude Sonnet: {len(rows)})")
    t.font = Font(bold=True, size=13)
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34

    # Шапка
    for j, (name, width) in enumerate(COLS, 1):
        c = ws.cell(2, j, name)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.row_dimensions[2].height = 40

    # сгруппировать по проекту (дисциплина + документ), внутри — по критичности
    rows = sorted(rows, key=lambda r: (r.get("discipline", ""), r.get("document", ""), -(r.get("_score") or 0)))
    grp_fill = PatternFill("solid", fgColor="D9E1F2")
    grp_font = Font(bold=True, size=11, color="1F4E78")

    # Данные с разделителями по проектам
    rr = 2
    num = 0
    cur = None
    from collections import Counter
    counts = Counter((r.get("discipline", ""), r.get("document", "")) for r in rows)
    for r in rows:
        key = (r.get("discipline", ""), r.get("document", ""))
        if key != cur:
            cur = key
            rr += 1
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=ncol)
            gc = ws.cell(rr, 1, f"▶ {key[0]} / {key[1]}   —   замечаний: {counts[key]}")
            gc.fill = grp_fill; gc.font = grp_font
            gc.alignment = Alignment(horizontal="left", vertical="center")
            for j in range(1, ncol + 1):
                ws.cell(rr, j).border = border
            ws.row_dimensions[rr].height = 22
        rr += 1
        num += 1
        doc = f"{r.get('document','')}"
        if r.get("version"):
            doc += f" ({r['version']})"
        vals = [
            num, r.get("discipline", ""), doc, r.get("item_id", ""), r.get("_sev", ""),
            (r.get("problem") or "").strip(),
            (r.get("rejection_reason") or "").strip(),
            _clean_finding(r.get("assistant_finding", "")).strip(),
            (r.get("arbiter_reason") or "").strip(),
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(rr, j, v)
            c.border = border
            c.alignment = center if j in (1, 2, 4) else wrap_top
        fill = SEV_FILL.get(r.get("_sev", ""))
        if fill:
            ws.cell(rr, 5).fill = PatternFill("solid", fgColor=fill)
        longest = max(len(str(v)) for v in vals[5:])
        ws.row_dimensions[rr].height = min(220, max(48, longest / 46 * 15))

    ws.freeze_panes = "A3"

    out = OUT_DIR / "EXPERT_AUDIT_ALIA_FINAL.xlsx"
    wb.save(out)
    print(f"[excel] {len(rows)} замечаний → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
