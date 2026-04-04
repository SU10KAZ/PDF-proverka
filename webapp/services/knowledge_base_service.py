"""
Сервис базы знаний — сбор экспертных решений, хранение, анализ паттернов.
"""
import json
import os
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from webapp.config import KNOWLEDGE_BASE_DIR, DECISIONS_LOG_FILE, PATTERNS_FILE
from webapp.models.expert_review import (
    ExpertDecision, KnowledgeBaseEntry, PatternSuggestion,
)
from webapp.services.project_service import resolve_project_dir


def _ensure_kb_dir():
    """Создать папку knowledge_base/ если не существует."""
    KNOWLEDGE_BASE_DIR.mkdir(parents=True, exist_ok=True)


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")


# ═══════════════════════════════════════════════════════════════════════════
# Чтение / запись JSON
# ═══════════════════════════════════════════════════════════════════════════

def _load_json(path: Path) -> dict | list | None:
    if not path.exists():
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, UnicodeDecodeError):
        return None


def _save_json(path: Path, data):
    _ensure_kb_dir()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ═══════════════════════════════════════════════════════════════════════════
# Экспертная оценка (per-project)
# ═══════════════════════════════════════════════════════════════════════════

def save_expert_review(project_id: str, decisions: list[ExpertDecision], reviewer: str = "") -> dict:
    """Сохранить решения эксперта по проекту.

    1. Записывает expert_review.json в _output/ проекта
    2. Обогащает решения контекстом из findings/optimization
    3. Добавляет записи в глобальный decisions_log.json
    """
    project_dir = resolve_project_dir(project_id)
    output_dir = project_dir / "_output"
    output_dir.mkdir(parents=True, exist_ok=True)

    # 1. Сохранить per-project файл
    review_data = {
        "project_id": project_id,
        "reviewer": reviewer,
        "reviewed_at": _now_iso(),
        "decisions": [d.model_dump() for d in decisions],
    }
    _save_json(output_dir / "expert_review.json", review_data)

    # 2. Обогатить и добавить в глобальный лог
    enriched = _enrich_decisions(project_id, decisions, reviewer)
    _append_to_decisions_log(enriched)

    return {
        "saved": len(decisions),
        "accepted": sum(1 for d in decisions if d.decision == "accepted"),
        "rejected": sum(1 for d in decisions if d.decision == "rejected"),
    }


def load_expert_review(project_id: str) -> Optional[dict]:
    """Загрузить сохранённые решения эксперта для проекта."""
    project_dir = resolve_project_dir(project_id)
    path = project_dir / "_output" / "expert_review.json"
    return _load_json(path)


def _enrich_decisions(project_id: str, decisions: list[ExpertDecision], reviewer: str) -> list[KnowledgeBaseEntry]:
    """Обогатить решения контекстом из findings/optimization JSON."""
    project_dir = resolve_project_dir(project_id)
    output_dir = project_dir / "_output"

    # Загрузить findings
    findings_map = {}
    for fname in ["03a_norms_verified.json", "03_findings.json"]:
        fpath = output_dir / fname
        fdata = _load_json(fpath)
        if fdata:
            for item in fdata.get("findings", fdata.get("items", [])):
                findings_map[item.get("id", "")] = item
            break

    # Загрузить optimization
    opt_map = {}
    opt_data = _load_json(output_dir / "optimization.json")
    if opt_data:
        for item in opt_data.get("items", []):
            opt_map[item.get("id", "")] = item

    # Загрузить project_info для section
    info = _load_json(project_dir / "project_info.json") or {}
    section = info.get("section", "")

    # Следующий ID
    existing_log = _load_decisions_log()
    next_num = len(existing_log) + 1

    entries = []
    for dec in decisions:
        source = findings_map.get(dec.item_id) or opt_map.get(dec.item_id) or {}

        # Извлечь norm_refs
        norm_refs = []
        norm = source.get("norm", source.get("norm_ref", ""))
        if norm:
            norm_refs = [norm] if isinstance(norm, str) else norm

        entry = KnowledgeBaseEntry(
            id=f"DEC-{next_num:04d}",
            source_project=project_id,
            section=section,
            item_id=dec.item_id,
            item_type=dec.item_type,
            severity=source.get("severity", ""),
            category=source.get("category", ""),
            summary=source.get("problem", source.get("description", source.get("summary", ""))),
            norm_refs=norm_refs,
            sheet=str(source.get("sheet", "")),
            page=source.get("page"),
            expert_decision=dec.decision,
            expert_reason=dec.rejection_reason or "",
            expert_reviewer=dec.reviewer or reviewer,
            expert_date=dec.timestamp or _now_iso(),
        )
        entries.append(entry)
        next_num += 1

    return entries


# ═══════════════════════════════════════════════════════════════════════════
# Глобальный лог решений (knowledge_base/decisions_log.json)
# ═══════════════════════════════════════════════════════════════════════════

def _load_decisions_log() -> list[dict]:
    data = _load_json(DECISIONS_LOG_FILE)
    if isinstance(data, dict):
        return data.get("entries", [])
    if isinstance(data, list):
        return data
    return []


def _save_decisions_log(entries: list[dict]):
    _save_json(DECISIONS_LOG_FILE, {"entries": entries})


def _append_to_decisions_log(new_entries: list[KnowledgeBaseEntry]):
    """Добавить записи в глобальный лог (дедупликация по project+item_id)."""
    existing = _load_decisions_log()
    existing_keys = {(e.get("source_project"), e.get("item_id")) for e in existing}

    # Обновить существующие или добавить новые
    updated_map = {(e.get("source_project"), e.get("item_id")): e for e in existing}
    for entry in new_entries:
        key = (entry.source_project, entry.item_id)
        updated_map[key] = entry.model_dump()

    _save_decisions_log(list(updated_map.values()))


def get_knowledge_base(
    status: Optional[str] = None,
    section: Optional[str] = None,
    item_type: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
) -> dict:
    """Получить записи базы знаний с фильтрацией."""
    entries = _load_decisions_log()

    # Вычислить status для каждой записи
    for e in entries:
        if e.get("customer_confirmed"):
            e["status"] = "customer_confirmed"
        else:
            e["status"] = e.get("expert_decision", "")

    # Фильтрация
    if status:
        entries = [e for e in entries if e.get("status") == status]
    if section:
        entries = [e for e in entries if e.get("section", "").upper() == section.upper()]
    if item_type:
        entries = [e for e in entries if e.get("item_type") == item_type]
    if search:
        s = search.lower()
        entries = [e for e in entries if s in json.dumps(e, ensure_ascii=False).lower()]

    total = len(entries)

    # Сортировка по дате (новые первые)
    entries.sort(key=lambda e: e.get("expert_date", ""), reverse=True)

    # Пагинация
    paginated = entries[offset:offset + limit]

    return {
        "total": total,
        "offset": offset,
        "limit": limit,
        "entries": paginated,
    }


def get_kb_stats() -> dict:
    """Счётчики по вкладкам."""
    entries = _load_decisions_log()
    stats = {"rejected": 0, "accepted": 0, "customer_confirmed": 0}
    for e in entries:
        if e.get("customer_confirmed"):
            stats["customer_confirmed"] += 1
        elif e.get("expert_decision") == "rejected":
            stats["rejected"] += 1
        elif e.get("expert_decision") == "accepted":
            stats["accepted"] += 1
    stats["total"] = sum(stats.values())
    return stats


def mark_customer_confirmed(entry_ids: list[str], note: str = "") -> int:
    """Отметить записи как согласованные заказчиком."""
    entries = _load_decisions_log()
    count = 0
    now = _now_iso()
    for e in entries:
        if e.get("id") in entry_ids and e.get("expert_decision") == "accepted":
            e["customer_confirmed"] = True
            e["customer_date"] = now
            if note:
                e["customer_note"] = note
            count += 1
    _save_decisions_log(entries)
    return count


def unmark_customer_confirmed(entry_ids: list[str]) -> int:
    """Снять отметку согласования заказчиком."""
    entries = _load_decisions_log()
    count = 0
    for e in entries:
        if e.get("id") in entry_ids and e.get("customer_confirmed"):
            e["customer_confirmed"] = False
            e["customer_date"] = None
            e["customer_note"] = None
            count += 1
    _save_decisions_log(entries)
    return count


def _find_project_dir(project_id: str) -> Optional[Path]:
    """Найти папку проекта — пробует resolve_project_dir, fallback через iter_project_dirs."""
    try:
        return resolve_project_dir(project_id)
    except Exception:
        pass
    # Fallback: поиск по имени
    try:
        from webapp.services.project_service import iter_project_dirs
        for pid, path in iter_project_dirs():
            if pid == project_id or pid.endswith("/" + project_id):
                return path
    except Exception:
        pass
    return None


def revoke_decision(entry_id: str, project_id: str, item_id: str) -> int:
    """Отменить решение — удалить из глобального лога и из expert_review проекта."""
    # 1. Удалить из decisions_log.json
    entries = _load_decisions_log()
    before = len(entries)
    entries = [e for e in entries if e.get("id") != entry_id]
    _save_decisions_log(entries)
    removed = before - len(entries)

    # 2. Удалить из expert_review.json проекта
    if project_id and item_id:
        project_dir = _find_project_dir(project_id)
        if project_dir:
            review_path = project_dir / "_output" / "expert_review.json"
            if review_path.exists():
                review_data = _load_json(review_path)
                if review_data and "decisions" in review_data:
                    review_data["decisions"] = [
                        d for d in review_data["decisions"]
                        if d.get("item_id") != item_id
                    ]
                    _save_json(review_path, review_data)

    return removed


# ═══════════════════════════════════════════════════════════════════════════
# Детекция паттернов из отклонённых решений
# ═══════════════════════════════════════════════════════════════════════════

def _load_patterns() -> list[dict]:
    data = _load_json(PATTERNS_FILE)
    if isinstance(data, dict):
        return data.get("patterns", [])
    if isinstance(data, list):
        return data
    return []


def _save_patterns(patterns: list[dict]):
    _save_json(PATTERNS_FILE, {"patterns": patterns})


def detect_patterns(min_frequency: int = 3) -> list[dict]:
    """Найти повторяющиеся паттерны среди отклонённых решений.

    Группирует по (section, category, norm_prefix) и ищет кластеры с >= min_frequency.
    """
    entries = _load_decisions_log()
    rejected = [e for e in entries if e.get("expert_decision") == "rejected"]

    if not rejected:
        return []

    # Группировка по (section, category)
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for e in rejected:
        section = e.get("section", "").upper()
        category = e.get("category", "").lower()
        # Извлечь prefix нормы (до пункта)
        norm_prefix = ""
        norms = e.get("norm_refs", [])
        if norms:
            norm_prefix = norms[0].split(",")[0].split("п.")[0].strip()
        key = (section, category, norm_prefix)
        groups[key].append(e)

    # Фильтр по частоте
    existing_patterns = _load_patterns()
    existing_ids = {p.get("pattern_id") for p in existing_patterns}
    next_num = len(existing_patterns) + 1

    new_patterns = []
    for (section, category, norm_prefix), items in groups.items():
        if len(items) < min_frequency:
            continue

        # Собрать уникальные причины отклонения
        reasons = [e.get("expert_reason", "") for e in items if e.get("expert_reason")]
        common_reason = reasons[0] if reasons else "Повторяющееся отклонение"

        # Уникальные проекты
        projects = list({e.get("source_project", "") for e in items})
        example_ids = [e.get("id", "") for e in items[:5]]

        pattern_id = f"PAT-{next_num:03d}"
        # Проверить что не дублирует существующий
        desc = f"[{section}] Категория '{category}'"
        if norm_prefix:
            desc += f", норма {norm_prefix}"
        desc += f" — {len(items)} отклонений"

        # Пропустить если уже есть паттерн с такой же description
        if any(p.get("description") == desc for p in existing_patterns):
            continue

        suggested_fix = f"Не генерировать замечания типа '{category}'"
        if norm_prefix:
            suggested_fix += f" со ссылкой на {norm_prefix}"
        suggested_fix += f". Причина: {common_reason}"

        target_file = f"disciplines/{section}/checklist.md" if section else ""

        new_patterns.append({
            "pattern_id": pattern_id,
            "section": section,
            "description": desc,
            "frequency": len(items),
            "projects_affected": projects,
            "example_ids": example_ids,
            "suggested_fix": suggested_fix,
            "target_file": target_file,
            "status": "pending",
            "proposed_at": _now_iso(),
            "decided_by": None,
            "decided_at": None,
        })
        next_num += 1

    # Сохранить новые + существующие
    if new_patterns:
        all_patterns = existing_patterns + new_patterns
        _save_patterns(all_patterns)

    return _load_patterns()


def get_patterns() -> list[dict]:
    """Получить все паттерны."""
    return _load_patterns()


def update_pattern_status(pattern_id: str, status: str, edited_fix: Optional[str] = None, decided_by: str = "") -> bool:
    """Обновить статус паттерна (approve/dismiss/edit)."""
    patterns = _load_patterns()
    for p in patterns:
        if p.get("pattern_id") == pattern_id:
            p["status"] = status
            p["decided_at"] = _now_iso()
            p["decided_by"] = decided_by
            if edited_fix is not None:
                p["suggested_fix"] = edited_fix
            _save_patterns(patterns)
            return True
    return False


# ═══════════════════════════════════════════════════════════════════════════
# Импорт решений из Excel
# ═══════════════════════════════════════════════════════════════════════════

def _resolve_project_id_from_sheet(ws_title: str) -> str:
    """Найти полный project_id по сокращённому имени листа Excel.

    Имя листа может быть '133_23-ГК-ГРЩ', а нужен 'EOM/133_23-ГК-ГРЩ'.
    """
    from webapp.services.project_service import iter_project_dirs

    # Убрать префикс "ОПТ " если есть
    name = ws_title
    if name.startswith("ОПТ "):
        name = name[4:]

    for pid, path in iter_project_dirs():
        # pid может быть "133_23-ГК-ГРЩ" или "EOM/133_23-ГК-ГРЩ"
        if pid == name or pid.endswith("/" + name) or pid.replace("/", "-") == name:
            return pid
    return name  # fallback — вернуть как есть


def import_decisions_from_excel(file_path: str) -> dict:
    """Импортировать решения из Excel-файла с колонками 'Решение эксперта' и 'Причина отклонения'.

    Возвращает {project_id: {saved, accepted, rejected}} для каждого обнаруженного проекта.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    results = {}

    for ws in wb.worksheets:
        if ws.title.upper() in ("ИНСТРУКЦИЯ", "СВОДКА", "SUMMARY"):
            continue

        # Найти колонки по заголовку (строка 1)
        headers = {}
        header_row = list(next(ws.iter_rows(min_row=1, max_row=1), []))
        for col_idx, c in enumerate(header_row):
            val = str(c.value or "").strip().lower()
            if val == "решение эксперта" or ("решение" in val and "эксперт" in val):
                headers["decision"] = col_idx
            elif "причина" in val and "отклон" in val:
                headers["reason"] = col_idx
            elif val == "id":
                headers["id"] = col_idx
            elif val == "№" and "id" not in headers:
                headers["num"] = col_idx  # fallback: если нет ID, используем №
            elif val == "project_id":
                headers["project_id"] = col_idx
            elif "тип" in val:
                headers["type"] = col_idx

        # Если нет колонки ID, но есть № — не подходит (в № порядковый номер, а не F-001)
        # Для листов оптимизации колонка ID есть всегда

        if "decision" not in headers or "id" not in headers:
            continue

        # Определить project_id: скрытый столбец (строка 2) → имя листа → fallback
        project_id = None
        if "project_id" in headers:
            row2 = list(next(ws.iter_rows(min_row=2, max_row=2), []))
            if row2 and headers["project_id"] < len(row2):
                pid_val = str(row2[headers["project_id"]].value or "").strip()
                if pid_val:
                    project_id = pid_val
        if not project_id:
            project_id = _resolve_project_id_from_sheet(ws.title)

        decisions = []
        for row in ws.iter_rows(min_row=3, values_only=False):
            cells = list(row)
            item_id = str(cells[headers["id"]].value or "").strip()
            decision_raw = str(cells[headers["decision"]].value or "").strip().lower()
            reason = ""
            if "reason" in headers:
                reason = str(cells[headers["reason"]].value or "").strip()

            if not item_id or not decision_raw:
                continue

            # Нормализация
            if decision_raw in ("принято", "accepted", "да", "yes", "+"):
                decision = "accepted"
            elif decision_raw in ("отклонено", "rejected", "нет", "no", "-"):
                decision = "rejected"
            else:
                continue

            # Определить тип
            item_type = "finding"
            if item_id.upper().startswith("OPT"):
                item_type = "optimization"
            if "type" in headers:
                type_val = str(cells[headers["type"]].value or "").strip().lower()
                if "opt" in type_val:
                    item_type = "optimization"

            decisions.append(ExpertDecision(
                item_id=item_id,
                item_type=item_type,
                decision=decision,
                rejection_reason=reason if decision == "rejected" else None,
                timestamp=_now_iso(),
            ))

        if decisions:
            result = save_expert_review(project_id, decisions)
            results[project_id] = result

    wb.close()
    return results
