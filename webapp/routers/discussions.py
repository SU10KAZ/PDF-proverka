"""
REST API для обсуждений замечаний/оптимизаций.
"""
import io
import json

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from webapp.config import DISCUSSION_MODELS, DISCUSSION_DEFAULT_MODEL
from webapp.models.discussion import ChatRequest, ResolutionRequest, ReviseRequest
from webapp.services import discussion_service

router = APIRouter(prefix="/api/discussions", tags=["discussions"])


@router.get("/models")
async def get_discussion_models():
    """Список доступных моделей для обсуждений."""
    return {
        "models": DISCUSSION_MODELS,
        "default": DISCUSSION_DEFAULT_MODEL,
    }


@router.get("/{project_id:path}/resolved/excel")
async def download_resolved_excel(project_id: str, type: str = "finding"):
    """Скачать Excel с отработанными (confirmed/revised) замечаниями."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    items = discussion_service.list_discussion_items(project_id, type)
    resolved = [item for item in items if item.get("discussion_status") in ("confirmed", "revised")]
    if not resolved:
        raise HTTPException(404, "Нет отработанных замечаний")

    from webapp.services.discussion_service import _load_json
    from webapp.config import PROJECTS_DIR
    proj_dir = PROJECTS_DIR / project_id / "_output"
    if type == "finding":
        fdata = _load_json(proj_dir / "03_findings.json")
        findings_data = fdata.get("findings", []) if fdata else []
    else:
        fdata = _load_json(proj_dir / "optimization.json")
        findings_data = fdata.get("optimizations", fdata.get("items", [])) if fdata else []

    wb = Workbook()
    ws = wb.active
    ws.title = "Отработанные замечания" if type == "finding" else "Отработанная оптимизация"

    header_font = Font(name="Times New Roman", bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell_font = Font(name="Times New Roman", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    if type == "finding":
        headers = ["ID", "Критичность", "Лист", "Замечание", "Норма", "Рекомендация", "Статус", "Резолюция"]
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 50
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 40
    else:
        headers = ["ID", "Тип", "Описание", "Предложение", "Экономия", "Статус", "Резолюция"]
        for i, w in enumerate([8, 14, 50, 50, 12, 14, 40]):
            ws.column_dimensions[chr(65 + i)].width = w

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align
        cell.border = thin_border

    row = 2
    for item in resolved:
        item_id = item.get("item_id", "")
        status = item.get("discussion_status", "")
        full = next((f for f in findings_data if (f.get("id") or f.get("item_id", "")) == item_id), {})
        disc = discussion_service.get_discussion(project_id, item_id)
        resolution = disc.resolution_summary if disc else ""

        if type == "finding":
            values = [
                item_id,
                full.get("severity", ""),
                full.get("sheet", ""),
                full.get("description") or full.get("problem") or full.get("finding", ""),
                full.get("norm", ""),
                full.get("solution") or full.get("recommendation", ""),
                status,
                resolution,
            ]
        else:
            values = [
                item_id,
                full.get("type", ""),
                full.get("current") or full.get("description", ""),
                full.get("proposed") or full.get("recommendation", ""),
                f"{full.get('savings_pct', '')}%" if full.get("savings_pct") else "",
                status,
                resolution,
            ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = cell_font
            cell.alignment = wrap_align
            cell.border = thin_border
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"resolved_{project_id.replace('/', '_')}_{type}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{project_id:path}/list")
async def list_items(project_id: str, type: str = "finding"):
    """Список замечаний/оптимизаций с индикатором статуса обсуждения."""
    if type not in ("finding", "optimization"):
        raise HTTPException(400, "type must be 'finding' or 'optimization'")
    items = discussion_service.list_discussion_items(project_id, type)
    return {"items": [item.model_dump() for item in items]}


@router.get("/{project_id:path}/{item_id}/estimate-tokens")
async def estimate_tokens(project_id: str, item_id: str, type: str = "finding"):
    """Оценить количество токенов контекста для обсуждения."""
    if type not in ("finding", "optimization"):
        raise HTTPException(400, "type must be 'finding' or 'optimization'")
    return discussion_service.estimate_context_tokens(project_id, item_id, type)


@router.get("/{project_id:path}/{item_id}")
async def get_discussion(project_id: str, item_id: str):
    """Получить историю обсуждения."""
    disc = discussion_service.get_discussion(project_id, item_id)
    if disc is None:
        return {"item_id": item_id, "messages": [], "status": None, "total_cost_usd": 0.0}
    return disc.model_dump()


@router.post("/{project_id:path}/{item_id}/chat")
async def send_message(project_id: str, item_id: str, req: ChatRequest, type: str = "finding"):
    """Отправить сообщение в чат."""
    if type not in ("finding", "optimization"):
        raise HTTPException(400, "type must be 'finding' or 'optimization'")
    if not req.message.strip():
        raise HTTPException(400, "message cannot be empty")

    result = await discussion_service.send_chat_message(
        project_id=project_id,
        item_id=item_id,
        item_type=type,
        user_message=req.message,
        model=req.model,
    )
    return result


@router.post("/{project_id:path}/{item_id}/chat/stream")
async def send_message_stream(project_id: str, item_id: str, req: ChatRequest, type: str = "finding"):
    """SSE стриминг ответа чата."""
    if type not in ("finding", "optimization"):
        raise HTTPException(400, "type must be 'finding' or 'optimization'")
    if not req.message.strip() and not req.image:
        raise HTTPException(400, "message or image required")

    async def event_generator():
        async for event in discussion_service.send_chat_message_stream(
            project_id=project_id,
            item_id=item_id,
            item_type=type,
            user_message=req.message,
            model=req.model,
            image=req.image,
        ):
            yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.post("/{project_id:path}/{item_id}/resolve")
async def resolve(project_id: str, item_id: str, req: ResolutionRequest, type: str = "finding"):
    """Установить резолюцию (confirmed / rejected / revised)."""
    if req.status not in ("confirmed", "rejected", "revised"):
        raise HTTPException(400, "status must be confirmed, rejected, or revised")
    if type not in ("finding", "optimization"):
        raise HTTPException(400, "type must be 'finding' or 'optimization'")

    try:
        result = discussion_service.set_resolution(
            project_id=project_id,
            item_id=item_id,
            item_type=type,
            status=req.status,
            summary=req.summary,
        )
        return result
    except Exception as e:
        raise HTTPException(500, f"Ошибка сохранения: {e}")


@router.post("/{project_id:path}/{item_id}/revise")
async def revise(project_id: str, item_id: str, req: ReviseRequest, type: str = "finding"):
    """Сгенерировать изменённую версию на основе диалога."""
    if type not in ("finding", "optimization"):
        raise HTTPException(400, "type must be 'finding' or 'optimization'")

    result = await discussion_service.generate_revised_version(
        project_id=project_id,
        item_id=item_id,
        item_type=type,
        model=req.model,
    )
    if "error" in result:
        raise HTTPException(400, result["error"])
    return result


@router.post("/{project_id:path}/{item_id}/apply-revision")
async def apply_revision(project_id: str, item_id: str, revised: dict, type: str = "finding"):
    """Применить изменённую версию замечания."""
    if type not in ("finding", "optimization"):
        raise HTTPException(400, "type must be 'finding' or 'optimization'")

    result = discussion_service.apply_revision(project_id, item_id, type, revised)
    if "error" in result:
        raise HTTPException(400, result["error"])
    return result


@router.post("/{project_id:path}/{item_id}/truncate")
async def truncate_discussion(project_id: str, item_id: str, body: dict):
    """Обрезать историю обсуждения до keep_count сообщений."""
    keep_count = body.get("keep_count", 0)
    result = discussion_service.truncate_messages(project_id, item_id, keep_count)
    return result


@router.delete("/{project_id:path}/{item_id}")
async def delete_discussion(project_id: str, item_id: str):
    """Удалить историю обсуждения."""
    path = discussion_service._discussion_path(project_id, item_id)
    if path.exists():
        path.unlink()
        return {"ok": True}
    raise HTTPException(404, "Discussion not found")


